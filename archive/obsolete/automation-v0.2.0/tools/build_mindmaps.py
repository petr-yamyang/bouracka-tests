#!/usr/bin/env python3
"""
build_mindmaps.py — render TestTargets + TestCases mindmaps from
BOURACKA-TESTPLAN-v0.2.xlsx into recon/diagrams/{tt,tc}-mindmap.{png,svg,pdf}.

Run from anywhere — paths are computed relative to this script's
location. Requires:
  - Python 3.9+
  - openpyxl   (pip install openpyxl)
  - Graphviz on PATH (Windows: winget install Graphviz; macOS: brew
    install graphviz; Ubuntu: apt-get install graphviz)
"""
from __future__ import annotations

import shutil
import subprocess
import sys
from pathlib import Path

# ── Path resolution (script-relative; works from any cwd) ────────────────────
HERE = Path(__file__).resolve().parent          # …/bouracka-tests/tools
ROOT = HERE.parent                               # …/bouracka-tests
XLSX = ROOT / "BOURACKA-TESTPLAN-v0.2.xlsx"
OUT  = ROOT / "recon" / "diagrams"

# ── Friendly pre-flight ──────────────────────────────────────────────────────
def fail(msg: str, hint: str = "") -> None:
    print(f"[FAIL] {msg}", file=sys.stderr)
    if hint:
        print(f"       hint: {hint}", file=sys.stderr)
    sys.exit(1)

if not XLSX.exists():
    fail(f"Excel not found: {XLSX}",
         "Run from inside the bouracka-tests/ checkout, or pass an absolute path.")

try:
    from openpyxl import load_workbook
except ImportError:
    fail("openpyxl is not installed.",
         "pip install openpyxl  (or:  py -m pip install openpyxl)")

dot_exe = shutil.which("dot")
if not dot_exe:
    fail("Graphviz 'dot' is not on PATH.",
         "Windows:  winget install Graphviz  (then restart shell)\n"
         "       macOS:    brew install graphviz\n"
         "       Ubuntu:   sudo apt-get install graphviz")

OUT.mkdir(parents=True, exist_ok=True)

# ── Palettes (matches Excel conditional formatting) ──────────────────────────
PRI_COLOR = {"A": "#F4B084", "B": "#FFD966", "C": "#C6E0B4", "D": "#D9D9D9"}
RELEASE_BORDER = {"R1": "#1F4E79", "R2": "#7F7F7F", "R3": "#7F7F7F", "R4+": "#7F7F7F"}
TYPE_SHAPE = {"happy": "ellipse", "failure": "octagon",
              "regression": "box", "smoke": "house"}

# ── Load Excel ───────────────────────────────────────────────────────────────
wb = load_workbook(XLSX, data_only=True)
ws_tt, ws_tc = wb["01_TestTargets"], wb["02_TestCases"]

def header_idx(ws, name):
    for i, c in enumerate(ws[1], start=1):
        if c.value == name:
            return i
    raise KeyError(name)

TT_COLS = {h: header_idx(ws_tt, h) for h in
           ["item_code", "item_name_cs", "release", "decomposition_kind",
            "behaviour_ref", "component_ref", "severity", "urgency",
            "priority", "state_machine_terminal_state", "env_coverage"]}
TC_COLS = {h: header_idx(ws_tc, h) for h in
           ["item_code", "item_name_cs", "test_target_ref", "type",
            "severity", "urgency", "priority",
            "state_machine_terminal_state", "state_error_subreason"]}

def rows(ws, hdrs):
    return [
        {h: ws.cell(row=r, column=hdrs[h]).value for h in hdrs}
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=hdrs["item_code"]).value
    ]

tts, tcs = rows(ws_tt, TT_COLS), rows(ws_tc, TC_COLS)

def compute_pri(sev, urg):
    if not sev or not urg or sev == "D" or urg == "D":
        return "D"
    n = (ord(sev) - 64) + (ord(urg) - 64)
    return "A" if n <= 3 else "B" if n == 4 else "C"

for r in tts: r["pri"] = compute_pri(r["severity"], r["urgency"])
for r in tcs: r["pri"] = compute_pri(r["severity"], r["urgency"])

def short(s, maxlen=42):
    s = str(s or "")
    return s if len(s) <= maxlen else s[:maxlen - 1].rstrip() + "…"

def esc(s):
    return str(s or "").replace('"', '\\"').replace("\n", " ")


# ── Diagram 1 — TestTargets ──────────────────────────────────────────────────
def emit_tt_dot():
    r1 = [t for t in tts if t["release"] == "R1"]
    r2 = [t for t in tts if t["release"] == "R2"]
    area_map = {"TT-CP-R1-A1": "auth", "TT-CP-R1-A2": "auth",
                "TT-CP-R1-B": "driver", "TT-CP-R1-C": "vehicle",
                "TT-CP-R1-D": "sign",   "TT-CP-R1-E": "e2e",
                "TT-CP-R1-F": "failures"}
    area_titles = {
        "auth":     "Vstup + Auth\\n(SMS PING + OTP)",
        "driver":   "Driver data\\n(zenID + AISPOV ROB/CRR)",
        "vehicle":  "Vehicle data\\n(zenID + AISPOV CRV/Pojištění)",
        "sign":     "Sign + Complete\\n(SMS-OTP + e-mail)",
        "e2e":      "E2E orchestration",
        "failures": "Failure envelope",
    }

    lines = [
        'digraph TT {',
        '  graph [layout=twopi, root="center", overlap=false, splines=true,',
        '         ranksep="2.0 1.6 1.4 1.2", bgcolor="white",',
        '         fontname="Arial", fontsize=10, label=<',
        '<B>Bouračka — TestTargets scope (R1 + R2)</B><BR/>',
        '<FONT POINT-SIZE="9">',
        f'{len(r1)} R1 (active) + {len(r2)} R2 (deferred) ',
        '· colour = priority (A=red B=yellow C=green D=gray) ',
        '· border = release (blue=R1 gray=R2)',
        '</FONT>>, labelloc="t"];',
        '  node [fontname="Arial", fontsize=10, style="filled,rounded",',
        '        shape=box, penwidth=2];',
        '  edge [color="#7F7F7F", arrowhead=none];',
        '',
        '  center [label=<<B>Bouračka</B><BR/>'
        '<FONT POINT-SIZE="9">tst.bouracka.cz<BR/>tst.demo.bouracka.cz</FONT>>,',
        '          shape=circle, fillcolor="#1F4E79", fontcolor="white",',
        '          width=1.6, fixedsize=true];',
        '',
    ]
    for area in ["auth", "driver", "vehicle", "sign", "e2e", "failures"]:
        lines += [
            f'  hub_r1_{area} [label="{area_titles[area]}",',
            f'                shape=box, fillcolor="#BDD7EE",',
            f'                fontcolor="#1F4E79", penwidth=2,',
            f'                color="#1F4E79", style="filled,rounded,bold"];',
            f'  center -> hub_r1_{area};',
        ]
    lines += [
        '',
        '  hub_r2 [label="R2 — deferred\\n(11 TestTargets)",',
        '          shape=box, fillcolor="#D9D9D9",',
        '          fontcolor="#404040", style="filled,rounded,dashed",',
        '          color="#7F7F7F", penwidth=2];',
        '  center -> hub_r2;',
        '',
    ]
    for tt in r1:
        code = tt["item_code"]
        area = area_map.get(code, "e2e")
        title = short(tt["item_name_cs"], 38)
        ts = tt.get("state_machine_terminal_state") or ""
        pri = tt["pri"]
        nid = code.replace("-", "_")
        label = (f'<<B>{code}</B><BR/>'
                 f'<FONT POINT-SIZE="9">{esc(title)}</FONT><BR/>'
                 f'<FONT POINT-SIZE="8" COLOR="#404040">→ {esc(short(ts, 26))}</FONT><BR/>'
                 f'<FONT POINT-SIZE="8" COLOR="#7F7F7F">Pri {pri}</FONT>>')
        lines += [
            f'  {nid} [label={label}, fillcolor="{PRI_COLOR[pri]}",',
            f'           color="{RELEASE_BORDER["R1"]}"];',
            f'  hub_r1_{area} -> {nid};',
        ]
    lines.append('')
    for tt in r2:
        code = tt["item_code"]
        title = short(tt["item_name_cs"], 38)
        pri = tt["pri"]
        nid = code.replace("-", "_")
        label = (f'<<B>{code}</B><BR/>'
                 f'<FONT POINT-SIZE="9">{esc(title)}</FONT><BR/>'
                 f'<FONT POINT-SIZE="8" COLOR="#7F7F7F">Pri {pri}</FONT>>')
        lines += [
            f'  {nid} [label={label}, fillcolor="{PRI_COLOR[pri]}",',
            f'           color="{RELEASE_BORDER["R2"]}", style="filled,rounded,dashed"];',
            f'  hub_r2 -> {nid};',
        ]
    lines.append('}')
    return "\n".join(lines)


# ── Diagram 2 — TestCases ────────────────────────────────────────────────────
def primary_tt(ttref):
    s = str(ttref or "")
    for token in s.replace(",", "+").split("+"):
        t = token.strip()
        if t.startswith("TT-CP-R1-"):
            return t
    return None

def emit_tc_dot():
    tt_used, seen = [], set()
    for tc in tcs:
        ptt = primary_tt(tc["test_target_ref"])
        if ptt and ptt not in seen:
            seen.add(ptt); tt_used.append(ptt)
    tt_name = {t["item_code"]: t["item_name_cs"] for t in tts}

    lines = [
        'digraph TC {',
        '  graph [layout=twopi, root="center", overlap=false, splines=true,',
        '         ranksep="1.6 1.4", bgcolor="white",',
        '         fontname="Arial", fontsize=10, label=<',
        '<B>Bouračka — TestCases — automation priority</B><BR/>',
        '<FONT POINT-SIZE="9">',
        f'{len(tcs)} R1 TestCases · ',
        'colour = priority (A=red — automate first; B=yellow — next; C=green — later) ',
        '· shape = type (○ happy · ⬡ failure · □ regression)',
        '</FONT>>, labelloc="t"];',
        '  node [fontname="Arial", fontsize=9, style="filled,rounded", penwidth=1.5];',
        '  edge [color="#7F7F7F", arrowhead=none];',
        '',
        '  center [label=<<B>Bouračka R1</B><BR/>'
        '<FONT POINT-SIZE="9">automation backlog</FONT>>,',
        '          shape=circle, fillcolor="#1F4E79", fontcolor="white",',
        '          width=1.6, fixedsize=true];',
        '',
    ]
    for ttcode in tt_used:
        hub = ttcode.replace("-", "_") + "_hub"
        title = short(tt_name.get(ttcode, ttcode), 32)
        lines += [
            f'  {hub} [label=<<B>{ttcode}</B><BR/>'
            f'<FONT POINT-SIZE="8">{esc(title)}</FONT>>,',
            f'              shape=box, fillcolor="#BDD7EE", color="#1F4E79",',
            f'              style="filled,rounded,bold", penwidth=2];',
            f'  center -> {hub};',
        ]
    lines.append('')
    for tc in tcs:
        ptt = primary_tt(tc["test_target_ref"])
        if not ptt:
            continue
        code = tc["item_code"]
        title = short(tc["item_name_cs"], 36)
        ts = tc.get("state_machine_terminal_state") or ""
        sub = tc.get("state_error_subreason") or ""
        pri = tc["pri"]
        ttype = tc.get("type") or "regression"
        shape = TYPE_SHAPE.get(ttype, "box")
        nid = code.replace("-", "_")
        hub = ptt.replace("-", "_") + "_hub"
        ts_part = f'<BR/><FONT POINT-SIZE="7" COLOR="#404040">→ {esc(short(ts, 22))}</FONT>'
        sub_part = (f'<BR/><FONT POINT-SIZE="7" COLOR="#C00000">{esc(short(sub, 22))}</FONT>'
                    if sub else '')
        label = (f'<<B>{code}</B><BR/>'
                 f'<FONT POINT-SIZE="8">{esc(title)}</FONT>'
                 f'{ts_part}{sub_part}<BR/>'
                 f'<FONT POINT-SIZE="7" COLOR="#7F7F7F">{ttype} · Pri {pri}</FONT>>')
        lines += [
            f'  {nid} [label={label}, shape={shape},',
            f'           fillcolor="{PRI_COLOR[pri]}", color="#1F4E79"];',
            f'  {hub} -> {nid};',
        ]
    lines.append('}')
    return "\n".join(lines)


def render(name, dot_text):
    dot_path = OUT / f"{name}.dot"
    dot_path.write_text(dot_text, encoding="utf-8")
    for fmt in ("png", "svg", "pdf"):
        out_path = OUT / f"{name}.{fmt}"
        result = subprocess.run(
            ["dot", f"-T{fmt}", "-Ktwopi", str(dot_path), "-o", str(out_path)],
            capture_output=True, text=True,
        )
        if result.returncode != 0:
            print(f"[FAIL] {name}.{fmt}: {result.stderr.strip()}")
        else:
            print(f"[OK] {name}.{fmt}  {out_path.stat().st_size:>9,d} bytes")


print(f"== build_mindmaps ==")
print(f"   xlsx in : {XLSX}")
print(f"   out dir : {OUT}")
print()
print("=== TT mindmap ===")
render("tt-mindmap", emit_tt_dot())
print()
print("=== TC mindmap ===")
render("tc-mindmap", emit_tc_dot())
print()
print("done.")

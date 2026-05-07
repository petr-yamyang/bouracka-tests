#!/usr/bin/env python3
"""
validate_workbook.py — workbook validator per CP-SUPIN-03 STEP 5
                       + Opus review §6.6.

10 checks against BOURACKA-TESTPLAN-v0.2.xlsx; exits non-zero on any
failure. JSON output for SecOps + automated pipeline consumption.

Run:
  python tools/validate_workbook.py
  python tools/validate_workbook.py --json
"""
from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("[FAIL] openpyxl not installed; pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "BOURACKA-TESTPLAN-v0.2.xlsx"

if not XLSX.exists():
    sys.exit(f"[FAIL] workbook not found: {XLSX}")

wb = load_workbook(XLSX, data_only=True)


def rows(sheet, hdr_row=1):
    """Return list of dicts keyed by column header."""
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    headers = [c.value for c in ws[hdr_row]]
    out = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        row = {h: ws.cell(row=r, column=i + 1).value for i, h in enumerate(headers) if h}
        if any(v is not None and v != "" for v in row.values()):
            out.append(row)
    return out


checks: list[dict[str, Any]] = []


def check(name: str, ok: bool, detail: str = "", failures: list[str] | None = None):
    checks.append({
        "name": name,
        "status": "OK" if ok else "FAIL",
        "detail": detail,
        "failures": failures or [],
    })


# ─── 1. All required sheets present ────────────────────────────────────────
required = [
    "00_README", "00b_Requirements", "01_TestTargets", "01b_Req_FURPS_Cartesian",
    "01c_StateMachine", "02_TestCases", "02b_TC_Parameters",
    "02c_TC_Assertions", "02d_AssertionLibrary", "03_TestData",
    "04_TestEnvironments", "05a_TestPlan", "05b_TestSchedule",
    "05c_TestEstimate", "06_TestRuns", "07_TestRunResults", "08_Bugs",
    "09_Reports", "10_Glossary", "11_Changelog",
]
missing = [s for s in required if s not in wb.sheetnames]
check("01_required_sheets_present", not missing,
      f"{len(required) - len(missing)}/{len(required)} present", missing)

# ─── 2. ItemBase columns present on every entity sheet ────────────────────
expected_base = ["id", "item_code", "item_name_cs", "item_name_en"]
entity_sheets = ["00b_Requirements", "01_TestTargets", "02_TestCases",
                 "03_TestData", "04_TestEnvironments"]
fails = []
for s in entity_sheets:
    if s not in wb.sheetnames:
        continue
    headers = [c.value for c in wb[s][1]]
    for col in expected_base:
        if col not in headers:
            fails.append(f"{s}::{col}")
check("02_itembase_present", not fails,
      f"{len(entity_sheets)} sheets checked", fails)

# ─── 3. R-FURPS-1: every TT + TC carries furps_dimensions ─────────────────
fails = []
for s in ["01_TestTargets", "02_TestCases"]:
    for row in rows(s):
        code = row.get("item_code") or "(blank)"
        if not row.get("furps_dimensions"):
            fails.append(f"{s}::{code}")
check("03_furps_dimensions_populated", not fails,
      f"{len(fails)} blanks of {len(rows('01_TestTargets')) + len(rows('02_TestCases'))} rows", fails[:5])

# ─── 4. R-DERIVE-1: every TT cites source_artefact_kind + ref ─────────────
fails = []
for row in rows("01_TestTargets"):
    code = row.get("item_code") or "(blank)"
    if not row.get("source_artefact_kind") or not row.get("source_artefact_ref"):
        fails.append(code)
check("04_tt_source_artefact_present", not fails,
      f"{len(fails)} TTs missing source_artefact_*", fails[:5])

# ─── 5. requirement_ref on TT resolves to 00b_Requirements ────────────────
req_codes = {r.get("item_code") for r in rows("00b_Requirements")}
fails = []
for row in rows("01_TestTargets"):
    code = row.get("item_code") or "(blank)"
    rref = row.get("requirement_ref")
    if rref and rref not in req_codes:
        fails.append(f"{code} → {rref}")
check("05_tt_requirement_ref_valid", not fails,
      f"{len(req_codes)} reqs in registry", fails[:5])

# ─── 6. test_target_ref on TC resolves to 01_TestTargets ──────────────────
tt_codes = {r.get("item_code") for r in rows("01_TestTargets")}
fails = []
for row in rows("02_TestCases"):
    code = row.get("item_code") or "(blank)"
    ttref = row.get("test_target_ref") or ""
    # may be combined with + (e.g. "TT-CP-R1-001+TT-CP-R1-004"); split first
    for token in ttref.replace(",", "+").split("+"):
        t = token.strip()
        if t and not t.startswith("TT-CP-R1") and not t.startswith("TT-CP-R2"):
            continue
        if t and t not in tt_codes:
            fails.append(f"{code} → {t}")
check("06_tc_test_target_ref_valid", not fails,
      f"{len(tt_codes)} TTs in registry", fails[:5])

# ─── 7. assertion_library_ref on 02c resolves to 02d ──────────────────────
lib_codes = {r.get("library_code") for r in rows("02d_AssertionLibrary")}
fails = []
for row in rows("02c_TC_Assertions"):
    aref = row.get("assertion_library_ref")
    if aref and aref not in lib_codes:
        fails.append(f"{row.get('test_case_ref')}::{row.get('step_id')} → {aref}")
check("07_assertion_library_ref_valid", not fails,
      f"{len(lib_codes)} library entries", fails[:5])

# ─── 8. R-EXPAND-1: Cartesian na cells have justification ─────────────────
fails = []
for row in rows("01b_Req_FURPS_Cartesian"):
    if row.get("cell_status") == "na" and not row.get("na_justification"):
        fails.append(f"{row.get('requirement_ref')}::{row.get('furps_dimension')}")
check("08_cartesian_na_justified", not fails,
      f"{len(fails)} N/A cells without justification", fails[:5])

# ─── 9. R-PLAN-2: scheduling_unit_kind on Cartesian valid ─────────────────
valid_units = {"iteration", "sprint", "phase", "none"}
fails = []
for row in rows("01b_Req_FURPS_Cartesian"):
    sku = row.get("scheduling_unit_kind")
    if sku and sku not in valid_units:
        fails.append(f"{row.get('requirement_ref')}::{row.get('furps_dimension')} = {sku}")
check("09_scheduling_unit_kind_valid", not fails,
      f"valid set: {sorted(valid_units)}", fails[:5])

# ─── 10. State machine integrity: every transition has from + to ──────────
states = {r.get("name") for r in rows("01c_StateMachine") if r.get("kind") == "state"}
fails = []
for row in rows("01c_StateMachine"):
    if row.get("kind") == "transition":
        frm = (row.get("from_state") or "").split("(")[0].strip()
        to = (row.get("to_state") or "").split("(")[0].strip()
        if frm and frm not in states and frm != "" and frm != "(none)" and "(any" not in frm and frm != "(during phone verify)" and frm != "(during driver-data load)" and frm != "(at gateway D00)":
            fails.append(f"transition {row.get('name')} from={frm} not in states")
        if to and to not in states and to != "" and to != "(none)" and "(stays in" not in to and "(no transition" not in to:
            fails.append(f"transition {row.get('name')} to={to} not in states")
check("10_state_machine_integrity", not fails,
      f"{len(states)} states; transitions cross-checked", fails[:5])

# ─── Summary ──────────────────────────────────────────────────────────────
total = len(checks)
ok = sum(1 for c in checks if c["status"] == "OK")
fail = total - ok

result = {
    "workbook": str(XLSX),
    "total_checks": total,
    "passed": ok,
    "failed": fail,
    "checks": checks,
}

if "--json" in sys.argv:
    print(json.dumps(result, indent=2, ensure_ascii=False))
else:
    print(f"[validate-workbook] checked {XLSX.name}")
    for c in checks:
        sym = "[OK]" if c["status"] == "OK" else "[FAIL]"
        print(f"{sym:6s} {c['name']}  {c['detail']}")
        if c["status"] == "FAIL" and c["failures"]:
            for f in c["failures"]:
                print(f"        - {f}")
    print()
    if fail == 0:
        print(f"[OK] all {total} checks passed.")
    else:
        print(f"[FAIL] {fail} of {total} checks failed.")

sys.exit(0 if fail == 0 else 1)

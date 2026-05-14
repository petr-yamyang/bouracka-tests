"""
Generate tools/tests/fixtures/synthetic-v0.4.3-mini.xlsx

A small deterministic workbook that mirrors the real BOURACKA-TESTPLAN-v0.4.3.xlsx
schema but contains only 3 TCs and 2 bugs.  Run this script to regenerate:

    python tools/tests/fixtures/make_synthetic_fixture.py

The fixture intentionally exercises:
  - TC with multi-line steps_summary (TC-SYN-001, 3 lines)
  - TC with single-line steps_summary (TC-SYN-002)
  - TC with empty steps_summary (TC-SYN-003)
  - Bug row with screenshot_ref + trace_ref (BUG-SYN-001)
  - Bug row with neither (BUG-SYN-002)
  - 02c_TC_Assertions row with non-empty step_id matching existing step (TC-SYN-001-S-01)
    -> this is an orphan against the patcher's STP-TC-SYN-001-01 format, used to test
       orphan reporting.
"""

from __future__ import annotations

import datetime
import pathlib
import sys

import openpyxl
from openpyxl import Workbook

OUTPUT = pathlib.Path(__file__).parent / "synthetic-v0.4.3-mini.xlsx"

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _add_sheet(wb: Workbook, title: str, headers: list[str]) -> openpyxl.worksheet.worksheet.Worksheet:
    ws = wb.create_sheet(title)
    ws.append(headers)
    return ws


def _stub_sheet(wb: Workbook, title: str) -> None:
    ws = wb.create_sheet(title)
    ws.append(["(stub)"])


# ---------------------------------------------------------------------------
# build
# ---------------------------------------------------------------------------

def build(output: pathlib.Path = OUTPUT) -> None:
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)

    # ---- 00_README ----
    _stub_sheet(wb, "00_README")

    # ---- 00b_Requirements ----
    _stub_sheet(wb, "00b_Requirements")

    # ---- 00c_VersionSanityRules ----
    _stub_sheet(wb, "00c_VersionSanityRules")

    # ---- 01_TestTargets ----
    _stub_sheet(wb, "01_TestTargets")

    # ---- 00e_BranchView ----
    _stub_sheet(wb, "00e_BranchView")

    # ---- 01b_Req_FURPS_Cartesian ----
    _stub_sheet(wb, "01b_Req_FURPS_Cartesian")

    # ---- 01c_StateMachine ----
    _stub_sheet(wb, "01c_StateMachine")

    # ---- 02_TestCases ----
    tc_headers = [
        "id", "item_code", "item_name_cs", "item_name_en", "item_descr_cs",
        "item_descr_en", "comments_KP_en", "env", "ext_ws", "item_type",
        "item_status", "severity", "urgency", "priority", "submitter",
        "submit_date", "test_target_ref", "type", "preconditions",
        "steps_summary", "expected_summary", "postconditions", "env_coverage",
        "viewport_spec", "framework_targets", "state_machine_terminal_state",
        "state_error_subreason", "dev_spec_path", "last_run_date",
        "last_run_verdict", "created_at", "updated_at", "notes",
        "furps_dimensions", "impulse_ref", "diligence", "env_constraints",
        "applies_to_demo", "applies_to_prod",
    ]
    ws_tc = _add_sheet(wb, "02_TestCases", tc_headers)

    now = datetime.datetime(2026, 1, 1, 12, 0, 0)

    def _tc_row(tc_id: int, code: str, steps: str | None) -> list:
        row = [None] * len(tc_headers)
        row[0] = tc_id                   # id
        row[1] = code                    # item_code
        row[2] = f"Syntetický TC {tc_id}"  # item_name_cs
        row[3] = f"Synthetic TC {tc_id}"   # item_name_en
        row[10] = "draft"                # item_status
        row[13] = "A"                    # priority
        row[19] = steps                  # steps_summary (index 19 = col 20)
        row[30] = now                    # created_at
        row[31] = now                    # updated_at
        return row

    ws_tc.append(_tc_row(1, "TC-SYN-001", "Step alpha\nStep beta\nStep gamma"))
    ws_tc.append(_tc_row(2, "TC-SYN-002", "Single step only"))
    ws_tc.append(_tc_row(3, "TC-SYN-003", None))  # empty steps_summary

    # ---- 02b_TC_Parameters ----
    _stub_sheet(wb, "02b_TC_Parameters")

    # ---- 02c_TC_Assertions ----
    ws_2c = _add_sheet(wb, "02c_TC_Assertions", [
        "id", "test_case_ref", "step_id", "assertion_library_ref",
        "furps_dimension", "expected", "notes",
    ])
    # row with step_id matching legacy format (orphan vs STP- convention)
    ws_2c.append([1, "TC-SYN-001", "TC-SYN-001-S-01", "ASSERT-001", "F", "form visible", ""])
    # row with empty step_id (TC-level assertion, should be left alone)
    ws_2c.append([2, "TC-SYN-002", None, "ASSERT-002", "R", "redirect happened", ""])

    # ---- 02d_AssertionLibrary ----
    _stub_sheet(wb, "02d_AssertionLibrary")

    # ---- 03_TestData ----
    _stub_sheet(wb, "03_TestData")

    # ---- 05a_TestPlan ----
    _stub_sheet(wb, "05a_TestPlan")

    # ---- 05b_TestSchedule ----
    _stub_sheet(wb, "05b_TestSchedule")

    # ---- 05c_TestEstimate ----
    _stub_sheet(wb, "05c_TestEstimate")

    # ---- 04_TestEnvironments ----
    _stub_sheet(wb, "04_TestEnvironments")

    # ---- 05_TestSets_DEPRECATED ----
    _stub_sheet(wb, "05_TestSets_DEPRECATED")

    # ---- 06_TestRuns ----
    _stub_sheet(wb, "06_TestRuns")

    # ---- 07_TestRunResults ----
    _stub_sheet(wb, "07_TestRunResults")

    # ---- 13_TestExecutionSummary ----
    _stub_sheet(wb, "13_TestExecutionSummary")

    # ---- 14_AssertionGateResults ----
    _stub_sheet(wb, "14_AssertionGateResults")

    # ---- 08_Bugs ----
    # Mirror the 39-column layout of the real workbook
    bug_headers = [
        "id", "item_code", "item_name_cs", "item_name_en", "item_descr_cs",
        "item_descr_en", "item_type", "item_status", "severity", "urgency",
        "priority", "submitter", "submit_date", "env_where_present",
        "linked_tc_ref", "linked_run_result_ref", "repro_steps", "expected",
        "actual", "workaround", "test_implication", "created_at", "updated_at",
        "notes", "diligence", "impulse_ref", "applies_to_demo", "applies_to_prod",
        "tc_ref", "assertion_code", "assertion_text", "first_seen", "last_seen",
        "occurrences", "envs_seen", "runs_seen", "screenshot_ref", "trace_ref",
        "error_message",
    ]
    ws_bugs = _add_sheet(wb, "08_Bugs", bug_headers)

    def _bug_row(bug_id: int, code: str, tc: str, ss: str | None, tr: str | None) -> list:
        row = [None] * len(bug_headers)
        row[0] = bug_id
        row[1] = code
        row[2] = f"Chyba {bug_id}"
        row[3] = f"Bug {bug_id}"
        row[14] = tc           # linked_tc_ref
        row[21] = now          # created_at
        row[22] = now          # updated_at
        row[36] = ss           # screenshot_ref (col 37, index 36)
        row[37] = tr           # trace_ref (col 38, index 37)
        return row

    ws_bugs.append(_bug_row(1, "BUG-SYN-001", "TC-SYN-001",
                             "runs/r-2026-01-01/artefacts/TC-SYN-001/screenshot.png",
                             "runs/r-2026-01-01/trace.zip"))
    ws_bugs.append(_bug_row(2, "BUG-SYN-002", "TC-SYN-002", None, None))

    # ---- 09_Reports ----
    _stub_sheet(wb, "09_Reports")

    # ---- 10_Glossary ----
    _stub_sheet(wb, "10_Glossary")

    # ---- 11_Changelog ----
    # Mirror the real workbook: title row, blank row, then header row, then data
    ws_cl = wb.create_sheet("11_Changelog")
    ws_cl.append(["Changelog", None, None, None])
    ws_cl.append([None, None, None, None])
    ws_cl.append(["Version", "Date", "Author", "Change summary"])
    ws_cl.append(["v0.4.3", datetime.datetime(2026, 1, 1),
                  "synthetic-fixture-generator",
                  "Synthetic baseline for patcher tests."])

    # ---- 01d_PrioritySevUrgMatrix ----
    _stub_sheet(wb, "01d_PrioritySevUrgMatrix")

    wb.save(output)
    print(f"Wrote {output} ({output.stat().st_size} bytes)")


if __name__ == "__main__":
    out = pathlib.Path(sys.argv[1]) if len(sys.argv) > 1 else OUTPUT
    build(out)

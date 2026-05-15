"""
Generate tools/tests/fixtures/synthetic-v0.4.3-with-user-data.xlsx

A v0.4.3-era workbook representing Kate's working copy: same schema as
synthetic-v0.4.3-mini.xlsx but with user-entered bugs, runs, and results
filled in. Used by Brief #001b (data migration) tests.

Run to regenerate:
    python tools/tests/fixtures/make_synthetic_user_data_fixture.py
"""
from __future__ import annotations

import datetime
import pathlib
import shutil
import sys

import openpyxl
from openpyxl import Workbook

MINI_FIXTURE = pathlib.Path(__file__).parent / "synthetic-v0.4.3-mini.xlsx"
OUTPUT = pathlib.Path(__file__).parent / "synthetic-v0.4.3-with-user-data.xlsx"

NOW = datetime.datetime(2026, 5, 14, 9, 0, 0)


def _add_data_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list]) -> None:
    ws = wb[title] if title in wb.sheetnames else wb.create_sheet(title)
    # If sheet is a stub (single col), wipe and rebuild
    if ws.max_column == 1 and ws.max_row <= 2:
        for row_num in range(ws.max_row, 0, -1):
            ws.delete_rows(row_num)
    ws.append(headers)
    for row in rows:
        ws.append(row)


def build(output: pathlib.Path = OUTPUT) -> None:
    # Start from mini fixture (has correct TC + assertion + changelog structure)
    shutil.copy2(MINI_FIXTURE, output)
    wb = openpyxl.load_workbook(str(output))

    # ---- 08_Bugs: Kate's 5 bugs ----
    bug_headers = [
        "id", "item_code", "item_name_cs", "item_name_en", "item_descr_cs",
        "item_descr_en", "item_type", "item_status", "severity", "urgency",
        "priority", "submitter", "submit_date", "env_where_present",
        "linked_tc_ref", "linked_run_result_ref", "repro_steps", "expected",
        "actual", "workaround", "test_implication", "created_at", "updated_at",
        "notes", "diligence", "impulse_ref", "applies_to_demo", "applies_to_prod",
        "tc_ref", "assertion_code", "assertion_text", "first_seen", "last_seen",
        "occurrences", "envs_seen", "runs_seen", "screenshot_ref", "trace_ref",
        "error_message",
    ]

    def _bug(i, code, tc, ss=None, tr=None, status="open", sev="B"):
        row = [None] * len(bug_headers)
        row[0] = i
        row[1] = code
        row[2] = f"Chyba Kate {i}"
        row[3] = f"Kate Bug {i}"
        row[7] = status
        row[8] = sev
        row[13] = "ENV-DMO"
        row[14] = tc
        row[21] = NOW
        row[22] = NOW
        row[36] = ss
        row[37] = tr
        return row

    bug_rows = [
        _bug(1, "BUG-KATE-001", "TC-SYN-001",
             ss="runs/r-2026-05-10/artefacts/TC-SYN-001/screenshot.png",
             tr="runs/r-2026-05-10/trace.zip"),
        _bug(2, "BUG-KATE-002", "TC-SYN-002", sev="A"),
        _bug(3, "BUG-KATE-003", "TC-SYN-001",
             ss="runs/r-2026-05-11/artefacts/TC-SYN-001/fail.png"),
        _bug(4, "BUG-KATE-004", "TC-SYN-003", status="closed", sev="C"),
        _bug(5, "BUG-KATE-005", "TC-SYN-002"),
    ]
    ws_bugs = wb["08_Bugs"]
    # Clear existing data rows (keep header)
    for r in range(ws_bugs.max_row, 1, -1):
        ws_bugs.delete_rows(r)
    for row in bug_rows:
        ws_bugs.append(row)

    # ---- 06_TestRuns: 3 runs ----
    run_headers = [
        "id", "run_id", "run_timestamp", "triggered_by", "framework_targets",
        "env", "status", "exit_code", "total_tcs", "passed", "failed", "skipped",
        "drift_skip_count", "parity_pass_count", "parity_divergence_count",
        "duration_ms", "envelope_path", "notes",
    ]
    run_rows = [
        [1, "run-2026-05-10T08-00-00Z-aaaaaaa", NOW, "kate", "playwright,selenium",
         "demo", "done", 0, 3, 3, 0, 0, 0, 3, 0, 45000, "runs/r-2026-05-10/envelope.json", ""],
        [2, "run-2026-05-10T14-00-00Z-bbbbbbb", NOW, "kate", "cypress",
         "demo", "done", 1, 3, 2, 1, 0, 0, 0, 1, 30000, "runs/r-2026-05-10b/envelope.json", ""],
        [3, "run-2026-05-11T09-00-00Z-ccccccc", NOW, "kate", "playwright",
         "tst", "done", 0, 2, 2, 0, 0, 0, 2, 0, 22000, "runs/r-2026-05-11/envelope.json", ""],
    ]
    ws_runs = wb["06_TestRuns"]
    # Wipe stub and write real data
    for r in range(ws_runs.max_row, 0, -1):
        ws_runs.delete_rows(r)
    ws_runs.append(run_headers)
    for row in run_rows:
        ws_runs.append(row)

    # ---- 07_TestRunResults: 5 results ----
    result_headers = [
        "id", "result_id", "run_id", "tc_ref", "framework", "verdict",
        "parity_status", "duration_ms", "exit_code", "error_message",
        "screenshot_ref", "trace_ref", "video_ref", "assertion_codes_passed",
        "assertion_codes_failed", "soft_pass_reason", "created_at", "notes",
    ]
    result_rows = [
        [1, "RES-001", "run-2026-05-10T08-00-00Z-aaaaaaa", "TC-SYN-001",
         "playwright", "pass", "agree", 1500, 0, None, None, None, None, "ASSERT-001", None, None, NOW, ""],
        [2, "RES-002", "run-2026-05-10T08-00-00Z-aaaaaaa", "TC-SYN-002",
         "playwright", "pass", "agree", 1200, 0, None, None, None, None, "ASSERT-002", None, None, NOW, ""],
        [3, "RES-003", "run-2026-05-10T08-00-00Z-aaaaaaa", "TC-SYN-003",
         "playwright", "pass", "agree", 900, 0, None, None, None, None, None, None, None, NOW, ""],
        [4, "RES-004", "run-2026-05-10T14-00-00Z-bbbbbbb", "TC-SYN-001",
         "cypress", "fail", "divergence", 2100, 1,
         "ALT-4 cypress mock failure",
         "runs/r-2026-05-10b/TC-SYN-001/fail.png", None, None, None, "ASSERT-001", None, NOW, ""],
        [5, "RES-005", "run-2026-05-11T09-00-00Z-ccccccc", "TC-SYN-001",
         "playwright", "pass", "agree", 1300, 0, None, None, None, None, "ASSERT-001", None, None, NOW, ""],
    ]
    ws_results = wb["07_TestRunResults"]
    for r in range(ws_results.max_row, 0, -1):
        ws_results.delete_rows(r)
    ws_results.append(result_headers)
    for row in result_rows:
        ws_results.append(row)

    # ---- 09_Reports: 1 report ----
    report_headers = ["id", "report_code", "run_id", "report_type", "report_path",
                      "generated_at", "notes"]
    report_rows = [
        [1, "RPT-001", "run-2026-05-10T08-00-00Z-aaaaaaa", "summary",
         "runs/r-2026-05-10/report.md", NOW, ""],
    ]
    ws_rpts = wb["09_Reports"]
    for r in range(ws_rpts.max_row, 0, -1):
        ws_rpts.delete_rows(r)
    ws_rpts.append(report_headers)
    for row in report_rows:
        ws_rpts.append(row)

    # ---- 13_TestExecutionSummary: 1 row ----
    tes_headers = ["id", "summary_code", "run_id", "total_tcs", "passed", "failed",
                   "skipped", "pass_rate", "parity_pass_count", "parity_divergence_count", "notes"]
    tes_rows = [
        [1, "TES-001", "run-2026-05-10T08-00-00Z-aaaaaaa", 3, 3, 0, 0, 1.0, 3, 0, ""],
    ]
    ws_tes = wb["13_TestExecutionSummary"]
    for r in range(ws_tes.max_row, 0, -1):
        ws_tes.delete_rows(r)
    ws_tes.append(tes_headers)
    for row in tes_rows:
        ws_tes.append(row)

    # ---- 14_AssertionGateResults: 2 rows ----
    agr_headers = ["id", "gate_code", "run_id", "assertion_code", "result",
                   "tc_ref", "step_ref", "notes"]
    agr_rows = [
        [1, "AGR-001", "run-2026-05-10T08-00-00Z-aaaaaaa", "ASSERT-001", "pass",
         "TC-SYN-001", None, ""],
        [2, "AGR-002", "run-2026-05-10T08-00-00Z-aaaaaaa", "ASSERT-002", "pass",
         "TC-SYN-002", None, ""],
    ]
    ws_agr = wb["14_AssertionGateResults"]
    for r in range(ws_agr.max_row, 0, -1):
        ws_agr.delete_rows(r)
    ws_agr.append(agr_headers)
    for row in agr_rows:
        ws_agr.append(row)

    wb.save(str(output))
    print(f"Wrote {output} ({output.stat().st_size} bytes)")


if __name__ == "__main__":
    out = pathlib.Path(sys.argv[1]) if len(sys.argv) > 1 else OUTPUT
    build(out)

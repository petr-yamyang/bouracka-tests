"""
pytest suite for workbook-v0.4.3-to-v0.4.4.py

Run:
    pytest tools/tests/ -v -m "not integration"   # fast (synthetic fixture)
    pytest tools/tests/ -v -m integration         # slow (real workbook, manual)
"""

from __future__ import annotations

import importlib.util
import pathlib
import re
import shutil

import openpyxl
import pytest

FIXTURES_DIR = pathlib.Path(__file__).parent / "fixtures"
REPO_ROOT = pathlib.Path(__file__).resolve().parent.parent.parent
REAL_SOURCE = REPO_ROOT / "BOURACKA-TESTPLAN-v0.4.3.xlsx"

STEP_COLS = [
    "id", "step_code", "tc_ref", "ordinal", "action_cs", "action_en",
    "expected_cs", "expected_en", "framework_hint", "assertion_lib_ref",
    "is_decision_point", "comments_KP_en", "created_at", "updated_at", "notes",
]

BUG_NEW_COLS = [
    "linked_step_ref",
    "evidence_screenshot_path",
    "evidence_video_path",
    "evidence_trace_path",
    "evidence_capture_kind",
    "evidence_capture_at",
]


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _load_patcher():
    spec_path = pathlib.Path(__file__).resolve().parent.parent / "workbook-v0.4.3-to-v0.4.4.py"
    spec = importlib.util.spec_from_file_location("workbook_patcher", spec_path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _col_map(ws, header_row: int = 1) -> dict[str, int]:
    row = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    return {v: i for i, v in enumerate(row, 1) if v is not None}


def _sheet_rows(ws, min_row: int = 2) -> list[tuple]:
    return [r for r in ws.iter_rows(min_row=min_row, values_only=True) if r[0] is not None]


# ---------------------------------------------------------------------------
# fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def patcher():
    return _load_patcher()


@pytest.fixture()
def patched(tmp_path, patcher):
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    shutil.copy2(FIXTURES_DIR / "synthetic-v0.4.3-mini.xlsx", src)
    rc = patcher.patch(src, dst, dry_run=False, verbose=False)
    assert rc in (0, 1)
    wb = openpyxl.load_workbook(dst)
    yield wb, src, dst
    wb.close()


# ---------------------------------------------------------------------------
# T-1  02e_TestSteps created with expected columns
# ---------------------------------------------------------------------------

def test_02e_TestSteps_created_with_expected_columns(patched):
    wb, _src, _dst = patched
    assert "02e_TestSteps" in wb.sheetnames, "02e_TestSteps sheet must exist"
    ws = wb["02e_TestSteps"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers == STEP_COLS, f"Column order mismatch: {headers}"


# ---------------------------------------------------------------------------
# T-2  Every TC has at least one step row
# ---------------------------------------------------------------------------

def test_each_TC_gets_at_least_one_step_row(patched):
    wb, _src, _dst = patched
    ws_tc = wb["02_TestCases"]
    ws_e = wb["02e_TestSteps"]

    tc_codes = {r[1] for r in ws_tc.iter_rows(min_row=2, values_only=True) if r[0] is not None}
    cmap_e = _col_map(ws_e)
    tc_ref_col = cmap_e["tc_ref"] - 1

    step_tcs = {r[tc_ref_col] for r in ws_e.iter_rows(min_row=2, values_only=True) if r[0] is not None}

    missing = tc_codes - step_tcs
    assert not missing, f"TCs with no step rows: {missing}"


# ---------------------------------------------------------------------------
# T-3  step_code matches STP-<tc>-NN format
# ---------------------------------------------------------------------------

def test_step_code_format_STP_dash_TC_dash_NN(patched):
    wb, _src, _dst = patched
    ws_e = wb["02e_TestSteps"]
    pattern = re.compile(r"^STP-.+-\d{2}$")
    bad = []
    for row in ws_e.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        step_code = row[1]  # col 2
        if not pattern.match(str(step_code or "")):
            bad.append(step_code)
    assert not bad, f"step_code values not matching STP-<tc>-NN: {bad}"


# ---------------------------------------------------------------------------
# T-4  Ordinals are gapless per TC
# ---------------------------------------------------------------------------

def test_ordinals_are_gapless_per_tc(patched):
    wb, _src, _dst = patched
    ws_e = wb["02e_TestSteps"]
    cmap = _col_map(ws_e)
    tc_ref_col = cmap["tc_ref"] - 1
    ord_col = cmap["ordinal"] - 1

    by_tc: dict[str, list[int]] = {}
    for row in ws_e.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        by_tc.setdefault(row[tc_ref_col], []).append(row[ord_col])

    for tc, ords in by_tc.items():
        expected = list(range(1, len(ords) + 1))
        assert sorted(ords) == expected, f"{tc}: ordinals {sorted(ords)} are not gapless 1..{len(ords)}"


# ---------------------------------------------------------------------------
# T-5  steps_count in 02_TestCases matches actual row count in 02e
# ---------------------------------------------------------------------------

def test_steps_count_column_matches_02e_row_count(patched):
    wb, _src, _dst = patched
    ws_tc = wb["02_TestCases"]
    ws_e = wb["02e_TestSteps"]

    cmap_e = _col_map(ws_e)
    tc_ref_col = cmap_e["tc_ref"] - 1

    steps_by_tc: dict[str, int] = {}
    for row in ws_e.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            tc = row[tc_ref_col]
            steps_by_tc[tc] = steps_by_tc.get(tc, 0) + 1

    cmap_tc = _col_map(ws_tc)
    sc_col = cmap_tc.get("steps_count")
    assert sc_col is not None, "steps_count column missing from 02_TestCases"

    item_code_col = cmap_tc["item_code"] - 1
    sc_col_idx = sc_col - 1
    mismatches = []
    for row in ws_tc.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        tc_code = row[item_code_col]
        expected = steps_by_tc.get(tc_code, 0)
        actual = row[sc_col_idx]
        if actual != expected:
            mismatches.append(f"{tc_code}: steps_count={actual} expected={expected}")
    assert not mismatches, "steps_count mismatches: " + "; ".join(mismatches)


# ---------------------------------------------------------------------------
# T-6  08_Bugs has all new evidence columns
# ---------------------------------------------------------------------------

def test_08_bugs_has_new_columns(patched):
    wb, _src, _dst = patched
    ws = wb["08_Bugs"]
    cmap = _col_map(ws)
    missing = [c for c in BUG_NEW_COLS if c not in cmap]
    assert not missing, f"08_Bugs missing columns: {missing}"


# ---------------------------------------------------------------------------
# T-7  Legacy screenshot_ref migrated to evidence_screenshot_path
# ---------------------------------------------------------------------------

def test_legacy_screenshot_ref_migrated(patched, tmp_path):
    # Use the synthetic fixture which has BUG-SYN-001 with screenshot_ref
    wb, src, dst = patched
    ws = wb["08_Bugs"]
    cmap = _col_map(ws)

    ss_col = cmap.get("screenshot_ref")
    esp_col = cmap.get("evidence_screenshot_path")
    eck_col = cmap.get("evidence_capture_kind")

    assert ss_col and esp_col and eck_col, "Expected columns not found"

    migrations_ok = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        ss_val = row[ss_col - 1]
        esp_val = row[esp_col - 1]
        eck_val = row[eck_col - 1]
        if ss_val:
            assert esp_val == ss_val, f"screenshot_ref not migrated: ss={ss_val!r} esp={esp_val!r}"
            assert eck_val == "manual-tester", f"evidence_capture_kind wrong: {eck_val!r}"
            migrations_ok.append(ss_val)

    assert migrations_ok, "No bug rows with screenshot_ref found in synthetic fixture"


# ---------------------------------------------------------------------------
# T-8  Legacy screenshot_ref and trace_ref preserved (soft-deprecation)
# ---------------------------------------------------------------------------

def test_legacy_screenshot_ref_preserved(patched):
    wb_src = openpyxl.load_workbook(FIXTURES_DIR / "synthetic-v0.4.3-mini.xlsx")
    wb_dst, _src, _dst = patched

    ws_src = wb_src["08_Bugs"]
    ws_dst = wb_dst["08_Bugs"]

    cmap_src = _col_map(ws_src)
    cmap_dst = _col_map(ws_dst)

    assert "screenshot_ref" in cmap_dst, "screenshot_ref column must be preserved in v0.4.4"
    assert "trace_ref" in cmap_dst, "trace_ref column must be preserved in v0.4.4"

    ss_src = cmap_src["screenshot_ref"] - 1
    tr_src = cmap_src["trace_ref"] - 1
    ss_dst = cmap_dst["screenshot_ref"] - 1
    tr_dst = cmap_dst["trace_ref"] - 1

    src_rows = list(ws_src.iter_rows(min_row=2, values_only=True))
    dst_rows = list(ws_dst.iter_rows(min_row=2, values_only=True))

    # compare row-by-row up to the source row count
    for i, (sr, dr) in enumerate(zip(src_rows, dst_rows)):
        if sr[0] is None:
            break
        assert sr[ss_src] == dr[ss_dst], f"Row {i+2}: screenshot_ref changed: {sr[ss_src]!r} → {dr[ss_dst]!r}"
        assert sr[tr_src] == dr[tr_dst], f"Row {i+2}: trace_ref changed: {sr[tr_src]!r} → {dr[tr_dst]!r}"

    wb_src.close()


# ---------------------------------------------------------------------------
# T-9  11_Changelog has a new row with version='v0.4.4'
# ---------------------------------------------------------------------------

def test_changelog_row_appended(patched):
    wb, _src, _dst = patched
    ws = wb["11_Changelog"]

    # scan all rows for version v0.4.4
    found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == "v0.4.4":
            found = True
            # 4th cell should be the change type embedded in summary or empty
            break
    assert found, "No row with Version='v0.4.4' found in 11_Changelog"


# ---------------------------------------------------------------------------
# T-10  Idempotency: patching twice produces identical 02e + evidence columns
# ---------------------------------------------------------------------------

def test_idempotent_when_run_twice(tmp_path, patcher):
    src = tmp_path / "src.xlsx"
    dst1 = tmp_path / "dst1.xlsx"
    dst2 = tmp_path / "dst2.xlsx"
    shutil.copy2(FIXTURES_DIR / "synthetic-v0.4.3-mini.xlsx", src)

    rc1 = patcher.patch(src, dst1, dry_run=False, verbose=False)
    assert rc1 in (0, 1)

    rc2 = patcher.patch(dst1, dst2, dry_run=False, verbose=False)
    assert rc2 in (0, 1)

    wb1 = openpyxl.load_workbook(dst1)
    wb2 = openpyxl.load_workbook(dst2)

    # 02e_TestSteps rows must be identical
    rows1 = list(wb1["02e_TestSteps"].iter_rows(min_row=2, values_only=True))
    rows2 = list(wb2["02e_TestSteps"].iter_rows(min_row=2, values_only=True))
    # strip timestamps (created_at/updated_at at indices 12,13) for comparison
    def _strip_ts(row):
        r = list(row)
        if len(r) > 13:
            r[12] = None
            r[13] = None
        return tuple(r)
    assert [_strip_ts(r) for r in rows1 if r[0] is not None] == \
           [_strip_ts(r) for r in rows2 if r[0] is not None], \
           "02e_TestSteps rows differ between first and second run"

    # steps_count values must match
    cmap1 = _col_map(wb1["02_TestCases"])
    cmap2 = _col_map(wb2["02_TestCases"])
    sc1_col = cmap1.get("steps_count")
    sc2_col = cmap2.get("steps_count")
    assert sc1_col and sc2_col
    tc_rows1 = [r for r in wb1["02_TestCases"].iter_rows(min_row=2, values_only=True) if r[0] is not None]
    tc_rows2 = [r for r in wb2["02_TestCases"].iter_rows(min_row=2, values_only=True) if r[0] is not None]
    sc_vals1 = [r[sc1_col - 1] for r in tc_rows1]
    sc_vals2 = [r[sc2_col - 1] for r in tc_rows2]
    assert sc_vals1 == sc_vals2, "steps_count values differ between runs"

    # evidence columns in 08_Bugs must match
    for col in BUG_NEW_COLS:
        if col == "evidence_capture_at":
            continue  # timestamp, skip
        cmap_b1 = _col_map(wb1["08_Bugs"])
        cmap_b2 = _col_map(wb2["08_Bugs"])
        c1 = cmap_b1.get(col)
        c2 = cmap_b2.get(col)
        if c1 and c2:
            vals1 = [r[c1 - 1] for r in wb1["08_Bugs"].iter_rows(min_row=2, values_only=True) if r[0] is not None]
            vals2 = [r[c2 - 1] for r in wb2["08_Bugs"].iter_rows(min_row=2, values_only=True) if r[0] is not None]
            assert vals1 == vals2, f"{col} differs between runs"

    wb1.close()
    wb2.close()


# ---------------------------------------------------------------------------
# T-11  --dry-run does not write dest
# ---------------------------------------------------------------------------

def test_dry_run_does_not_write_dest(tmp_path, patcher):
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    shutil.copy2(FIXTURES_DIR / "synthetic-v0.4.3-mini.xlsx", src)

    rc = patcher.patch(src, dst, dry_run=True, verbose=False)
    assert rc in (0, 1)
    assert not dst.exists(), f"--dry-run must not write dest file; {dst} exists"


# ---------------------------------------------------------------------------
# T-12  Orphan step_id reported in PATCH-REPORT, row unchanged in dest
# ---------------------------------------------------------------------------

def test_orphan_02c_step_id_reported_not_modified(patched, tmp_path):
    wb_dst, src, dst = patched

    # synthetic fixture has TC-SYN-001-S-01 as an orphan step_id (legacy format)
    ws_dst_2c = wb_dst["02c_TC_Assertions"]
    cmap = _col_map(ws_dst_2c)
    step_id_col = cmap["step_id"] - 1

    # confirm orphan row still has original value in dest
    wb_src = openpyxl.load_workbook(FIXTURES_DIR / "synthetic-v0.4.3-mini.xlsx")
    ws_src_2c = wb_src["02c_TC_Assertions"]
    cmap_src = _col_map(ws_src_2c)
    si_src = cmap_src["step_id"] - 1

    for r_src, r_dst in zip(
        ws_src_2c.iter_rows(min_row=2, values_only=True),
        ws_dst_2c.iter_rows(min_row=2, values_only=True),
    ):
        if r_src[0] is None:
            break
        assert r_src[si_src] == r_dst[step_id_col], \
            f"02c_TC_Assertions step_id was modified: {r_src[si_src]!r} → {r_dst[step_id_col]!r}"

    wb_src.close()

    # confirm PATCH-REPORT mentions the orphan
    report_dir = pathlib.Path(__file__).resolve().parent.parent / "patcher-reports"
    reports = sorted(report_dir.glob("PATCH-REPORT-v0.4.3-to-v0.4.4-*.md"))
    assert reports, "No PATCH-REPORT found"
    latest = reports[-1].read_text(encoding="utf-8")
    assert "TC-SYN-001-S-01" in latest or "orphan" in latest.lower(), \
        "Orphan step_id not mentioned in PATCH-REPORT"


# ---------------------------------------------------------------------------
# T-13  Empty steps_summary → placeholder step row
# ---------------------------------------------------------------------------

def test_empty_steps_summary_produces_placeholder(patched):
    wb, _src, _dst = patched
    ws_e = wb["02e_TestSteps"]
    cmap = _col_map(ws_e)
    tc_ref_col = cmap["tc_ref"] - 1
    action_col = cmap["action_cs"] - 1
    comment_col = cmap["comments_KP_en"] - 1

    placeholder_found = False
    for row in ws_e.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        if row[tc_ref_col] == "TC-SYN-003":  # TC with empty steps_summary in fixture
            assert "(no steps_summary" in str(row[action_col] or ""), \
                f"Expected placeholder action_cs for TC-SYN-003, got: {row[action_col]!r}"
            assert "KP: define steps" in str(row[comment_col] or ""), \
                f"Expected KP comment for TC-SYN-003, got: {row[comment_col]!r}"
            placeholder_found = True
    assert placeholder_found, "No placeholder row found for TC-SYN-003 (empty steps_summary TC)"


# ---------------------------------------------------------------------------
# Integration test — real workbook
# ---------------------------------------------------------------------------

@pytest.mark.integration
def test_integration_real_workbook(tmp_path):
    """
    Patches the real BOURACKA-TESTPLAN-v0.4.3.xlsx and asserts structural
    correctness.  Run manually:  pytest tools/tests/ -v -m integration
    """
    if not REAL_SOURCE.exists():
        pytest.skip(f"Real source workbook not found: {REAL_SOURCE}")

    patcher = _load_patcher()
    dest = tmp_path / "BOURACKA-TESTPLAN-v0.4.4.xlsx"
    rc = patcher.patch(REAL_SOURCE, dest, dry_run=False, verbose=True)
    assert rc in (0, 1), f"Patcher returned {rc}"
    assert dest.exists()

    wb = openpyxl.load_workbook(dest)
    assert "02e_TestSteps" in wb.sheetnames

    ws_e = wb["02e_TestSteps"]
    cmap_e = _col_map(ws_e)
    tc_ref_col = cmap_e["tc_ref"] - 1

    step_rows = [r for r in ws_e.iter_rows(min_row=2, values_only=True) if r[0] is not None]
    assert len(step_rows) >= 24, f"Expected ≥24 step rows, got {len(step_rows)}"

    # P0 smoke TCs from BOURACKA-P0-SMOKE-TC-RANKING ranks 1-4 must have step rows
    p0_tcs = {"TC-CP-008", "TC-CP-015", "TC-CP-019", "TC-CP-013"}
    tcs_with_steps = {r[tc_ref_col] for r in step_rows}
    missing = p0_tcs - tcs_with_steps
    assert not missing, f"P0 TCs missing step rows: {missing}"

    wb.close()

    # idempotency: run again on the output
    dest2 = tmp_path / "BOURACKA-TESTPLAN-v0.4.4-run2.xlsx"
    rc2 = patcher.patch(dest, dest2, dry_run=False, verbose=False)
    assert rc2 in (0, 1)

    wb1 = openpyxl.load_workbook(dest)
    wb2 = openpyxl.load_workbook(dest2)

    rows_e1 = [r for r in wb1["02e_TestSteps"].iter_rows(min_row=2, values_only=True) if r[0] is not None]
    rows_e2 = [r for r in wb2["02e_TestSteps"].iter_rows(min_row=2, values_only=True) if r[0] is not None]

    def _strip(row):
        r = list(row)
        for i in [12, 13]:
            if i < len(r):
                r[i] = None
        return tuple(r)

    assert [_strip(r) for r in rows_e1] == [_strip(r) for r in rows_e2], \
        "Idempotency failure: 02e_TestSteps differs between run 1 and run 2"

    wb1.close()
    wb2.close()

"""
workbook-v0.4.3-to-v0.4.4.py — one-shot idempotent schema patcher.

Upgrades BOURACKA-TESTPLAN-v0.4.3.xlsx to v0.4.4:
  F-1  Create 02e_TestSteps from 02_TestCases.steps_summary
  F-2  Add steps_count column to 02_TestCases
  F-3  Validate 02c_TC_Assertions.step_id FK; report orphans (do not modify)
  F-4  Add evidence_* columns to 08_Bugs; migrate screenshot_ref / trace_ref
  F-5  Append changelog row to 11_Changelog
  F-6  Write PATCH-REPORT-*.md

Usage:
    python tools/workbook-v0.4.3-to-v0.4.4.py [--source PATH] [--dest PATH]
                                               [--dry-run] [--report-only] [-v]

Exit codes:
    0 — success (no orphans, no >20-step warnings)
    1 — success with warnings
    2 — input validation failure
    3 — output write failure
"""

from __future__ import annotations

import argparse
import datetime
import hashlib
import os
import pathlib
import re
import shutil
import sys
import tempfile
import textwrap
from typing import Any

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# constants
# ---------------------------------------------------------------------------

REPO_ROOT = pathlib.Path(__file__).resolve().parent.parent
DEFAULT_SOURCE = REPO_ROOT / "BOURACKA-TESTPLAN-v0.4.3.xlsx"
DEFAULT_DEST = REPO_ROOT / "BOURACKA-TESTPLAN-v0.4.4.xlsx"
REPORT_DIR = pathlib.Path(__file__).resolve().parent / "patcher-reports"

STEPS_COL_HEADER = "steps_summary"
TC_CODE_COL_HEADER = "item_code"
STEPS_COUNT_COL_HEADER = "steps_count"

STEP_COLS = [
    "id", "step_code", "tc_ref", "ordinal", "action_cs", "action_en",
    "expected_cs", "expected_en", "framework_hint", "assertion_lib_ref",
    "is_decision_point", "comments_KP_en", "created_at", "updated_at", "notes",
]

BUG_NEW_COLS = [
    "linked_step_ref",
    "evidence_screenshot_path",
    "evidence_video_path",
    "evidence_trace_path",
    "evidence_capture_kind",
    "evidence_capture_at",
]

CHANGELOG_HEADER_ROW = 3   # row where Version/Date/Author/Change summary live
CHANGELOG_COLS = ["Version", "Date", "Author", "Change summary"]

# ---------------------------------------------------------------------------
# §F-1m..§F-7m data migration constants
# ---------------------------------------------------------------------------

# Sheets to migrate from source-data; value = primary-key column for dup-check
# (None means no dup-check performed for that sheet)
DATA_MIGRATION_SHEETS: dict[str, str | None] = {
    "08_Bugs":                 "item_code",
    "06_TestRuns":             "run_id",
    "07_TestRunResults":       None,
    "09_Reports":              None,
    "13_TestExecutionSummary": None,
    "14_AssertionGateResults": None,
}

# Sheets explicitly NOT migrated from source-data (schema owned by repo)
SCHEMA_ONLY_SHEETS = {
    "02_TestCases", "02e_TestSteps", "02b_TC_Parameters", "02c_TC_Assertions",
    "02d_AssertionLibrary", "00_README", "10_Glossary", "11_Changelog",
}

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _col_index(ws: openpyxl.worksheet.worksheet.Worksheet, header: str, header_row: int = 1) -> int | None:
    """Return 1-based column index of *header* in *header_row*, or None."""
    row = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    for i, val in enumerate(row, 1):
        if val == header:
            return i
    return None


def _col_map(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int = 1) -> dict[str, int]:
    """Return {header: 1-based col index} for all non-None headers in header_row."""
    row = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    return {v: i for i, v in enumerate(row, 1) if v is not None}


def _sha256(path: pathlib.Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _split_steps(text: str | None) -> list[str]:
    if not text:
        return []
    lines = re.split(r"\r\n|\r|\n", str(text))
    return [l.strip() for l in lines if l.strip()]


def _log(verbose: bool, msg: str) -> None:
    if verbose:
        print(msg, file=sys.stderr)


# ---------------------------------------------------------------------------
# §F-1m..§F-7m data migration helpers
# ---------------------------------------------------------------------------

def _rows_as_dicts(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict]:
    """Return non-empty data rows as dicts keyed by column header."""
    cmap = _col_map(ws)
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        result.append({col: row[idx - 1] for col, idx in cmap.items() if idx - 1 < len(row)})
    return result


def _validate_source_data_duplicates(wb_data: openpyxl.Workbook) -> list[str]:
    """Risk gates 2 + 6: return error strings for any duplicate PKs in data sheets."""
    errors: list[str] = []
    for sheet_name, pk_col in DATA_MIGRATION_SHEETS.items():
        if pk_col is None or sheet_name not in wb_data.sheetnames:
            continue
        ws = wb_data[sheet_name]
        cmap = _col_map(ws)
        if pk_col not in cmap:
            continue
        pk_idx = cmap[pk_col] - 1
        seen: dict[str, int] = {}
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or row[0] is None:
                continue
            pk_val = row[pk_idx] if pk_idx < len(row) else None
            if pk_val is None:
                continue
            key = str(pk_val)
            if key in seen:
                errors.append(
                    f"{sheet_name}: duplicate {pk_col}={pk_val!r} at rows {seen[key]} and {row_num}"
                )
            else:
                seen[key] = row_num
    return errors


def _is_stub_sheet(dest_ws: openpyxl.worksheet.worksheet.Worksheet) -> bool:
    """True when dest sheet is a placeholder stub (generated by _stub_sheet())."""
    if dest_ws.max_column != 1:
        return False
    first_cell = dest_ws.cell(row=1, column=1).value
    return first_cell is None or str(first_cell).startswith("(")


def _clear_data_rows(dest_ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    """Delete all non-header rows; return count of rows removed."""
    removed = 0
    for row_num in range(dest_ws.max_row, 1, -1):
        row_vals = [c.value for c in dest_ws[row_num]]
        if any(v is not None for v in row_vals):
            dest_ws.delete_rows(row_num)
            removed += 1
    return removed


def _migrate_one_sheet(
    dest_ws: openpyxl.worksheet.worksheet.Worksheet,
    src_rows: list[dict],
    verbose: bool,
) -> dict:
    """Replace dest data rows with src_rows (header-mapped). Returns stats dict."""
    legacy_evidence: list[dict] = []

    if not src_rows:
        collisions = _clear_data_rows(dest_ws)
        _log(verbose, f"  [{dest_ws.title}] 0 source rows; cleared {collisions} existing rows")
        return {"migrated": 0, "replaced": collisions, "legacy_evidence": []}

    # Stub sheet: adopt source-data schema entirely
    if _is_stub_sheet(dest_ws):
        src_headers = list(src_rows[0].keys())
        _clear_data_rows(dest_ws)
        dest_ws.delete_rows(1)  # remove stub header row
        dest_ws.append(src_headers)
        for row_dict in src_rows:
            dest_ws.append([row_dict.get(h) for h in src_headers])
        _log(verbose, f"  [{dest_ws.title}] stub replaced; {len(src_rows)} rows written")
        return {"migrated": len(src_rows), "replaced": 0, "legacy_evidence": []}

    # Normal path: dest has a real column layout (e.g. 08_Bugs after F-4)
    dest_cmap = _col_map(dest_ws)
    dest_headers = [h for h, _ in sorted(dest_cmap.items(), key=lambda x: x[1])]
    collisions = _clear_data_rows(dest_ws)

    esp_col = dest_cmap.get("evidence_screenshot_path")
    etp_col = dest_cmap.get("evidence_trace_path")
    eck_col = dest_cmap.get("evidence_capture_kind")

    for row_dict in src_rows:
        new_row = [row_dict.get(h) for h in dest_headers]

        # F-4 legacy evidence migration (§F-2m)
        if esp_col and etp_col and eck_col:
            ss_val = row_dict.get("screenshot_ref")
            tr_val = row_dict.get("trace_ref")
            if ss_val and not row_dict.get("evidence_screenshot_path"):
                new_row[esp_col - 1] = ss_val
            if tr_val and not row_dict.get("evidence_trace_path"):
                new_row[etp_col - 1] = tr_val
            if (ss_val or tr_val) and not row_dict.get("evidence_capture_kind"):
                new_row[eck_col - 1] = "manual-tester"
            if ss_val or tr_val:
                legacy_evidence.append({
                    "code": row_dict.get("item_code"),
                    "screenshot_ref": ss_val,
                    "trace_ref": tr_val,
                })

        dest_ws.append(new_row)

    _log(verbose, f"  [{dest_ws.title}] {len(src_rows)} rows written ({collisions} schema rows replaced)")
    return {"migrated": len(src_rows), "replaced": collisions, "legacy_evidence": legacy_evidence}


def _migrate_data(
    wb_dest: openpyxl.Workbook,
    wb_data: openpyxl.Workbook,
    verbose: bool,
) -> dict:
    """Migrate all data sheets from wb_data into wb_dest. Returns full stats."""
    stats: dict[str, dict] = {}
    for sheet_name in DATA_MIGRATION_SHEETS:
        if sheet_name not in wb_data.sheetnames:
            _log(verbose, f"  [{sheet_name}] not in source-data workbook — skipped")
            stats[sheet_name] = {"migrated": 0, "replaced": 0, "legacy_evidence": []}
            continue
        src_rows = _rows_as_dicts(wb_data[sheet_name])
        if sheet_name not in wb_dest.sheetnames:
            _log(verbose, f"  [{sheet_name}] not in dest workbook — skipped")
            stats[sheet_name] = {"migrated": 0, "replaced": 0, "legacy_evidence": []}
            continue
        stats[sheet_name] = _migrate_one_sheet(wb_dest[sheet_name], src_rows, verbose)
    return stats


# ---------------------------------------------------------------------------
# core analysis
# ---------------------------------------------------------------------------

def _analyse(wb: openpyxl.Workbook, run_ts: datetime.datetime, verbose: bool) -> dict:
    """
    Walk the workbook and collect everything the patcher needs to know.
    Returns a dict of analysis results (no mutation of wb).
    """
    ws_tc = wb["02_TestCases"]
    cmap_tc = _col_map(ws_tc)
    item_code_col = cmap_tc.get(TC_CODE_COL_HEADER, 2) - 1      # 0-based
    steps_sum_col = cmap_tc.get(STEPS_COL_HEADER, 20) - 1       # 0-based

    # --- build step rows from 02_TestCases ---
    step_rows: list[dict] = []
    step_id_counter = 1
    tcs_kp_review: list[str] = []
    tcs_over_20: list[str] = []
    tcs_empty_steps: list[str] = []

    for row in ws_tc.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        tc_code = row[item_code_col]
        steps_text = row[steps_sum_col] if steps_sum_col < len(row) else None
        lines = _split_steps(steps_text)

        if not lines:
            tcs_empty_steps.append(tc_code)
            lines = ["(no steps_summary in source workbook)"]

        if len(lines) > 20:
            tcs_over_20.append(tc_code)

        multi_line_original = len(lines) > 1

        for ordinal, line in enumerate(lines, 1):
            nn = f"{ordinal:02d}"
            step_code = f"STP-{tc_code}-{nn}"
            needs_review = multi_line_original
            placeholder = line == "(no steps_summary in source workbook)"

            if placeholder:
                comment = "(KP: define steps for this TC)"
            elif needs_review:
                comment = "(machine-split — KP review needed)"
            else:
                comment = ""

            step_rows.append({
                "id": step_id_counter,
                "step_code": step_code,
                "tc_ref": tc_code,
                "ordinal": ordinal,
                "action_cs": line,
                "action_en": "",
                "expected_cs": "",
                "expected_en": "",
                "framework_hint": "",
                "assertion_lib_ref": "",
                "is_decision_point": False,
                "comments_KP_en": comment,
                "created_at": run_ts,
                "updated_at": run_ts,
                "notes": "(generated by workbook-v0.4.3-to-v0.4.4.py)",
            })
            step_id_counter += 1
            _log(verbose, f"  {step_code}")

        if needs_review or (not lines[0].startswith("(no steps")):
            if tc_code not in tcs_kp_review and any(
                s["tc_ref"] == tc_code and s["comments_KP_en"] for s in step_rows
            ):
                tcs_kp_review.append(tc_code)

    step_code_set = {s["step_code"] for s in step_rows}
    steps_by_tc: dict[str, list[dict]] = {}
    for s in step_rows:
        steps_by_tc.setdefault(s["tc_ref"], []).append(s)

    # --- orphan check in 02c_TC_Assertions ---
    ws_2c = wb["02c_TC_Assertions"]
    cmap_2c = _col_map(ws_2c)
    step_id_col_idx = cmap_2c.get("step_id", 3) - 1  # 0-based

    orphans: list[dict] = []
    for row_num, row in enumerate(ws_2c.iter_rows(min_row=2, values_only=True), 2):
        if row[0] is None:
            continue
        step_id_val = row[step_id_col_idx] if step_id_col_idx < len(row) else None
        if step_id_val and step_id_val not in step_code_set:
            orphans.append({"row": row_num, "step_id": step_id_val,
                            "tc_ref": row[cmap_2c.get("test_case_ref", 2) - 1]})

    # --- bug migration analysis ---
    ws_bugs = wb["08_Bugs"]
    cmap_bugs = _col_map(ws_bugs)
    ss_col = cmap_bugs.get("screenshot_ref")
    tr_col = cmap_bugs.get("trace_ref")
    item_code_col_bugs = cmap_bugs.get("item_code", 2) - 1  # 0-based

    bug_migrations: list[dict] = []
    for row_num, row in enumerate(ws_bugs.iter_rows(min_row=2, values_only=True), 2):
        if row[0] is None:
            continue
        ss_val = row[ss_col - 1] if ss_col and ss_col - 1 < len(row) else None
        tr_val = row[tr_col - 1] if tr_col and tr_col - 1 < len(row) else None
        if ss_val or tr_val:
            bug_code = row[item_code_col_bugs] if item_code_col_bugs < len(row) else None
            bug_migrations.append({
                "row": row_num,
                "bug_code": bug_code,
                "screenshot_ref": ss_val,
                "trace_ref": tr_val,
            })

    return {
        "step_rows": step_rows,
        "steps_by_tc": steps_by_tc,
        "step_code_set": step_code_set,
        "tcs_kp_review": tcs_kp_review,
        "tcs_over_20": tcs_over_20,
        "tcs_empty_steps": tcs_empty_steps,
        "orphans": orphans,
        "bug_migrations": bug_migrations,
    }


# ---------------------------------------------------------------------------
# mutation
# ---------------------------------------------------------------------------

def _apply_f1_f2(wb: openpyxl.Workbook, analysis: dict, run_ts: datetime.datetime,
                 verbose: bool) -> dict:
    """
    F-1: Create or validate 02e_TestSteps.
    F-2: Add / update steps_count in 02_TestCases.
    Returns idempotency info.
    """
    step_rows = analysis["step_rows"]
    steps_by_tc = analysis["steps_by_tc"]
    idempotency_drift: list[str] = []
    already_existed = False

    # ---- F-1 ----
    if "02e_TestSteps" in wb.sheetnames:
        already_existed = True
        _log(verbose, "[F-1] 02e_TestSteps already exists — validating idempotency")
        ws_e = wb["02e_TestSteps"]
        existing_codes = set()
        for row in ws_e.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and row[1] is not None:
                existing_codes.add(row[1])  # step_code is col 2 (index 1)
        expected_codes = {s["step_code"] for s in step_rows}
        drift = (existing_codes - expected_codes) | (expected_codes - existing_codes)
        if drift:
            idempotency_drift = sorted(drift)
    else:
        _log(verbose, "[F-1] Creating 02e_TestSteps")
        ws_e = wb.create_sheet("02e_TestSteps")
        ws_e.append(STEP_COLS)
        for s in step_rows:
            ws_e.append([s[c] for c in STEP_COLS])

    # ---- F-2 ----
    ws_tc = wb["02_TestCases"]
    cmap_tc = _col_map(ws_tc)
    sc_col = _col_index(ws_tc, STEPS_COUNT_COL_HEADER)
    if sc_col is None:
        _log(verbose, "[F-2] Adding steps_count column to 02_TestCases")
        max_col = ws_tc.max_column
        sc_col = max_col + 1
        ws_tc.cell(row=1, column=sc_col, value=STEPS_COUNT_COL_HEADER)

    item_code_col = cmap_tc.get(TC_CODE_COL_HEADER, 2)
    for row in ws_tc.iter_rows(min_row=2):
        tc_code = row[item_code_col - 1].value
        if tc_code is None:
            continue
        count = len(steps_by_tc.get(tc_code, []))
        row[sc_col - 1].value = count

    return {"idempotency_drift": idempotency_drift, "f1_already_existed": already_existed}


def _apply_f4(wb: openpyxl.Workbook, analysis: dict, verbose: bool) -> None:
    """F-4: Add evidence_* columns to 08_Bugs; migrate legacy values."""
    ws_bugs = wb["08_Bugs"]
    cmap = _col_map(ws_bugs)
    ss_col = cmap.get("screenshot_ref")
    tr_col = cmap.get("trace_ref")

    # Ensure new columns exist (append right if missing)
    added_cols: dict[str, int] = {}
    for col_name in BUG_NEW_COLS:
        existing = _col_index(ws_bugs, col_name)
        if existing is None:
            max_col = ws_bugs.max_column
            new_col = max_col + 1
            ws_bugs.cell(row=1, column=new_col, value=col_name)
            added_cols[col_name] = new_col
            _log(verbose, f"[F-4] Added column {col_name} at col {new_col}")
        else:
            added_cols[col_name] = existing

    ls_col = added_cols["linked_step_ref"]
    esp_col = added_cols["evidence_screenshot_path"]
    evp_col = added_cols["evidence_video_path"]
    etp_col = added_cols["evidence_trace_path"]
    eck_col = added_cols["evidence_capture_kind"]
    eca_col = added_cols["evidence_capture_at"]

    for row in ws_bugs.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        ss_val = row[ss_col - 1].value if ss_col else None
        tr_val = row[tr_col - 1].value if tr_col else None

        esp_cell = row[esp_col - 1]
        etp_cell = row[etp_col - 1]
        eck_cell = row[eck_col - 1]

        # only migrate when typed columns are empty AND legacy is non-empty
        if ss_val and not esp_cell.value:
            esp_cell.value = ss_val
        if tr_val and not etp_cell.value:
            etp_cell.value = tr_val
        if (ss_val or tr_val) and not eck_cell.value:
            eck_cell.value = "manual-tester"
        elif not eck_cell.value:
            eck_cell.value = "none"


def _apply_f5(wb: openpyxl.Workbook, run_ts: datetime.datetime, verbose: bool) -> None:
    """F-5: Append changelog row to 11_Changelog."""
    ws = wb["11_Changelog"]
    # Find column indices from the header row (row 3 in real workbook)
    # Scan rows 1..5 for the row that looks like a header
    header_row_idx = None
    for i in range(1, 6):
        row_vals = [c.value for c in ws[i]]
        if "Version" in row_vals:
            header_row_idx = i
            break
    if header_row_idx is None:
        _log(verbose, "[F-5] Could not find Version header row; appending raw row")
        ws.append([
            "v0.4.4", run_ts,
            "workbook-v0.4.3-to-v0.4.4.py",
            "Added 02e_TestSteps; added evidence columns to 08_Bugs; "
            "added steps_count to 02_TestCases; soft-deprecated screenshot_ref + trace_ref in 08_Bugs.",
        ])
        return

    cmap = _col_map(ws, header_row=header_row_idx)
    ver_col = cmap.get("Version", 1)
    date_col = cmap.get("Date", 2)
    auth_col = cmap.get("Author", 3)
    summ_col = cmap.get("Change summary", 4)

    max_cols = max(ver_col, date_col, auth_col, summ_col)
    new_row = [None] * max_cols
    new_row[ver_col - 1] = "v0.4.4"
    new_row[date_col - 1] = run_ts
    new_row[auth_col - 1] = "workbook-v0.4.3-to-v0.4.4.py"
    new_row[summ_col - 1] = (
        "Added 02e_TestSteps; added evidence columns to 08_Bugs; "
        "added steps_count to 02_TestCases; soft-deprecated screenshot_ref + trace_ref in 08_Bugs."
    )
    ws.append(new_row)
    _log(verbose, "[F-5] Appended changelog row for v0.4.4")


# ---------------------------------------------------------------------------
# PATCH-REPORT
# ---------------------------------------------------------------------------

def _write_report(
    analysis: dict,
    idempotency_info: dict,
    source_path: pathlib.Path,
    dest_path: pathlib.Path | None,
    run_ts: datetime.datetime,
    dry_run: bool,
    migration_stats: dict | None = None,
    source_data_path: pathlib.Path | None = None,
) -> pathlib.Path:
    ts_str = run_ts.strftime("%Y%m%d-%H%M%S")
    report_path = REPORT_DIR / f"PATCH-REPORT-v0.4.3-to-v0.4.4-{ts_str}.md"
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    step_rows = analysis["step_rows"]
    steps_by_tc = analysis["steps_by_tc"]
    tcs_kp = analysis["tcs_kp_review"]
    tcs_over_20 = analysis["tcs_over_20"]
    tcs_empty = analysis["tcs_empty_steps"]
    orphans = analysis["orphans"]
    bug_migrations = analysis["bug_migrations"]
    drift = idempotency_info.get("idempotency_drift", [])

    src_sha = _sha256(source_path)
    dest_sha = _sha256(dest_path) if (dest_path and dest_path.exists()) else "(not written — dry-run)"

    lines: list[str] = []
    lines.append(f"# PATCH-REPORT v0.4.3 → v0.4.4")
    lines.append(f"")
    lines.append(f"Run timestamp: {run_ts.isoformat()}")
    lines.append(f"Source: `{source_path}`")
    lines.append(f"Dest:   `{dest_path}` {'(dry-run — not written)' if dry_run else ''}")
    lines.append(f"")

    # §1 Summary
    lines.append(f"## §1 Summary")
    lines.append(f"")
    lines.append(f"| Metric | Count |")
    lines.append(f"|--------|-------|")
    lines.append(f"| TCs scanned | {len(steps_by_tc)} |")
    lines.append(f"| Steps generated | {len(step_rows)} |")
    lines.append(f"| Bugs with evidence migrated | {len(bug_migrations)} |")
    lines.append(f"| Orphan 02c step_id refs | {len(orphans)} |")
    lines.append(f"| TCs needing KP step-cleanup | {len(tcs_kp)} |")
    lines.append(f"| TCs with empty steps_summary | {len(tcs_empty)} |")
    lines.append(f"| TCs with >20 steps | {len(tcs_over_20)} |")
    if drift:
        lines.append(f"| Idempotency drift items | {len(drift)} |")
    lines.append(f"")
    if drift:
        lines.append(f"**IDEMPOTENCY DRIFT DETECTED** — existing 02e_TestSteps differs "
                     f"from what a fresh run would produce. Step codes in drift:")
        for code in drift:
            lines.append(f"  - `{code}`")
        lines.append(f"")

    # §2 TCs needing KP step-cleanup
    lines.append(f"## §2 TCs needing KP step-cleanup")
    lines.append(f"")
    if tcs_kp:
        for tc in tcs_kp:
            lines.append(f"- `{tc}`")
    else:
        lines.append(f"_(none)_")
    lines.append(f"")

    # §3 TCs with >20 steps
    lines.append(f"## §3 TCs with >20 steps (cap-exceeded — review needed)")
    lines.append(f"")
    if tcs_over_20:
        for tc in tcs_over_20:
            lines.append(f"- `{tc}` ({len(steps_by_tc.get(tc, []))} steps)")
    else:
        lines.append(f"_(none)_")
    lines.append(f"")

    # §4 TCs with empty steps_summary
    lines.append(f"## §4 TCs with empty steps_summary (placeholder rows generated)")
    lines.append(f"")
    if tcs_empty:
        for tc in tcs_empty:
            lines.append(f"- `{tc}`")
    else:
        lines.append(f"_(none)_")
    lines.append(f"")

    # §5 Orphan step_id refs
    lines.append(f"## §5 Orphan 02c_TC_Assertions.step_id values")
    lines.append(f"")
    if orphans:
        lines.append(f"These rows were **not modified**. KP decides whether to relink, delete, or backfill.")
        lines.append(f"")
        lines.append(f"| Row # | tc_ref | step_id (orphan) |")
        lines.append(f"|-------|--------|-----------------|")
        for o in orphans:
            lines.append(f"| {o['row']} | `{o['tc_ref']}` | `{o['step_id']}` |")
    else:
        lines.append(f"_(none)_")
    lines.append(f"")

    # §6 Bug evidence migrations
    lines.append(f"## §6 Bugs migrated from legacy screenshot_ref / trace_ref")
    lines.append(f"")
    if bug_migrations:
        lines.append(f"| Row # | bug_code | screenshot_ref → evidence_screenshot_path | trace_ref → evidence_trace_path |")
        lines.append(f"|-------|----------|-------------------------------------------|---------------------------------|")
        for m in bug_migrations:
            ss = m["screenshot_ref"] or ""
            tr = m["trace_ref"] or ""
            lines.append(f"| {m['row']} | `{m['bug_code']}` | `{ss}` | `{tr}` |")
        lines.append(f"")
        lines.append(f"**Soft-deprecated columns** — `screenshot_ref` and `trace_ref` were "
                     f"**NOT deleted** from 08_Bugs. Existing readers continue to work until "
                     f"v0.1.5 ships. They are marked soft-deprecated by this patch.")
    else:
        lines.append(f"_(none)_")
    lines.append(f"")

    # §7 Idempotency stamp
    lines.append(f"## §7 Idempotency stamp")
    lines.append(f"")
    lines.append(f"| File | SHA-256 |")
    lines.append(f"|------|---------|")
    lines.append(f"| source (v0.4.3) | `{src_sha}` |")
    lines.append(f"| dest (v0.4.4)   | `{dest_sha}` |")
    lines.append(f"")

    # §8 What was NOT done
    lines.append(f"## §8 What was NOT done")
    lines.append(f"")
    lines.append(textwrap.dedent("""\
        - `tc_ref` (08_Bugs col 29) deprecation rewire deferred to patcher v0.4.4 → v0.4.5
        - KP translations of `action_en`, `expected_cs/en` in 02e_TestSteps — left empty
        - Step-level `expected_cs/en` harvest from `expected_summary` — left empty
        - `bouracka_ui/workbook_io.py::list_steps()` — deferred to next brief
        - `02b_TC_Parameters` step-scoped params — out of scope (v0.2 if needed)
        - API endpoints (§4.2/§4a.5 of design notes) — out of scope for this brief
        - Dispatcher artefact-copy logic (FR-K-004) — out of scope
    """))

    # §9-§12: data migration sections (only when --source-data was used)
    if migration_stats is not None:
        lines.append(f"## §9 Data migration summary")
        lines.append(f"")
        lines.append(f"Source-data: `{source_data_path}`")
        lines.append(f"")
        lines.append(f"| Sheet | Rows migrated | Schema rows replaced |")
        lines.append(f"|-------|---------------|----------------------|")
        for sheet_name, s in migration_stats.items():
            lines.append(f"| `{sheet_name}` | {s['migrated']} | {s['replaced']} |")
        lines.append(f"")

        # §10 Legacy evidence migrations from 08_Bugs
        lines.append(f"## §10 08_Bugs rows with legacy evidence migrated")
        lines.append(f"")
        all_legacy = migration_stats.get("08_Bugs", {}).get("legacy_evidence", [])
        if all_legacy:
            lines.append(f"| bug_code | screenshot_ref | trace_ref |")
            lines.append(f"|----------|---------------|-----------|")
            for e in all_legacy:
                lines.append(f"| `{e['code']}` | `{e['screenshot_ref'] or ''}` | `{e['trace_ref'] or ''}` |")
        else:
            lines.append(f"_(none)_")
        lines.append(f"")

        # §11 Row-code collisions (source-data row replaced dest schema-source row)
        lines.append(f"## §11 Row-code collisions (schema rows overwritten by user data)")
        lines.append(f"")
        total_replaced = sum(s["replaced"] for s in migration_stats.values())
        if total_replaced:
            for sheet_name, s in migration_stats.items():
                if s["replaced"]:
                    lines.append(f"- `{sheet_name}`: {s['replaced']} schema row(s) replaced")
        else:
            lines.append(f"_(none — dest data sheets were empty before migration)_")
        lines.append(f"")

        # §12 Sheets NOT migrated
        lines.append(f"## §12 Sheets explicitly NOT migrated (schema-owned)")
        lines.append(f"")
        lines.append(f"The following sheets are schema/canonical content owned by the repo. "
                     f"Source-data versions were ignored.")
        lines.append(f"")
        for name in sorted(SCHEMA_ONLY_SHEETS):
            lines.append(f"- `{name}`")
        lines.append(f"")

    report_path.write_text("\n".join(lines), encoding="utf-8")
    return report_path


# ---------------------------------------------------------------------------
# report-only mode
# ---------------------------------------------------------------------------

def _report_only(dest_path: pathlib.Path, run_ts: datetime.datetime, verbose: bool) -> int:
    """Re-emit PATCH-REPORT from an existing v0.4.4 workbook."""
    if not dest_path.exists():
        print(f"ERROR: --report-only requires dest workbook to exist: {dest_path}", file=sys.stderr)
        return 2

    _log(verbose, f"[report-only] Loading {dest_path}")
    wb = load_workbook(dest_path, read_only=True)
    analysis = _analyse(wb, run_ts, verbose)
    wb.close()
    report = _write_report(analysis, {}, dest_path, dest_path, run_ts, dry_run=False)
    print(f"Report: {report}")
    return 0


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def patch(
    source: pathlib.Path,
    dest: pathlib.Path,
    dry_run: bool = False,
    report_only: bool = False,
    verbose: bool = False,
    source_data: pathlib.Path | None = None,
) -> int:
    run_ts = datetime.datetime.now().replace(microsecond=0)

    if report_only:
        return _report_only(dest, run_ts, verbose)

    # --- validate source ---
    if not source.exists():
        print(f"ERROR: source workbook not found: {source}", file=sys.stderr)
        return 2

    # --- validate source-data (F-1m) ---
    wb_data: openpyxl.Workbook | None = None
    if source_data is not None:
        if not source_data.exists():
            print(f"ERROR: --source-data workbook not found: {source_data}", file=sys.stderr)
            return 2
        if source_data.resolve() == source.resolve():
            print(f"WARNING: --source-data and --source point to the same file; "
                  f"data migration is a no-op (risk gate 5)", file=sys.stderr)
        _log(verbose, f"Loading source-data: {source_data}")
        try:
            wb_data = load_workbook(str(source_data))
        except Exception as exc:
            print(f"ERROR: openpyxl failed to read {source_data}: {exc}", file=sys.stderr)
            return 2
        dup_errors = _validate_source_data_duplicates(wb_data)
        if dup_errors:
            for err in dup_errors:
                print(f"ERROR (data integrity): {err}", file=sys.stderr)
            wb_data.close()
            return 4

    _log(verbose, f"Loading source: {source}")
    try:
        wb_source = load_workbook(source, read_only=True)
    except Exception as exc:
        print(f"ERROR: openpyxl failed to read {source}: {exc}", file=sys.stderr)
        return 2

    required_sheets = {"02_TestCases", "02c_TC_Assertions", "08_Bugs", "11_Changelog"}
    missing = required_sheets - set(wb_source.sheetnames)
    if missing:
        print(f"ERROR: source workbook missing sheets: {missing}", file=sys.stderr)
        wb_source.close()
        return 2

    # --- analyse (read-only) ---
    _log(verbose, "Analysing…")
    try:
        analysis = _analyse(wb_source, run_ts, verbose)
    except Exception as exc:
        print(f"ERROR during analysis: {exc}", file=sys.stderr)
        wb_source.close()
        return 2
    wb_source.close()

    _log(verbose, f"  {len(analysis['steps_by_tc'])} TCs → {len(analysis['step_rows'])} steps")
    _log(verbose, f"  {len(analysis['orphans'])} orphan 02c refs")
    _log(verbose, f"  {len(analysis['bug_migrations'])} bugs with legacy evidence refs")

    migration_stats: dict | None = None

    if dry_run:
        _log(verbose, "[dry-run] Skipping workbook write")
        report = _write_report(analysis, {}, source, None, run_ts, dry_run=True,
                               migration_stats=migration_stats, source_data_path=source_data)
        print(f"[dry-run] Report: {report}")
    else:
        source = source.resolve()
        dest = dest.resolve()

        # Load source into memory (non-read-only so it can be mutated and saved)
        _log(verbose, f"Loading {source} for mutation…")
        try:
            wb = load_workbook(source)
        except Exception as exc:
            print(f"ERROR: openpyxl failed to load {source}: {exc}", file=sys.stderr)
            return 3

        _log(verbose, "Applying F-1/F-2 (02e_TestSteps + steps_count)…")
        idem_info = _apply_f1_f2(wb, analysis, run_ts, verbose)

        _log(verbose, "Applying F-4 (08_Bugs evidence columns)…")
        _apply_f4(wb, analysis, verbose)

        _log(verbose, "Applying F-5 (11_Changelog row)…")
        _apply_f5(wb, run_ts, verbose)

        # --- F-1m..F-5m: data migration (only when --source-data provided) ---
        if wb_data is not None:
            _log(verbose, "Applying data migration (F-1m..F-5m)…")
            migration_stats = _migrate_data(wb, wb_data, verbose)
            # Second changelog row for data migration
            ws_cl = wb["11_Changelog"]
            header_row_idx = None
            for i in range(1, 6):
                row_vals = [c.value for c in ws_cl[i]]
                if "Version" in row_vals:
                    header_row_idx = i
                    break
            if header_row_idx:
                cmap_cl = _col_map(ws_cl, header_row=header_row_idx)
                ver_c = cmap_cl.get("Version", 1)
                date_c = cmap_cl.get("Date", 2)
                auth_c = cmap_cl.get("Author", 3)
                summ_c = cmap_cl.get("Change summary", 4)
                max_c = max(ver_c, date_c, auth_c, summ_c)
                cl_row = [None] * max_c
                cl_row[ver_c - 1] = "v0.4.4-data"
                cl_row[date_c - 1] = run_ts
                cl_row[auth_c - 1] = "workbook-v0.4.3-to-v0.4.4.py --source-data"
                total = sum(s["migrated"] for s in migration_stats.values())
                cl_row[summ_c - 1] = (
                    f"Data migration from {source_data.name}: "
                    f"{total} rows across {len(migration_stats)} sheets."
                )
                ws_cl.append(cl_row)
            wb_data.close()
            wb_data = None

        _log(verbose, f"Saving {dest}…")
        try:
            # Save to a sibling temp file first, then rename — avoids Windows
            # permission-denied when saving back to the same path as the loaded copy.
            tmp_fd, tmp_path_str = tempfile.mkstemp(
                dir=dest.parent, prefix=".tmp-patcher-", suffix=".xlsx"
            )
            tmp_out = pathlib.Path(tmp_path_str)
            try:
                os.close(tmp_fd)
                wb.save(tmp_out)
                tmp_out.replace(dest)
            except Exception:
                tmp_out.unlink(missing_ok=True)
                raise
        except Exception as exc:
            print(f"ERROR: failed to save {dest}: {exc}", file=sys.stderr)
            return 3

        report = _write_report(analysis, idem_info, source, dest, run_ts, dry_run=False,
                               migration_stats=migration_stats, source_data_path=source_data)
        print(f"Output: {dest}")
        print(f"Report: {report}")

    if wb_data is not None:
        wb_data.close()

    # determine exit code
    has_warnings = (
        bool(analysis["orphans"])
        or bool(analysis["tcs_over_20"])
        or bool(analysis.get("tcs_kp_review"))
    )
    return 1 if has_warnings else 0


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(
        description="Patch BOURACKA-TESTPLAN-v0.4.3.xlsx to v0.4.4",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--source", type=pathlib.Path, default=DEFAULT_SOURCE)
    p.add_argument("--dest", type=pathlib.Path, default=DEFAULT_DEST)
    p.add_argument("--source-data", type=pathlib.Path, default=None,
                   dest="source_data",
                   help="tester's workbook with user data to migrate (F-1m..F-7m)")
    p.add_argument("--dry-run", action="store_true",
                   help="analyse only; write report but not dest workbook")
    p.add_argument("--report-only", action="store_true",
                   help="re-emit report from existing dest workbook (no schema work)")
    p.add_argument("-v", "--verbose", action="store_true")
    args = p.parse_args(argv)
    return patch(args.source, args.dest, args.dry_run, args.report_only, args.verbose,
                 source_data=args.source_data)


if __name__ == "__main__":
    sys.exit(main())

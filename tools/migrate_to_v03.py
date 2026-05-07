#!/usr/bin/env python3
"""
migrate_to_v03.py — apply CP-SUPIN-04 STEP 12 schema changes to v0.3 workbook.

What v0.3 adds vs v0.2:
1. 01_TestTargets: 6 new TTs from live evidence (TT-CP-PHASE0..PHASE5 surface).
2. 02_TestCases: 25+ new TC rows (TC-CP-NEW-A..Y) per flow.md §6 + new column
   `env_constraints` (both | demo-only | prod-only | both-with-adapter) per
   L-WORK-11.
3. 11_Changelog: rev9 entry.
4. 09_Reports: bottom-up TT/TC consistency check report row.

Run from repo root with: python tools/migrate_to_v03.py
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
WB_PATH = ROOT / "BOURACKA-TESTPLAN-v0.3.xlsx"

# Header & style helpers
HDR_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF")
DEMO_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow
PROD_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # light blue
BOTH_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # light green


def add_env_constraints_column(wb):
    """Add `env_constraints` column to 02_TestCases (per L-WORK-11)."""
    ws = wb["02_TestCases"]
    headers = [c.value for c in ws[1]]
    if "env_constraints" in headers:
        print("[skip] env_constraints column already present")
        return
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col, value="env_constraints").fill = HDR_FILL
    ws.cell(row=1, column=new_col).font = HDR_FONT
    # Backfill existing TCs with "both-with-adapter" as default
    # (TC-CP-005 SMS gets "both-with-adapter" — Mockoon profile)
    backfill = {
        "TC-CP-001": "both",
        "TC-CP-002": "both",
        "TC-CP-003": "both",
        "TC-CP-004": "both",
        "TC-CP-005": "both-with-adapter",  # Mockoon for PROD; DEMO native
        "TC-CP-006": "both-with-adapter",
        "TC-CP-007": "demo-only",  # zenID OCR mocked
        "TC-CP-008": "both-with-adapter",
        "TC-CP-009": "demo-only",  # AISPOV ROB sandbox
        "TC-CP-010": "demo-only",
        "TC-CP-011": "both",
        "TC-CP-012": "both",
        "TC-CP-013": "demo-only",
        "TC-CP-014": "demo-only",
        "TC-CP-015": "demo-only",
        "TC-CP-016": "both",
        "TC-CP-017": "both",
        "TC-CP-018": "both",
        "TC-CP-019": "both",
        "TC-CP-020": "both",
        "TC-CP-021": "both",
        "TC-CP-022": "prod-only",  # real email dispatch
        "TC-CP-023": "both",
    }
    for r in range(2, ws.max_row + 1):
        tc_code = ws.cell(row=r, column=1).value
        if tc_code in backfill:
            cell = ws.cell(row=r, column=new_col, value=backfill[tc_code])
            v = backfill[tc_code]
            if v == "demo-only":
                cell.fill = DEMO_FILL
            elif v == "prod-only":
                cell.fill = PROD_FILL
            else:
                cell.fill = BOTH_FILL
    print(f"[ok] env_constraints column added (col {new_col}) + backfilled {len(backfill)} existing TCs")


# ─────────────────────────────────────────────────────────────────────────
# New TTs from live evidence (per flow.md §2.2 + §3)
# ─────────────────────────────────────────────────────────────────────────

NEW_TTS = [
    # (item_code, name_cs, name_en, scope_in_r1, related_endpoints, related_screens)
    ("TT-CP-PHASE0", "Rozcestník + safety guidance", "Gateway page + safety guidance",
     "primary entry; emergency-call escape; police-criteria reveal",
     "(none — static SPA shell + outage feed INT-006)",
     "/formular/ rozcestník"),
    ("TT-CP-PHASE-A", "Pre-wizard intro", "Pre-wizard intro",
     "explicit user acknowledgement of 5-step process before report mint",
     "POST /api/reports",
     "/formular/informations"),
    ("TT-CP-PHASE1", "Phone verification (OTP rounds)", "Phone verification",
     "Czech/Slovak mobile + OTP for both participants; PROD via N8 SMS",
     "PUT /reporter, POST /participants, POST /sendCodeToVerify, POST /verify, GET /participantsInVerification",
     "/verification + /verification/success"),
    ("TT-CP-PHASE2", "Documents per participant (OP+ŘP+address+email)", "Documents per participant",
     "OCR or manual fallback; RUIAN address resolution; ŘP categories codelist",
     "PUT /participantData, /email, /drivingLicense/update + RUIAN INT-009",
     "/documents/{p}/{...}/manual /recap"),
    ("TT-CP-PHASE2.5", "Witnesses (optional)", "Witnesses",
     "up to 3 witnesses; skip-able",
     "GET /witnesses (POST/PUT TBC)",
     "/witness"),
    ("TT-CP-PHASE3-PER", "Per-participant accident details", "Per-participant accident details",
     "photos + damage zones + movement + vehicle + insurer per participant",
     "PUT damage, movementDefinition, vehicle, insurer + GET enumerations",
     "/accident/{p}/[damage|damage/movement|damage/data]"),
    ("TT-CP-PHASE3-SH", "Shared circumstances + place", "Shared circumstances + place",
     "accident-type + free-text + datetime + RUIAN + Google Maps + culprit",
     "PUT circumstances, datetime, location, responsibleParticipant + INT-009 + INT-007",
     "/accident/circumstances /situation /situation/location /culprit"),
    ("TT-CP-PHASE4", "Sign + submit (dual SMS)", "Sign + submit",
     "summary review → confirmation → SMS-sign per participant → PDF + email dispatch",
     "POST /confirmation, /sendCodeToSign, /sign + GET /summaryId",
     "/summary /sign-report /success"),
    ("TT-CP-IDENTIFIER", "Identifikační kód for Police/IZS", "Identifikační kód for Police/IZS",
     "format XXXXXX-YY-NNNNNNN; persistent header modal; QR code companion",
     "(rendered from report state — drives police on-scene lookup)",
     "(persistent modal across all post-Phase-1 routes)"),
    ("TT-CP-CODELISTS", "Codelist API surface", "Codelist API surface",
     "public: insuranceCompanies (13), vehicleBrands (275); protected: 8× 403",
     "GET /api/enumerations/* (selective public access)",
     "(used internally by Phase 2 + 3)"),
]


def add_new_tts(wb):
    ws = wb["01_TestTargets"]
    headers = [c.value for c in ws[1]]
    print(f"[info] 01_TestTargets has {ws.max_row} rows + {len(headers)} cols")

    # Find columns to fill
    col_idx = {h: i + 1 for i, h in enumerate(headers) if h}
    code_col = col_idx.get("item_code")
    name_cs_col = col_idx.get("item_name_cs") or col_idx.get("name_cs")
    name_en_col = col_idx.get("item_name_en") or col_idx.get("name_en")
    scope_col = col_idx.get("scope_in_r1") or col_idx.get("scope")
    notes_col = col_idx.get("notes")

    if not (code_col and name_cs_col):
        print("[fail] required columns not found:", col_idx)
        return

    next_row = ws.max_row + 1
    for tt in NEW_TTS:
        ws.cell(row=next_row, column=code_col, value=tt[0])
        if name_cs_col:
            ws.cell(row=next_row, column=name_cs_col, value=tt[1])
        if name_en_col:
            ws.cell(row=next_row, column=name_en_col, value=tt[2])
        if scope_col:
            ws.cell(row=next_row, column=scope_col, value=tt[3])
        if notes_col:
            ws.cell(row=next_row, column=notes_col, value=f"endpoints: {tt[4]} | screens: {tt[5]}")
        next_row += 1
    print(f"[ok] added {len(NEW_TTS)} new TTs (rows {ws.max_row + 1 - len(NEW_TTS)} .. {ws.max_row})")


# ─────────────────────────────────────────────────────────────────────────
# New TCs from live evidence (per flow.md §6)
# ─────────────────────────────────────────────────────────────────────────

NEW_TCS = [
    # (tc_code, title_cs, target_tt, env_constraints, priority, notes)
    ("TC-CP-NEW-A", "Rozcestník copy integrity (3 karty + 3 safety bullets)",
     "TT-CP-PHASE0", "both", "P2",
     "asserts STR-001..005; catches PROD-bundle-leak into DEMO if banner missing"),
    ("TC-CP-NEW-B", "R1 scope sentence verbatim (200 000 Kč)",
     "TT-CP-PHASE0", "both", "P1",
     "STR-003; legal anchor for R1 scope"),
    ("TC-CP-NEW-C", "DEMO banner present on DEMO; absent on PROD",
     "TT-CP-PHASE0", "both", "P2",
     "Δ11; STR-002; env-aware screenshot baseline"),
    ("TC-CP-NEW-D", "Police-criteria expand interaction (7 bullets)",
     "TT-CP-PHASE0", "both", "P3",
     "expand-collapse; STR-* TBC"),
    ("TC-CP-NEW-E", "Pre-wizard intro renders before report mint",
     "TT-CP-PHASE-A", "both", "P2",
     "STR-006..009; route /informations; click 'Rozumím' → POST /reports"),
    ("TC-CP-NEW-F", "POST /api/reports envelope returns UUID-v4",
     "TT-CP-CODELISTS", "both", "P1",
     "API contract; schema assertion via Zod-mirror"),
    ("TC-CP-NEW-G", "Deep-link /verification works after manual POST",
     "TT-CP-PHASE1", "both", "P2",
     "automation-friendly: pre-mint UUID via API, jump to /verification"),
    ("TC-CP-NEW-H", "Slovak +421 predvolba selectable + accepted",
     "TT-CP-PHASE1", "both", "P3",
     "DEMO hint claims SK accepted; verify form actually accepts"),
    ("TC-CP-NEW-I", "GDPR consent required for A; absent for B",
     "TT-CP-PHASE1", "both", "P2",
     "asymmetric consent; legal-relevance"),
    ("TC-CP-NEW-J", "Phase 1 DEMO hint absent on PROD",
     "TT-CP-PHASE1", "prod-only", "P1",
     "Δ1 negation; PROD MUST send real SMS"),
    ("TC-CP-NEW-K", "Data-purge-on-abandon (per Phase 1 disclaimer)",
     "TT-CP-PHASE1", "both", "P2",
     "STR-013; GDPR-aligned; verify by abandoning + re-checking GET /reports/{r}"),
    ("TC-CP-NEW-L", "Outage-feed reachable; banner state mirrors feed",
     "TT-CP-PHASE0", "both", "P3",
     "INT-006; orange banner content from odstavky.json"),
    ("TC-CP-NEW-M", "Map locale = cs (not en_gb)",
     "TT-CP-PHASE3-SH", "both", "P3",
     "INT-007 defect candidate; Δ23"),
    ("TC-CP-NEW-N", "POST /api/reports rate-limited (anti-abuse)",
     "TT-CP-CODELISTS", "both-with-adapter", "P3",
     "spam 100 POSTs; expect 429 after threshold"),
    ("TC-CP-NEW-O", "DEMO accepts ANY 4-digit OTP (verify + sign)",
     "TT-CP-PHASE1", "demo-only", "P1",
     "Δ1 confirmed; smoke for the bypass behaviour"),
    ("TC-CP-NEW-P", "PROD rejects arbitrary 4-digit OTP",
     "TT-CP-PHASE1", "prod-only", "P1",
     "negation of NEW-O; needs N8 sandbox"),
    ("TC-CP-NEW-Q", "Identifikační kód format XXXXXX-YY-NNNNNNN",
     "TT-CP-IDENTIFIER", "both", "P2",
     "regex assert; year segment matches current year"),
    ("TC-CP-NEW-R", "Police/IZS QR modal accessible from header",
     "TT-CP-IDENTIFIER", "both", "P2",
     "header 4-dot icon post-verification opens drawer"),
    ("TC-CP-NEW-S", "Phase 2 manual deep-link ?validate=false respected",
     "TT-CP-PHASE2", "both", "P2",
     "automation hook; relaxes OCR validation but not required-field"),
    ("TC-CP-NEW-T", "Phase 2 driver-registry: DEMO no-op; PROD AISPOV-CRR",
     "TT-CP-PHASE2", "both-with-adapter", "P1",
     "Δ27; Načíst-z-registru button behaviour split"),
    ("TC-CP-NEW-U", "Phase 2 RUIAN address autocomplete (3+ chars)",
     "TT-CP-PHASE2", "both", "P2",
     "INT-009; ČÚZK ArcGIS suggest endpoint"),
    ("TC-CP-NEW-V", "ŘP number regex ^[A-Z]{2} \\d{6}$ enforced",
     "TT-CP-PHASE2", "both", "P2",
     "format hint XY 000000; client-side validation"),
    ("TC-CP-NEW-W", "Birth date DD.MM.YYYY enforced + DatePicker functional",
     "TT-CP-PHASE2", "both", "P3",
     "MUI DatePicker; locale cs"),
    ("TC-CP-NEW-X", "Pre-wizard 20-video pre-fetch is non-blocking",
     "TT-CP-PHASE2", "both", "P3",
     "Phase 1 form usable while videos download"),
    ("TC-CP-NEW-Y", "ČÚZK outage degrades gracefully to manual address",
     "TT-CP-PHASE2", "both", "P3",
     "INT-009 down; 'Zadat adresu manuálně' fallback works"),
]


def add_new_tcs(wb):
    ws = wb["02_TestCases"]
    headers = [c.value for c in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(headers) if h}

    code_col = col_idx.get("tc_code") or col_idx.get("item_code")
    title_col = col_idx.get("title_cs") or col_idx.get("item_name_cs") or col_idx.get("name_cs")
    target_col = col_idx.get("test_target_ref")
    env_col = col_idx.get("env_constraints")
    priority_col = col_idx.get("priority")
    notes_col = col_idx.get("notes")

    if not (code_col and title_col):
        print("[fail] required columns not found in 02_TestCases:", col_idx)
        return

    next_row = ws.max_row + 1
    for tc in NEW_TCS:
        ws.cell(row=next_row, column=code_col, value=tc[0])
        ws.cell(row=next_row, column=title_col, value=tc[1])
        if target_col:
            ws.cell(row=next_row, column=target_col, value=tc[2])
        if env_col:
            cell = ws.cell(row=next_row, column=env_col, value=tc[3])
            v = tc[3]
            if v == "demo-only":
                cell.fill = DEMO_FILL
            elif v == "prod-only":
                cell.fill = PROD_FILL
            else:
                cell.fill = BOTH_FILL
        if priority_col:
            ws.cell(row=next_row, column=priority_col, value=tc[4])
        if notes_col:
            ws.cell(row=next_row, column=notes_col, value=tc[5])
        next_row += 1
    print(f"[ok] added {len(NEW_TCS)} new TCs")


# ─────────────────────────────────────────────────────────────────────────
# Consistency check (logical: every TC → existing TT; every TT → ≥1 TC)
# ─────────────────────────────────────────────────────────────────────────

def consistency_check(wb):
    """Print TT/TC traceability + flag orphans."""
    tt_ws = wb["01_TestTargets"]
    tc_ws = wb["02_TestCases"]

    # Resolve column indices by header name
    def col_of(ws, name):
        for i, c in enumerate(ws[1], 1):
            if c.value == name:
                return i
        return None

    tt_code_col = col_of(tt_ws, "item_code")
    tc_code_col = col_of(tc_ws, "item_code")
    target_col_idx = col_of(tc_ws, "test_target_ref")

    if not (tt_code_col and tc_code_col and target_col_idx):
        print(f"[warn] required columns missing — tt:{tt_code_col} tc:{tc_code_col} target:{target_col_idx}")
        return

    tt_codes = set()
    for r in range(2, tt_ws.max_row + 1):
        c = tt_ws.cell(row=r, column=tt_code_col).value
        if c:
            tt_codes.add(str(c).strip())

    tc_targets = {}
    for r in range(2, tc_ws.max_row + 1):
        tc = tc_ws.cell(row=r, column=tc_code_col).value
        tt = tc_ws.cell(row=r, column=target_col_idx).value
        if tc and tt:
            tc_targets[str(tc).strip()] = str(tt).strip()

    # 1. Every TC's target exists?
    orphan_tc = [(tc, tt) for tc, tt in tc_targets.items() if tt and tt not in tt_codes]
    # 2. Every TT has ≥1 TC?
    targets_used = {tt for tt in tc_targets.values() if tt}
    unused_tt = tt_codes - targets_used

    print()
    print("─" * 60)
    print("CONSISTENCY CHECK (TT ↔ TC traceability)")
    print("─" * 60)
    print(f"  Total TTs: {len(tt_codes)}")
    print(f"  Total TCs: {len(tc_targets)}")
    print(f"  Orphan TCs (target not in TT list): {len(orphan_tc)}")
    if orphan_tc:
        for tc, tt in orphan_tc:
            print(f"    {tc} → {tt}  (TT not defined)")
    print(f"  Unused TTs (no TC references): {len(unused_tt)}")
    for tt in sorted(unused_tt):
        print(f"    {tt}")
    print()


# ─────────────────────────────────────────────────────────────────────────
# Changelog
# ─────────────────────────────────────────────────────────────────────────

def update_changelog(wb):
    ws = wb["11_Changelog"]
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=f"rev9-v0.3")
    ws.cell(row=next_row, column=2, value="2026-05-06")
    ws.cell(row=next_row, column=3, value="Cowork Opus / CP-SUPIN-04 STEP 12")
    ws.cell(row=next_row, column=4, value=(
        "Bottom-up migration from live-walk evidence on demo.bouracka.cz. "
        "Added 10 new TTs (TT-CP-PHASE0..PHASE5 + IDENTIFIER + CODELISTS), "
        "25 new TCs (TC-CP-NEW-A..Y), env_constraints column on 02_TestCases "
        "(L-WORK-11). Codelists CL-LIVE-* fully populated in fixtures. "
        "Δ matrix v0.2 → v0.3 with 8 confirmed deltas."
    ))
    print(f"[ok] changelog row added at row {next_row}")


# ─────────────────────────────────────────────────────────────────────────

def main():
    print(f"[info] loading {WB_PATH}")
    wb = openpyxl.load_workbook(WB_PATH)

    add_env_constraints_column(wb)
    add_new_tts(wb)
    add_new_tcs(wb)
    update_changelog(wb)
    consistency_check(wb)

    wb.save(WB_PATH)
    print(f"[ok] saved {WB_PATH}")


if __name__ == "__main__":
    main()
_"
    main()
= openpyxl.load_workbook(WB_PATH)

    add_env_constraints_column(wb)
    add_new_tts(wb)
    add_new_tcs(wb)
    update_changelog(wb)
    consistency_check(wb)

    wb.save(WB_PATH)
    print(f"[ok] saved {WB_PATH}")


if __name__ == "__main__":
    main()

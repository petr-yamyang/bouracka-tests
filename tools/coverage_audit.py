#!/usr/bin/env python3
"""
coverage_audit.py — audit TestTarget × TestCase coverage matrix.

CP-SUPIN-05 v0.1 — Phase 0 (informational only). Reads:
  - Sheet 15_VModelAssemblyMap (TT enumeration with layer)
  - Sheet 16_CoverageMatrix    (TC × TT bindings with strength)

Writes report to runs/coverage-audit-{date}.json + console.

Phases planned (from _specs/COVERAGE-RULE-STRATEGY-v0.1-CS.md):
  v0.1 — Phase 0 (informational, exit 0 always)
  v0.2 — Phase 1 (warnings on uncovered TT-FUNC-*)
  v0.3 — Phase 2 (--enforce flag, exit 2 on TT-FUNC gap)
  v0.4 — Phase 3 (full strict, per-layer thresholds)
"""
import argparse
import json
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path


def audit(xlsx_path: Path, out_dir: Path) -> int:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[coverage] openpyxl not installed — run: py -m pip install openpyxl")
        return 3

    if not xlsx_path.exists():
        print(f"[coverage] Excel not found: {xlsx_path}")
        return 4

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    if "15_VModelAssemblyMap" not in wb.sheetnames:
        print("[coverage] sheet 15_VModelAssemblyMap missing")
        print("[coverage] expected after Excel schema bump v0.5.1; running Phase 0 — OK")
        return 0
    if "16_CoverageMatrix" not in wb.sheetnames:
        print("[coverage] sheet 16_CoverageMatrix missing")
        print("[coverage] expected after Excel schema bump v0.5.1; running Phase 0 — OK")
        return 0

    tt_sheet = wb["15_VModelAssemblyMap"]
    cov_sheet = wb["16_CoverageMatrix"]

    # tt_code -> tt_layer
    tt_layer = {}
    headers = [c.value for c in next(tt_sheet.iter_rows(max_row=1))]
    if "tt_code" not in headers or "tt_layer" not in headers:
        print(f"[coverage] sheet 15 missing required cols; got {headers}")
        return 5
    code_idx = headers.index("tt_code")
    layer_idx = headers.index("tt_layer")
    for row in tt_sheet.iter_rows(min_row=2, values_only=True):
        if row and row[code_idx]:
            tt_layer[row[code_idx]] = row[layer_idx]

    # tc_code × tt_code -> coverage_strength
    matrix = defaultdict(list)
    headers = [c.value for c in next(cov_sheet.iter_rows(max_row=1))]
    if "tc_code" not in headers or "tt_code" not in headers or "coverage_strength" not in headers:
        print(f"[coverage] sheet 16 missing required cols; got {headers}")
        return 5
    tc_idx = headers.index("tc_code")
    tt_idx = headers.index("tt_code")
    str_idx = headers.index("coverage_strength")
    for row in cov_sheet.iter_rows(min_row=2, values_only=True):
        if row and row[tc_idx] and row[tt_idx]:
            matrix[row[tt_idx]].append((row[tc_idx], row[str_idx]))

    # Compute coverage per layer
    layer_stats = defaultdict(
        lambda: {"total": 0, "covered": 0, "uncovered": [], "tcs": set()}
    )
    for tt_code, layer in tt_layer.items():
        if not layer:
            continue
        s = layer_stats[layer]
        s["total"] += 1
        bindings = matrix.get(tt_code, [])
        full_or_partial = [b for b in bindings if b[1] in ("full", "partial")]
        if full_or_partial:
            s["covered"] += 1
            for tc, _ in full_or_partial:
                s["tcs"].add(tc)
        else:
            s["uncovered"].append(tt_code)

    # Console report
    print(f"[coverage-audit] {xlsx_path.name} — {date.today().isoformat()}")
    print(f"[coverage-audit] Phase 0 (informational only — gating not yet enabled)")
    print()
    for layer in sorted(layer_stats.keys()):
        s = layer_stats[layer]
        pct = (s["covered"] / s["total"] * 100) if s["total"] else 0
        bar = "█" * int(pct / 5) + "·" * (20 - int(pct / 5))
        print(f"  TT-{layer}-* {bar}  {s['covered']:>3}/{s['total']:<3}  ({pct:>5.1f}%)  {len(s['tcs'])} unique TCs")
        if s["uncovered"]:
            preview = ", ".join(s["uncovered"][:5])
            more = f" +{len(s['uncovered']) - 5} more" if len(s["uncovered"]) > 5 else ""
            print(f"    gaps: {preview}{more}")
    print()

    # JSON report
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"coverage-audit-{date.today().isoformat()}.json"
    serializable = {
        layer: {
            "total": s["total"],
            "covered": s["covered"],
            "uncovered": s["uncovered"],
            "tcs": sorted(s["tcs"]),
        }
        for layer, s in layer_stats.items()
    }
    with open(out_path, "w", encoding="utf-8") as h:
        json.dump(
            {
                "date": date.today().isoformat(),
                "xlsx": str(xlsx_path),
                "phase": 0,
                "by_layer": serializable,
            },
            h,
            indent=2,
            ensure_ascii=False,
        )
    print(f"[coverage-audit] report: {out_path}")
    return 0  # Phase 0 — never fail


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", type=Path, help="Path to BOURACKA-TESTPLAN-vX.Y.Z.xlsx")
    ap.add_argument(
        "--out-dir",
        type=Path,
        default=Path("runs"),
        help="Output dir for JSON report (default: runs/)",
    )
    ap.add_argument(
        "--enforce",
        action="store_true",
        help="Phase 2+: exit non-zero on uncovered FUNC TTs (no-op in v0.1)",
    )
    args = ap.parse_args()
    sys.exit(audit(args.xlsx, args.out_dir))

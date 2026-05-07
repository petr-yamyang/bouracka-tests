#!/usr/bin/env python3
"""
migrate_to_v04_branch_tagging.py — CP-SUPIN-04 STEP 19.

User direction: master workbook with parametric branch view (Bouračka/PROD vs
DEMO Bouračka). Hide rows by toggle. Single source of truth, one test suite,
later switch to pre-live tests by parameter.

What this migration does (v0.3.2 → v0.4.0):

  1. Adds two BOOLEAN columns to every ItemBase sheet:
       applies_to_demo  TRUE/FALSE  — row is in scope for DEMO branch
       applies_to_prod  TRUE/FALSE  — row is in scope for PROD branch

     Derivation rule (from env_constraints column on 02_TestCases) +
     manually-set on 00b_Requirements / 01_TestTargets:

       env_constraints           applies_to_demo  applies_to_prod
       ─────────────────────────────────────────────────────────
       both                      TRUE             TRUE
       both-with-adapter         TRUE             TRUE
       demo-only                 TRUE             FALSE
       prod-only                 FALSE            TRUE
       (blank — TestTargets/Reqs)  per row hint

  2. Pre-configures Excel AutoFilter on every ItemBase sheet so the user
     gets clickable filter dropdowns for both columns. To switch view:
       - Click the AutoFilter dropdown on `applies_to_demo` → uncheck FALSE
         to hide PROD-only rows.
       - Same on `applies_to_prod` for inverse.
       - Both checked = show everything.

  3. Adds a new sheet `00e_BranchView` with:
       - Quick TT/TC counts per branch
       - "Switch view" instructions in CS
       - Cross-link to AutoFilter usage

  4. Updates `00_README!A1` to include the branched-master tagline.

  5. Bumps the version to v0.4.0 (minor bump — schema change).

  6. Adds a rev11 changelog entry.

Output: BOURACKA-TESTPLAN-v0.4.0.xlsx
"""
from __future__ import annotations
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

ROOT = Path(__file__).resolve().parent.parent
SRC  = ROOT / "BOURACKA-TESTPLAN-v0.3.2.xlsx"
DST  = ROOT / "BOURACKA-TESTPLAN-v0.4.0.xlsx"

NEW_VERSION = "v0.4.0"
NEW_REV = "rev11"
NEW_DATE = date(2026, 5, 6)
TODAY = datetime(2026, 5, 6)

ITEMBASE_SHEETS = ['00b_Requirements', '01_TestTargets', '02_TestCases', '08_Bugs']

HDR_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF")
TRUE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FALSE_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

# Manual hints for sheets that don't have env_constraints (00b, 01)
# Default rule: TRUE for both branches unless explicitly overridden.
# Overrides keyed on item_code:
TT_REQ_OVERRIDES = {
    # R2 TTs that are PROD-only (mostly things tied to live integrations)
    'TT-CP-R2-DEMO-RELAX': ('TRUE', 'FALSE'),  # DEMO-RELAX is by definition demo-only
    # Everything else defaults to (TRUE, TRUE)
}


def col_of(ws, name: str):
    for i, c in enumerate(ws[1], 1):
        if c.value == name:
            return i
    return None


def derive_from_env_constraints(env: str | None) -> tuple[str, str]:
    if env is None or str(env).strip() == '':
        return ('TRUE', 'TRUE')
    e = str(env).strip().lower()
    if e == 'both' or e == 'both-with-adapter':
        return ('TRUE', 'TRUE')
    if e == 'demo-only':
        return ('TRUE', 'FALSE')
    if e == 'prod-only':
        return ('FALSE', 'TRUE')
    return ('TRUE', 'TRUE')


def add_branch_columns(wb):
    counts = {}
    for sheet in ITEMBASE_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        # Skip if columns already added
        existing_demo = col_of(ws, 'applies_to_demo')
        existing_prod = col_of(ws, 'applies_to_prod')
        if existing_demo and existing_prod:
            print(f"[skip] {sheet}: branch columns already present")
            continue

        # Add at end
        demo_col = ws.max_column + 1
        prod_col = ws.max_column + 2
        ws.cell(row=1, column=demo_col, value='applies_to_demo').fill = HDR_FILL
        ws.cell(row=1, column=demo_col).font = HDR_FONT
        ws.cell(row=1, column=prod_col, value='applies_to_prod').fill = HDR_FILL
        ws.cell(row=1, column=prod_col).font = HDR_FONT

        env_col = col_of(ws, 'env_constraints')
        code_col = col_of(ws, 'item_code')

        n_demo = 0
        n_prod = 0
        for r in range(2, ws.max_row + 1):
            code = ws.cell(row=r, column=code_col).value if code_col else None
            if env_col:
                env = ws.cell(row=r, column=env_col).value
                d, p = derive_from_env_constraints(env)
            else:
                # 00b_Requirements / 01_TestTargets / 08_Bugs — use overrides
                d, p = TT_REQ_OVERRIDES.get(code, ('TRUE', 'TRUE'))

            cell_d = ws.cell(row=r, column=demo_col, value=d)
            cell_d.fill = TRUE_FILL if d == 'TRUE' else FALSE_FILL
            cell_d.alignment = Alignment(horizontal='center')
            cell_d.font = Font(bold=True)

            cell_p = ws.cell(row=r, column=prod_col, value=p)
            cell_p.fill = TRUE_FILL if p == 'TRUE' else FALSE_FILL
            cell_p.alignment = Alignment(horizontal='center')
            cell_p.font = Font(bold=True)

            if d == 'TRUE': n_demo += 1
            if p == 'TRUE': n_prod += 1

        # Configure AutoFilter on the whole header row
        last_col_letter = get_column_letter(ws.max_column)
        last_row = ws.max_row
        ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

        counts[sheet] = (n_demo, n_prod, last_row - 1)
        print(f"[ok] {sheet}: branch cols {get_column_letter(demo_col)}+{get_column_letter(prod_col)}; "
              f"DEMO={n_demo}/{last_row-1}, PROD={n_prod}/{last_row-1}; AutoFilter set")

    return counts


def add_branch_view_sheet(wb, counts: dict):
    name = '00e_BranchView'
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, index=4)

    ws['A1'] = 'Branched master view — Bouračka (PROD) vs DEMO Bouračka'
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = 'Architecture'
    ws['A3'].font = Font(bold=True)
    ws['A4'] = 'Single master workbook — switch view via AutoFilter on applies_to_demo / applies_to_prod columns.'
    ws['A5'] = 'Both branches share the same TestTargets / TestCases where env_constraints = "both" or "both-with-adapter".'
    ws['A6'] = 'Rows specific to one branch are tagged via env_constraints (demo-only / prod-only).'

    ws['A8'] = 'Counts per branch (entries with applies_to_X = TRUE)'
    ws['A8'].font = Font(bold=True)

    ws['A10'] = 'Sheet'
    ws['B10'] = 'Total rows'
    ws['C10'] = 'DEMO branch'
    ws['D10'] = 'PROD branch'
    ws['E10'] = 'DEMO-only'
    ws['F10'] = 'PROD-only'
    ws['G10'] = 'Both'
    for c in ['A10','B10','C10','D10','E10','F10','G10']:
        ws[c].font = HDR_FONT
        ws[c].fill = HDR_FILL
        ws[c].alignment = Alignment(horizontal='center')

    row = 11
    total_demo = 0
    total_prod = 0
    total_both = 0
    total_rows = 0
    for sheet in ITEMBASE_SHEETS:
        if sheet not in counts: continue
        n_demo, n_prod, n_total = counts[sheet]
        n_demo_only = n_demo - (n_demo + n_prod - n_total)
        n_prod_only = n_prod - (n_demo + n_prod - n_total)
        n_both = (n_demo + n_prod - n_total)
        ws.cell(row=row, column=1, value=sheet)
        ws.cell(row=row, column=2, value=n_total)
        ws.cell(row=row, column=3, value=n_demo)
        ws.cell(row=row, column=4, value=n_prod)
        ws.cell(row=row, column=5, value=n_demo_only)
        ws.cell(row=row, column=6, value=n_prod_only)
        ws.cell(row=row, column=7, value=n_both)
        for col in range(2, 8):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        total_demo += n_demo; total_prod += n_prod; total_rows += n_total; total_both += n_both
        row += 1

    # Totals
    ws.cell(row=row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=row, column=2, value=total_rows).font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_demo).font = Font(bold=True)
    ws.cell(row=row, column=4, value=total_prod).font = Font(bold=True)
    ws.cell(row=row, column=7, value=total_both).font = Font(bold=True)
    for col in range(2, 8):
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')

    # Instructions
    ws['A18'] = 'Jak přepnout pohled (CS)'
    ws['A18'].font = Font(bold=True, size=12)

    instr = [
        ('A19', '1. Otevřete list (např. 02_TestCases).'),
        ('A20', '2. Klikněte na šipku AutoFilter v záhlaví sloupce `applies_to_demo`.'),
        ('A21', '3. Pro pohled "DEMO branch only" — zaškrtněte JEN TRUE.'),
        ('A22', '4. Pro pohled "PROD branch only" — udělejte stejné na sloupci `applies_to_prod`.'),
        ('A23', '5. Pro reset (zobrazit vše) — klikněte na šipku → "Vymazat filtr z..."'),
        ('A25', 'Tip — pro rychlé přepnutí lze definovat Excel makro / VBA proceduru,'),
        ('A26', 'která nastaví filtr na jednom místě a aplikuje napříč všemi listy.'),
        ('A27', 'Vzor makra je v `_install/EXCEL-MACRO-BRANCH-FILTER.bas` (po roll-out).'),
        ('A29', 'Sloupec env_constraints — co znamenají hodnoty (jen 02_TestCases)'),
        ('A29', 'Sloupec env_constraints — co znamenají hodnoty (jen 02_TestCases)'),
    ]
    for coord, text in instr:
        ws[coord] = text

    ws['A29'].font = Font(bold=True, size=12)
    legend = [
        ('A30', '`both`              — TC běží na obou branchech identicky (DEMO i PROD)'),
        ('A31', '`both-with-adapter` — TC běží na obou, na PROD s Mockoon / N8 sandbox adaptérem'),
        ('A32', '`demo-only`         — TC vyžaduje DEMO chování (např. mock OTP); na PROD nemá smysl'),
        ('A33', '`prod-only`         — TC vyžaduje reálné integrace; na DEMO by produkoval false signal'),
    ]
    for coord, text in legend:
        ws[coord] = text

    # Column widths
    ws.column_dimensions['A'].width = 28
    for c in 'BCDEFG':
        ws.column_dimensions[c].width = 14

    print(f"[ok] added sheet '{name}' with branch counts + CS instructions")


def bump_readme(wb):
    ws = wb['00_README']
    ws['A1'].value = f"BOURACKA-TESTPLAN — {NEW_VERSION} (master with branch tagging)"
    if ws['B3'].value:
        ws['B3'].value = NEW_VERSION
    print(f"[ok] 00_README bumped to {NEW_VERSION}")


def add_changelog(wb):
    ws = wb['11_Changelog']
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=f"{NEW_REV}-{NEW_VERSION}")
    ws.cell(row=next_row, column=2, value=NEW_DATE)
    ws.cell(row=next_row, column=3, value="Cowork Opus / CP-SUPIN-04 STEP 19")
    ws.cell(row=next_row, column=4, value=(
        "Branch tagging — added applies_to_demo + applies_to_prod boolean columns "
        "across all ItemBase sheets (00b_Requirements, 01_TestTargets, 02_TestCases, "
        "08_Bugs). Derived from env_constraints (TC sheet) or per-row hint elsewhere. "
        "AutoFilter pre-configured for one-click branch view switching. Added "
        "00e_BranchView sheet with counts + CS instructions. Master workbook is now "
        "the single source of truth for both Bouračka (PROD) and DEMO Bouračka branches."
    ))
    print(f"[ok] changelog row added at {next_row}")


def main():
    print(f"[info] loading {SRC}")
    wb = openpyxl.load_workbook(SRC)

    counts = add_branch_columns(wb)
    add_branch_view_sheet(wb, counts)
    bump_readme(wb)
    add_changelog(wb)

    wb.save(DST)
    print(f"[ok] saved {DST}")
    print(f"[summary] branch-tagged {sum(c[2] for c in counts.values())} ItemBase rows")


if __name__ == '__main__':
    main()

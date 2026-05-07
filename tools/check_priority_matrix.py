#!/usr/bin/env python3
"""
check_priority_matrix.py — standalone consistency check for severity × urgency → priority.

Per CP-SUPIN-04 STEP 17. Run after any edit to ItemBase rows to ensure no
row's priority drifts from the canonical matrix in 01d_PrioritySevUrgMatrix.

Used by:
  - manual review pass before each delivery
  - CI gate (exit non-zero on any violation)

Usage:
  python tools/check_priority_matrix.py [path/to/workbook.xlsx]

Default workbook: BOURACKA-TESTPLAN-v0.3.1.xlsx (or v0.3.x, latest auto-detect).
"""
from __future__ import annotations
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent

# Canonical matrix (also documented in 01d_PrioritySevUrgMatrix sheet)
MATRIX = {
    ('A', 'A'): 'A',  ('A', 'B'): 'A',  ('A', 'C'): 'B',  ('A', 'D'): 'C',
    ('B', 'A'): 'A',  ('B', 'B'): 'B',  ('B', 'C'): 'C',  ('B', 'D'): 'D',
    ('C', 'A'): 'B',  ('C', 'B'): 'C',  ('C', 'C'): 'D',  ('C', 'D'): 'D',
    ('D', 'A'): 'C',  ('D', 'B'): 'D',  ('D', 'C'): 'D',  ('D', 'D'): 'D',
}

ITEMBASE_SHEETS = ['00b_Requirements', '01_TestTargets', '02_TestCases', '08_Bugs']


def col_of(ws, name: str) -> int | None:
    for i, c in enumerate(ws[1], 1):
        if c.value == name:
            return i
    return None


CORRECTED_FORMULA_FRAGMENT = '<=1,"A",IF((CODE('  # signature of corrected formula
BUGGY_FORMULA_FRAGMENT     = '<=3,"A",IF((CODE('  # signature of v0.2/v0.3 buggy formula


def check_workbook(path: Path) -> int:
    print(f"[info] checking {path}")
    # Two-pass: data_only=True reads cached values (post-Excel-save);
    # data_only=False reads formula text (pre-Excel-save / headless).
    wb_values  = openpyxl.load_workbook(path, data_only=True)
    wb_formula = openpyxl.load_workbook(path, data_only=False)
    violations = 0
    rows_checked = 0
    for sheet in ITEMBASE_SHEETS:
        if sheet not in wb_values.sheetnames:
            continue
        ws_v = wb_values[sheet]
        ws_f = wb_formula[sheet]
        sev_col = col_of(ws_v, 'severity')
        urg_col = col_of(ws_v, 'urgency')
        pri_col = col_of(ws_v, 'priority')
        code_col = col_of(ws_v, 'item_code')
        if not (sev_col and urg_col and pri_col):
            continue
        for r in range(2, ws_v.max_row + 1):
            sev = ws_v.cell(row=r, column=sev_col).value
            urg = ws_v.cell(row=r, column=urg_col).value
            cached_pri = ws_v.cell(row=r, column=pri_col).value
            formula_pri = ws_f.cell(row=r, column=pri_col).value
            if not (sev and urg):
                continue
            rows_checked += 1
            expected = MATRIX.get((sev, urg))
            code = ws_v.cell(row=r, column=code_col).value
            # Path A: Excel computed the cache — judge cached value
            if cached_pri is not None and not (isinstance(cached_pri, str) and cached_pri.startswith('=')):
                if cached_pri != expected:
                    violations += 1
                    print(f"  [BAD-cached] {sheet}::{code} — sev={sev} urg={urg} "
                          f"pri={cached_pri} (expected {expected})")
                continue
            # Path B: no cache yet — judge formula text instead
            if isinstance(formula_pri, str):
                if BUGGY_FORMULA_FRAGMENT in formula_pri:
                    violations += 1
                    print(f"  [BAD-formula] {sheet}::{code} — buggy v0.2/v0.3 formula still present "
                          f"(sev={sev} urg={urg}); run tools/fix_priority_matrix.py")
                elif CORRECTED_FORMULA_FRAGMENT not in formula_pri:
                    violations += 1
                    print(f"  [BAD-formula] {sheet}::{code} — formula doesn't match corrected pattern "
                          f"(sev={sev} urg={urg})")
            else:
                # Plain text but not a formula — should still equal expected
                if formula_pri != expected:
                    violations += 1
                    print(f"  [BAD-static] {sheet}::{code} — sev={sev} urg={urg} "
                          f"pri={formula_pri} (expected {expected})")
    print()
    print(f"[summary] {rows_checked} rows checked, {violations} violations")
    if violations == 0:
        print("[ok] all rows consistent with canonical matrix.")
    return violations


def main():
    if len(sys.argv) > 1:
        path = Path(sys.argv[1])
    else:
        # Auto-detect latest v0.3.x workbook
        candidates = sorted(ROOT.glob('BOURACKA-TESTPLAN-v0.3*.xlsx'))
        if not candidates:
            sys.exit("[FAIL] no v0.3.x workbook found in repo root")
        path = candidates[-1]
    if not path.exists():
        sys.exit(f"[FAIL] workbook not found: {path}")
    sys.exit(check_workbook(path))


if __name__ == '__main__':
    main()

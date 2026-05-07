#!/usr/bin/env python3
"""
migrate_08bugs_v04_1.py — CP-SUPIN-04 STEP 23.

Adds dedup-key columns to 08_Bugs sheet per BUG-NAMING-CONVENTION-v0.1.md:
  - tc_ref          (which TC surfaced this bug)
  - assertion_code  (which assertion failed; e.g. S4)
  - assertion_text  (verbatim failing expression)
  - first_seen      (DATE)
  - last_seen       (DATE)
  - occurrences     (INT)
  - envs_seen       (CSV of envs)
  - runs_seen       (CSV of last 10 RUN-* IDs)
  - screenshot_ref  (most recent test-failed-1.png path)
  - trace_ref       (most recent trace.zip path)
  - error_message   (most recent verbatim error)

Output: BOURACKA-TESTPLAN-v0.4.1.xlsx (rev12 — schema bump for bugs)
"""
from __future__ import annotations
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "BOURACKA-TESTPLAN-v0.4.0.xlsx"
DST = ROOT / "BOURACKA-TESTPLAN-v0.4.1.xlsx"

NEW_VERSION = "v0.4.1"
NEW_REV = "rev12"

HDR_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF")

# New columns to add (in order)
NEW_COLUMNS = [
    "tc_ref",
    "assertion_code",
    "assertion_text",
    "first_seen",
    "last_seen",
    "occurrences",
    "envs_seen",
    "runs_seen",
    "screenshot_ref",
    "trace_ref",
    "error_message",
]


def col_of(ws, name):
    for i, c in enumerate(ws[1], 1):
        if c.value == name:
            return i
    return None


def add_columns(wb):
    if "08_Bugs" not in wb.sheetnames:
        print("[skip] 08_Bugs not present")
        return
    ws = wb["08_Bugs"]
    next_col = ws.max_column + 1
    for name in NEW_COLUMNS:
        if col_of(ws, name):
            print(f"[skip] {name} already present")
            continue
        cell = ws.cell(row=1, column=next_col, value=name)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")
        next_col += 1
    print(f"[ok] 08_Bugs: added {len(NEW_COLUMNS)} dedup columns")


def bump_readme(wb):
    ws = wb["00_README"]
    ws["A1"].value = f"BOURACKA-TESTPLAN — {NEW_VERSION} (master with branch tagging + bug dedup)"
    if ws["B3"].value:
        ws["B3"].value = NEW_VERSION
    print(f"[ok] 00_README bumped to {NEW_VERSION}")


def add_changelog(wb):
    ws = wb["11_Changelog"]
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=f"{NEW_REV}-{NEW_VERSION}")
    ws.cell(row=next_row, column=2, value=date(2026, 5, 6))
    ws.cell(row=next_row, column=3, value="Cowork Opus / CP-SUPIN-04 STEP 23")
    ws.cell(row=next_row, column=4, value=(
        "Bug naming convention + dedup. Added 11 columns to 08_Bugs (tc_ref, "
        "assertion_code, assertion_text, first_seen, last_seen, occurrences, "
        "envs_seen, runs_seen, screenshot_ref, trace_ref, error_message). "
        "ID pattern: BUG-CP-{TC_CODE}-{ASSERT_CODE}. "
        "Detail: _specs/BUG-NAMING-CONVENTION-v0.1.md."
    ))
    print(f"[ok] changelog rev12 added")


def main():
    print(f"[info] loading {SRC}")
    wb = openpyxl.load_workbook(SRC)
    add_columns(wb)
    bump_readme(wb)
    add_changelog(wb)
    wb.save(DST)
    print(f"[ok] saved {DST}")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
append_test_run_result.py — UPSERT test-run rows into 07_TestRunResults.

Per CP-SUPIN-04 STEP 28. Called by playwright/reporters/excel-row-writer.ts
at end of every Playwright run with the path to results.json.

Dedup key: (tc_code, run_id, retry). Same key → updates existing row.
Different key → appends new row.

Auto-detects newest BOURACKA-TESTPLAN-*.xlsx in repo root.

Usage:
  py tools/append_test_run_result.py path/to/results.json
"""
from __future__ import annotations
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parent.parent

# Color-code by status for at-a-glance review
STATUS_FILLS = {
    "passed":      PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # green
    "failed":      PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # red
    "timedOut":    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "skipped":     PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),  # blue
    "interrupted": PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),  # amber
}

EXPECTED_HEADERS = [
    "tc_code", "test_title", "run_id", "env", "viewport", "framework",
    "status", "duration_ms", "retry", "started_at", "ended_at",
    "error_message", "screenshot_ref", "trace_ref",
]


def find_workbook() -> Path:
    candidates = sorted(ROOT.glob("BOURACKA-TESTPLAN-*.xlsx"))
    if not candidates:
        raise FileNotFoundError("no BOURACKA-TESTPLAN-*.xlsx in repo root")
    return candidates[-1]


def ensure_headers(ws):
    headers = [c.value for c in ws[1]]
    if not headers or all(h is None for h in headers):
        # empty sheet — write headers
        for i, h in enumerate(EXPECTED_HEADERS, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        return {h: i for i, h in enumerate(EXPECTED_HEADERS, 1)}

    # ensure all expected headers exist; append missing
    col_idx = {}
    next_col = ws.max_column + 1
    for h in EXPECTED_HEADERS:
        if h in headers:
            col_idx[h] = headers.index(h) + 1
        else:
            ws.cell(row=1, column=next_col, value=h).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=1, column=next_col).fill = PatternFill(
                start_color="1F4E78", end_color="1F4E78", fill_type="solid"
            )
            col_idx[h] = next_col
            next_col += 1
    return col_idx


def find_existing_row(ws, col_idx, tc_code, run_id, retry):
    """Look for row matching (tc_code, run_id, retry)."""
    tc_c = col_idx["tc_code"]
    run_c = col_idx["run_id"]
    ret_c = col_idx["retry"]
    for r in range(2, ws.max_row + 1):
        if (
            ws.cell(row=r, column=tc_c).value == tc_code
            and ws.cell(row=r, column=run_c).value == run_id
            and ws.cell(row=r, column=ret_c).value == retry
        ):
            return r
    return None


def upsert_row(ws, col_idx, row: dict, target_row: int):
    """Write all fields of row dict into target_row."""
    for key, value in row.items():
        if key in col_idx:
            cell = ws.cell(row=target_row, column=col_idx[key], value=value)
            if key == "status":
                fill = STATUS_FILLS.get(value)
                if fill:
                    cell.fill = fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")


def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: append_test_run_result.py <results.json>")

    json_path = Path(sys.argv[1])
    if not json_path.exists():
        sys.exit(f"[FAIL] not found: {json_path}")

    rows = json.loads(json_path.read_text(encoding="utf-8"))
    if not rows:
        print("[info] no test rows; nothing to do")
        return 0

    wb_path = find_workbook()
    print(f"[info] target workbook: {wb_path}")
    wb = openpyxl.load_workbook(wb_path)

    if "07_TestRunResults" not in wb.sheetnames:
        ws = wb.create_sheet("07_TestRunResults", index=10)
    else:
        ws = wb["07_TestRunResults"]

    col_idx = ensure_headers(ws)

    n_inserted = 0
    n_updated = 0
    for row in rows:
        existing = find_existing_row(
            ws, col_idx, row.get("tc_code"), row.get("run_id"), row.get("retry", 0)
        )
        if existing:
            upsert_row(ws, col_idx, row, existing)
            n_updated += 1
        else:
            target_row = ws.max_row + 1
            upsert_row(ws, col_idx, row, target_row)
            n_inserted += 1

    wb.save(wb_path)
    print(f"[ok] {n_inserted} inserted, {n_updated} updated; saved {wb_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

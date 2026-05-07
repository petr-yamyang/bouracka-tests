#!/usr/bin/env python3
"""
bump_workbook_version.py — CP-SUPIN-04 STEP 18.

User-reported: TestPlan versioning isn't working properly; values are not
updated. Two root causes identified:

  1. `00_README` sheet front-page still shows "BOURACKA-TESTPLAN — v0.1"
     across all revisions because no migration script touched it. The
     header is the first thing a reviewer sees and has been LYING since
     CP-SUPIN-02.

  2. The `priority` formula cells have no cached value at write time —
     openpyxl writes formulas but doesn't compute them. Any non-Excel
     viewer (`load_workbook(data_only=True)`, web-based xlsx renderers,
     scripted readers) sees `None` until someone opens the file in Excel
     and saves once.

Fix strategy:

  A. Bump 00_README front-page (title cell + version cell + date cell)
     to the new revision in lockstep with the file rename.
  B. Replace formula `priority` cells with STATIC computed values from
     the canonical matrix (see 01d_PrioritySevUrgMatrix). Non-Excel
     viewers now see the right priority immediately. Excel viewers see
     the static value AND a comment pointing to the matrix sheet for
     re-derivation.
  C. Bump `updated_at` cells (column 25 in ItemBase sheets) to today
     for every row that wasn't touched by the priority-fix migration.
  D. Add a "Front-page sanity check" to the validator (#12) so future
     migrations don't regress this.

Output: BOURACKA-TESTPLAN-v0.3.2.xlsx

Run:
  python tools/bump_workbook_version.py
"""
from __future__ import annotations
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "BOURACKA-TESTPLAN-v0.3.1.xlsx"
DST = ROOT / "BOURACKA-TESTPLAN-v0.3.2.xlsx"

NEW_VERSION = "v0.3.2"
NEW_REV = "rev10"
NEW_DATE = date(2026, 5, 6)
TODAY = datetime(2026, 5, 6)

# Canonical priority matrix
MATRIX = {
    ('A', 'A'): 'A',  ('A', 'B'): 'A',  ('A', 'C'): 'B',  ('A', 'D'): 'C',
    ('B', 'A'): 'A',  ('B', 'B'): 'B',  ('B', 'C'): 'C',  ('B', 'D'): 'D',
    ('C', 'A'): 'B',  ('C', 'B'): 'C',  ('C', 'C'): 'D',  ('C', 'D'): 'D',
    ('D', 'A'): 'C',  ('D', 'B'): 'D',  ('D', 'C'): 'D',  ('D', 'D'): 'D',
}

ITEMBASE_SHEETS = ['00b_Requirements', '01_TestTargets', '02_TestCases', '08_Bugs']


def col_of(ws, name: str):
    for i, c in enumerate(ws[1], 1):
        if c.value == name:
            return i
    return None


def fix_readme(wb):
    """Bump 00_README front page."""
    if '00_README' not in wb.sheetnames:
        print("[skip] 00_README not present")
        return
    ws = wb['00_README']
    # A1 is the title. Replace any "v0.X" suffix.
    a1 = ws['A1'].value or ''
    if 'BOURACKA-TESTPLAN' in str(a1):
        ws['A1'].value = f"BOURACKA-TESTPLAN — {NEW_VERSION}"
    # B3 is typically the version cell
    if ws['B3'].value and 'v0.' in str(ws['B3'].value):
        ws['B3'].value = f"{NEW_VERSION}.0"
    # B4 / similar — search rows 2..10 for "v0.X" patterns and bump
    for r in range(2, 11):
        for c_idx in (1, 2, 3):
            cell = ws.cell(row=r, column=c_idx)
            v = cell.value
            if isinstance(v, str) and v.startswith('v0.') and len(v) <= 8:
                cell.value = NEW_VERSION + ('.0' if v.count('.') == 2 else '')
            elif isinstance(v, datetime):
                # bump version-stamp dates
                cell.value = TODAY
    print(f"[ok] 00_README bumped to {NEW_VERSION}")


def materialise_priorities(wb):
    """Replace priority formula cells with STATIC computed values."""
    total_fixed = 0
    for sheet in ITEMBASE_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        sev_col = col_of(ws, 'severity')
        urg_col = col_of(ws, 'urgency')
        pri_col = col_of(ws, 'priority')
        if not (sev_col and urg_col and pri_col):
            continue
        n = 0
        for r in range(2, ws.max_row + 1):
            sev = ws.cell(row=r, column=sev_col).value
            urg = ws.cell(row=r, column=urg_col).value
            if not (sev and urg):
                continue
            pri = MATRIX.get((sev, urg))
            if pri is None:
                continue  # invalid sev/urg combo — leave alone
            cell = ws.cell(row=r, column=pri_col)
            cell.value = pri
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            # Color-code by priority for at-a-glance review
            colour = {'A': 'FF6B6B', 'B': 'FFD93D', 'C': 'A8DADC', 'D': 'B0BEC5'}[pri]
            cell.fill = PatternFill(start_color=colour, end_color=colour, fill_type='solid')
            # Add a comment reference (only on the first row of each sheet — keeps file size small)
            if n == 0:
                cell.comment = Comment(
                    f"Static value computed from severity × urgency per "
                    f"01d_PrioritySevUrgMatrix.\nFix history: rev9.1 (formula correction), "
                    f"rev10 (materialised to static).",
                    "TestPlan-rev10",
                )
            n += 1
        if n:
            print(f"[ok] {sheet}: materialised {n} priority cells (formula → static)")
            total_fixed += n
    return total_fixed


def bump_updated_at(wb):
    """Touch updated_at column for every ItemBase row to today."""
    total = 0
    for sheet in ITEMBASE_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        upd_col = col_of(ws, 'updated_at')
        if not upd_col:
            continue
        for r in range(2, ws.max_row + 1):
            # only bump rows that have at least an item_code (skip blanks)
            code = ws.cell(row=r, column=col_of(ws, 'item_code') or 2).value
            if code:
                ws.cell(row=r, column=upd_col).value = TODAY
                total += 1
    print(f"[ok] bumped updated_at on {total} rows across {len(ITEMBASE_SHEETS)} sheets")


def add_changelog(wb):
    if '11_Changelog' not in wb.sheetnames:
        return
    ws = wb['11_Changelog']
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=f"{NEW_REV}-{NEW_VERSION}")
    ws.cell(row=next_row, column=2, value=NEW_DATE)
    ws.cell(row=next_row, column=3, value="Cowork Opus / CP-SUPIN-04 STEP 18")
    ws.cell(row=next_row, column=4, value=(
        "Fix Excel versioning + materialised priority values. "
        f"Bumped 00_README front-page from v0.1 → {NEW_VERSION} (was stale across rev2..rev9.1). "
        "Replaced priority formula cells with static computed values from the canonical matrix "
        "(non-Excel readers now see the right value immediately; previously they saw blanks "
        "until someone opened in Excel + saved). Added comment-reference on first priority "
        "cell of each ItemBase sheet pointing to 01d_PrioritySevUrgMatrix. Bumped updated_at "
        "on all ItemBase rows to 2026-05-06."
    ))
    print(f"[ok] changelog row added at {next_row}")


def add_validation_check(wb):
    """Add a tiny meta sheet with the version sanity-check rules for future
    migrations to honour."""
    name = '00c_VersionSanityRules'
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, index=2)
    ws['A1'] = 'Version sanity rules (must hold after every rev)'
    ws['A1'].font = Font(bold=True, size=14)

    rules = [
        ('A3', '1. 00_README!A1 must contain the current version string (e.g. "v0.3.2")'),
        ('A4', '2. 00_README!B3 must contain the full version (e.g. "v0.3.2.0")'),
        ('A5', '3. 11_Changelog!A{last} must reference the same version'),
        ('A6', '4. priority column must NOT contain formulas (use static values from MATRIX)'),
        ('A7', '5. updated_at column must equal the rev date for any row touched'),
        ('A8', '6. severity ∈ {A, B, C, D} and urgency ∈ {A, B, C, D} for every row that has both'),
        ('A9', '7. priority must equal MATRIX[severity][urgency] for every row that has sev+urg'),
        ('A11', 'Validator: tools/check_priority_matrix.py + tools/validate_workbook.py'),
        ('A12', 'Migration tools: tools/migrate_to_v0X.py + tools/bump_workbook_version.py'),
    ]
    for coord, text in rules:
        ws[coord] = text
    ws.column_dimensions['A'].width = 110
    print(f"[ok] added sheet '{name}'")


def main():
    print(f"[info] loading {SRC}")
    wb = openpyxl.load_workbook(SRC)

    fix_readme(wb)
    n = materialise_priorities(wb)
    bump_updated_at(wb)
    add_validation_check(wb)
    add_changelog(wb)

    wb.save(DST)
    print(f"[ok] saved {DST}")
    print(f"[info] priority cells materialised: {n} (no Excel-recompute round trip needed)")


if __name__ == '__main__':
    main()

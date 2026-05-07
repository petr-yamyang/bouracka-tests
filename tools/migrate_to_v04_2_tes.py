#!/usr/bin/env python3
"""
migrate_to_v04_2_tes.py — CP-SUPIN-04 STEP 32.

Schema migration v0.4.1 → v0.4.2 — adds TestExecution Summary sheets.

What this migration does:
  1. Adds sheet `13_TestExecutionSummary` — TC-level run records
     (per VUP-grade format spec; one row per TC × run)
  2. Adds sheet `14_AssertionGateResults` — per-assertion detail
     (one row per assertion within step within TC × run)
  3. Bumps 00_README to v0.4.2
  4. Adds rev13 changelog entry
  5. Updates `00c_VersionSanityRules` if present

Output: BOURACKA-TESTPLAN-v0.4.2.xlsx
"""
from __future__ import annotations
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "BOURACKA-TESTPLAN-v0.4.1.xlsx"
DST = ROOT / "BOURACKA-TESTPLAN-v0.4.2.xlsx"

NEW_VERSION = "v0.4.2"
NEW_REV = "rev13"

HDR_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF")

TES_COLUMNS = [
    "run_id", "tc_code", "tc_title",
    "framework", "env", "viewport",
    "started_at", "ended_at", "duration_ms",
    "verdict",
    "step_count", "assertion_count",
    "assertion_pass_count", "assertion_fail_count",
    "failed_at_step", "failed_at_assertion",
    "error_message", "retry",
    "screenshot_path", "trace_path",
    "bug_ref", "tester", "created_at",
]

AGR_COLUMNS = [
    "run_id", "tc_code", "step_no", "step_kind", "step_description",
    "assertion_code", "assertion_kind",
    "assertion_expected", "assertion_actual",
    "verdict", "duration_ms", "evidence_ref", "notes",
]


def add_sheet_with_headers(wb, name, headers, after=None):
    if name in wb.sheetnames:
        print(f"[skip] {name} already exists")
        return wb[name]
    if after:
        ws = wb.create_sheet(name, index=wb.sheetnames.index(after) + 1)
    else:
        ws = wb.create_sheet(name)
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")
    print(f"[ok] sheet '{name}' added with {len(headers)} columns")
    return ws


def bump_readme(wb):
    if "00_README" not in wb.sheetnames:
        return
    ws = wb["00_README"]
    ws["A1"].value = f"BOURACKA-TESTPLAN — {NEW_VERSION} (master + branch tagging + bug dedup + TES)"
    if ws["B3"].value:
        ws["B3"].value = NEW_VERSION
    print(f"[ok] 00_README bumped to {NEW_VERSION}")


def add_changelog(wb):
    ws = wb["11_Changelog"]
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=f"{NEW_REV}-{NEW_VERSION}")
    ws.cell(row=next_row, column=2, value=date(2026, 5, 6))
    ws.cell(row=next_row, column=3, value="Cowork Opus / CP-SUPIN-04 STEP 32")
    ws.cell(row=next_row, column=4, value=(
        "Multi-platform testing + TestExecution Summary VUP-grade artefact. "
        f"Added 13_TestExecutionSummary sheet ({len(TES_COLUMNS)} cols, TC-level run records) "
        f"and 14_AssertionGateResults sheet ({len(AGR_COLUMNS)} cols, per-assertion detail). "
        "Per spec _specs/TEST-EXECUTION-SUMMARY-FORMAT-v0.1-CS.md. "
        "Strategy detail _specs/MULTI-PLATFORM-TESTING-STRATEGY-v0.1-CS.md."
    ))
    print(f"[ok] changelog rev13 added")


def main():
    print(f"[info] loading {SRC}")
    wb = openpyxl.load_workbook(SRC)
    add_sheet_with_headers(wb, "13_TestExecutionSummary", TES_COLUMNS, after="07_TestRunResults")
    add_sheet_with_headers(wb, "14_AssertionGateResults", AGR_COLUMNS, after="13_TestExecutionSummary")
    bump_readme(wb)
    add_changelog(wb)
    wb.save(DST)
    print(f"[ok] saved {DST}")


if __name__ == "__main__":
    main()

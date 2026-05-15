"""Workbook reader for bouracka-ui — Excel-backed Repository implementation.

v0.1 Phase 1 — Runnable Mock: returns synthetic data based on the schema
of BOURACKA-TESTPLAN-v0.4.2.xlsx but doesn't actually read the file. Phase 2
implements real reads via openpyxl.

REPO-HOOK module — this module's public functions are the swap-in boundary
where the future OracleRepository (per `bouracka_ui/dwh/`) replaces the
Excel-backed reads with Oracle-backed reads. Callers (server.py) MUST go
through this module's functions; direct openpyxl imports in server.py are
forbidden by discipline (per BOURACKA-DATA-STORE-EVOLUTION-PLAN §M2).

Public surface (REPO-HOOKs):
  list_envs(wb_path)
  list_tcs(wb_path, env, framework)
  list_bugs(wb_path, status, severity)
  get_bug(wb_path, code)
  append_bug(wb_path, bug)
  update_bug(wb_path, code, fields)
  set_bug_retest_run_id(wb_path, code, run_id)

Reference:
  _config/BOURACKA-UI-DESIGN-v0.1-2026-05-10.md §4
  _config/BOURACKA-DWH-DEV-PLAN-v0.1-EN.md §6 (REPO-HOOK boundaries)
"""
from __future__ import annotations

import datetime as _dt
from pathlib import Path

# ──────────────────────────────────────────────────────────────────────────
# ENV code → schema env enum mapping (per design doc §4.1)
# ──────────────────────────────────────────────────────────────────────────
# 2026-05-12 Kate-drop alignment: empirical inspection of KP-reviewed workbook
# v0.4.3 showed ENV-DMO base_url = "https://tst.demo.bouracka.cz/" (NOT public
# demo.bouracka.cz). So ENV-DMO in the workbook IS the SUPIN-internal driving-
# school DEMO. To avoid the silent-wrong-target trap (workbook says ENV-DMO is
# tst.demo but dispatcher resolves schema "demo" to public demo.bouracka.cz),
# ENV-DMO now maps to "tst-demo" schema, matching the workbook's actual URL.
# Public demo.bouracka.cz is surfaced via the supplemental ENV-DMO-PUB.
ENV_CODE_TO_SCHEMA = {
    "ENV-PUB":     "prod-readonly",
    "ENV-TST":     "tst",
    "ENV-DMO":     "tst-demo",      # workbook reality: tst.demo.bouracka.cz
    "ENV-TST-DMO": "tst-demo",      # alias for explicit naming if workbook adds it
    "ENV-DMO-PUB": "demo",          # supplemental — public demo.bouracka.cz
    # Future: ENV-UAT → uat (when env-tagging migration v0.5.2 lands)
    # Future: ENV-PRD-W → prod-writable
}

# Synthetic envs the UI surfaces even when not present in the workbook —
# guarantees Kate has BOTH the SUPIN-internal AND public DEMO targets
# selectable without requiring a v0.4.x schema bump. Added 2026-05-12 Kate drop.
SUPPLEMENTAL_ENVS = [
    {
        "code": "ENV-DMO-PUB",
        "name_cs": "DEMO (veřejné, demo.bouracka.cz)",
        "name_en": "DEMO (public, demo.bouracka.cz)",
        "descr_en": "Public DEMO at demo.bouracka.cz. Cíl-1 baseline target. "
                    "Reachable without SUPIN network. Reduced validation; watermarked PDF.",
        "base_url": "https://demo.bouracka.cz",
        "recaptcha_posture": "live",
        "schema_env": "demo",
        "status": "live",
    },
]


class WorkbookLockedError(RuntimeError):
    """Raised when the workbook is open in Excel and we can't write."""


# ──────────────────────────────────────────────────────────────────────────
# Header-name → 0-based column index helper.
#
# BUG-K-001 (2026-05-13 Kate first-round): hardcoded column indexes in
# list_envs / list_tcs were drifting after KP's 2026-05-11 review added 3
# columns (comments_KP_en, env, ext_ws) to 02_TestCases. Result: the
# framework-filter dropdown returned 0 TCs because r[21] was reading the
# wrong column. Fix: read row-1 headers, map by lowercase name, look up.
# Same pattern as _BUG_COL for the bugs sheet.
# ──────────────────────────────────────────────────────────────────────────

def _column_map(ws) -> dict[str, int]:
    """Map row-1 header text (lowercase, stripped) → 0-based column index.
    Robust against column reordering / insertion in the workbook."""
    out: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col_idx).value
        if v is not None:
            out[str(v).strip().lower()] = col_idx - 1
    return out


def _safe_get(row: tuple, col_idx: int, default=None):
    """Bounds-checked tuple access."""
    if col_idx is None or col_idx < 0 or col_idx >= len(row):
        return default
    return row[col_idx]


# ──────────────────────────────────────────────────────────────────────────
# §4.1 read paths
# ──────────────────────────────────────────────────────────────────────────

def list_envs(wb_path: Path) -> list[dict]:
    """Read 04_TestEnvironments → list of dicts.

    Phase 1 mock: returns synthetic data matching the workbook's actual shape.
    Phase 2 implements real openpyxl read.
    """
    if not wb_path.exists():
        return _mock_envs()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(wb_path), read_only=True, data_only=True)
        if "04_TestEnvironments" not in wb.sheetnames:
            return _mock_envs()
        ws = wb["04_TestEnvironments"]
        # Header-based column lookup — robust to v0.4.3 KP-review column drift.
        # BUG-K-001 fix (2026-05-13).
        cols = _column_map(ws)
        c_code           = cols.get("item_code", 1)
        c_name_cs        = cols.get("name_cs", 2)
        c_name_en        = cols.get("name_en", 3)
        c_descr_en       = cols.get("descr_en", 5)
        c_status         = cols.get("item_status", 7)
        c_base_url       = cols.get("base_url", 13)
        c_recaptcha      = cols.get("recaptcha_posture", 17)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        out = []
        for r in rows:
            if not r or not _safe_get(r, c_code):
                continue
            code = _safe_get(r, c_code)
            out.append({
                "code": code,
                "name_cs": _safe_get(r, c_name_cs),
                "name_en": _safe_get(r, c_name_en),
                "descr_en": _safe_get(r, c_descr_en),
                "base_url": _safe_get(r, c_base_url),
                "recaptcha_posture": _safe_get(r, c_recaptcha),
                "schema_env": ENV_CODE_TO_SCHEMA.get(code, "demo"),
                "status": _safe_get(r, c_status),
            })
        # Merge in supplemental envs the workbook doesn't carry yet —
        # primarily ENV-TST-DMO for the Kate drop tst.demo.bouracka.cz access.
        existing_codes = {e["code"] for e in out}
        for supp in SUPPLEMENTAL_ENVS:
            if supp["code"] not in existing_codes:
                out.append(supp)
        return out
    except Exception as e:
        # Workbook unreadable — fall back to mock so UI still loads
        return _mock_envs()


def list_tcs(wb_path: Path, env: str | None = None,
             framework: str | None = None) -> list[dict]:
    """Read 02_TestCases. Filter by env (applies_to_demo / applies_to_prod)
    and framework (substring in framework_targets) if provided.

    Phase 1 mock unless workbook readable.
    """
    if not wb_path.exists():
        return _mock_tcs(env, framework)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(wb_path), read_only=True, data_only=True)
        if "02_TestCases" not in wb.sheetnames:
            return _mock_tcs(env, framework)
        ws = wb["02_TestCases"]
        # Header-based column lookup — robust to KP-review column shifts.
        # BUG-K-001 fix (2026-05-13 Kate first-round).
        cols = _column_map(ws)
        c_code              = cols.get("item_code", 1)
        c_name_cs           = cols.get("name_cs", 2)
        c_name_en           = cols.get("name_en", 3)
        c_status            = cols.get("item_status", 7)
        c_severity          = cols.get("severity", 8)
        c_urgency           = cols.get("urgency", 9)
        c_priority          = cols.get("priority", 10)
        c_type              = cols.get("type", 14)
        c_framework_targets = cols.get("framework_targets", 21)
        c_applies_demo      = cols.get("applies_to_demo", 34)
        c_applies_prod      = cols.get("applies_to_prod", 35)
        # KP-review (v0.4.3) added these — surface to UI so it can filter
        c_env_shorthand     = cols.get("env", None)             # 'STANDARD' / 'DEMO, STANDARD' / etc.
        c_ext_ws            = cols.get("ext_ws", None)          # 'N8' / 'AISPOV' / 'D8WS' / etc.
        c_comments_kp_en    = cols.get("comments_kp_en", None)  # KP acceptance criteria
        c_steps_count       = cols.get("steps_count", None)     # v0.4.4: denormalized count
        c_steps_summary     = cols.get("steps_summary", None)   # legacy fallback

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        out = []
        for r in rows:
            if not r or not _safe_get(r, c_code):
                continue
            code = _safe_get(r, c_code)
            framework_targets_raw = _safe_get(r, c_framework_targets, "") or ""
            applies_to_demo_val = _safe_get(r, c_applies_demo)
            applies_to_prod_val = _safe_get(r, c_applies_prod)
            applies_to_demo = bool(applies_to_demo_val) if applies_to_demo_val is not None else None
            applies_to_prod = bool(applies_to_prod_val) if applies_to_prod_val is not None else None
            # Env filter
            if env == "demo" and applies_to_demo is False:
                continue
            if env == "prod-readonly" and applies_to_prod is False:
                continue
            # Framework filter — defensive parse:
            #   - empty cell → include row (assume applies to all frameworks)
            #   - comma-separated → split + trim + lowercase + set-membership
            # BUG-K-001 fix: prior version did substring match on r[21] which
            # was wrong column after KP review; this version is header-based +
            # tolerant of empty cells.
            if framework:
                ft_str = str(framework_targets_raw).strip().lower()
                if ft_str:  # only filter if column is populated
                    targets = {t.strip() for t in ft_str.split(",") if t.strip()}
                    if framework.lower() not in targets:
                        continue
                # else: empty framework_targets → include (assume any framework)
            # steps_count: primary from column, fallback from newline-count in steps_summary
            raw_steps_count = _safe_get(r, c_steps_count) if c_steps_count is not None else None
            if raw_steps_count is None and c_steps_summary is not None:
                summary_raw = _safe_get(r, c_steps_summary, "") or ""
                raw_steps_count = len([ln for ln in str(summary_raw).split("\n") if ln.strip()])
            out.append({
                "code": code,
                "name_cs": _safe_get(r, c_name_cs),
                "name_en": _safe_get(r, c_name_en),
                "severity": _safe_get(r, c_severity),
                "urgency": _safe_get(r, c_urgency),
                "priority": _safe_get(r, c_priority),
                "status": _safe_get(r, c_status),
                "type": _safe_get(r, c_type),
                "framework_targets": framework_targets_raw,
                "applies_to_demo": applies_to_demo,
                "applies_to_prod": applies_to_prod,
                # KP-review enrichment (v0.4.3) — exposed to UI for filter+display
                "env_shorthand": _safe_get(r, c_env_shorthand),
                "ext_ws": _safe_get(r, c_ext_ws),
                "comments_kp_en": _safe_get(r, c_comments_kp_en),
                "steps_count": raw_steps_count,
            })
        return out
    except Exception:
        return _mock_tcs(env, framework)


def list_bugs(wb_path: Path, status: str | None = None,
              severity: str | None = None) -> list[dict]:
    """Read 08_Bugs → JIRA-style list of dicts."""
    if not wb_path.exists():
        return _mock_bugs(status, severity)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(wb_path), read_only=True, data_only=True)
        if "08_Bugs" not in wb.sheetnames:
            return _mock_bugs(status, severity)
        ws = wb["08_Bugs"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        # Header positions (per investigation 2026-05-10):
        # 0:id 1:code 2:name_cs 3:name_en 7:status 8:severity 9:urgency 10:priority
        # 13:env_where_present 14:linked_tc_ref 15:linked_run_result_ref
        # 36:screenshot_ref 37:trace_ref 38:error_message
        out = []
        for r in rows:
            if not r or not r[1]:
                continue
            row_status = r[7]
            row_severity = r[8]
            if status and row_status != status:
                continue
            if severity and row_severity != severity:
                continue
            out.append({
                "code": r[1],
                "name_cs": r[2],
                "name_en": r[3],
                "status": row_status,
                "severity": row_severity,
                "urgency": r[9],
                "priority": r[10],
                "env_where_present": r[13] if len(r) > 13 else None,
                "linked_tc_ref": r[14] if len(r) > 14 else None,
                "linked_run_result_ref": r[15] if len(r) > 15 else None,
                "screenshot_ref": r[36] if len(r) > 36 else None,
                "error_message": r[38] if len(r) > 38 else None,
            })
        return out
    except Exception:
        return _mock_bugs(status, severity)


# Bug column index map (08_Bugs sheet, 0-based). Per the append_bug layout below
# and the read mapping in list_bugs(). Single source of truth — both readers
# and writers consult this. v0.1.3 (Block 2 bug-edit work).
_BUG_COL = {
    "id":                      0,
    "code":                    1,
    "name_cs":                 2,
    "name_en":                 3,
    "descr_cs":                4,
    "descr_en":                5,
    "item_type":               6,
    "status":                  7,
    "severity":                8,
    "urgency":                 9,
    "priority":               10,
    "submitter":              11,
    "submitted_at":           12,
    "env_where_present":      13,
    "linked_tc_ref":          14,
    "linked_run_result_ref":  15,
    "repro_steps":            16,
    "expected":               17,
    "actual":                 18,
    "created_at":             21,
    "updated_at":             22,
    "last_retest_run_id":     23,   # new col 2026-05-13 (Block 2 retest)
    "screenshot_ref":         36,
    "trace_ref":              37,
    "error_message":          38,
}


def get_bug(wb_path: Path, code: str) -> dict | None:
    """Read a single bug row from 08_Bugs by code. None if not found.
    v0.1.3 (Block 2 — bug-edit workflow)."""
    if not wb_path.exists():
        # mock fallback
        mocks = _mock_bugs(None, None)
        for b in mocks:
            if b.get("code") == code:
                return {**b, "repro_steps": "(mock)", "expected": "(mock)",
                        "actual": "(mock)", "descr_en": "(mock)",
                        "created_at": None, "updated_at": None,
                        "last_retest_run_id": None}
        return None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(wb_path), read_only=True, data_only=True)
        if "08_Bugs" not in wb.sheetnames:
            return None
        ws = wb["08_Bugs"]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r and r[_BUG_COL["code"]] == code:
                return _row_to_bug_dict(r)
    except Exception:
        return None
    return None


def _row_to_bug_dict(r: tuple) -> dict:
    """Convert a workbook row tuple to a bug dict, full fields for edit form."""
    def at(idx):
        return r[idx] if len(r) > idx else None
    return {
        "code":                  at(_BUG_COL["code"]),
        "name_cs":               at(_BUG_COL["name_cs"]),
        "name_en":               at(_BUG_COL["name_en"]),
        "descr_cs":              at(_BUG_COL["descr_cs"]),
        "descr_en":              at(_BUG_COL["descr_en"]),
        "status":                at(_BUG_COL["status"]),
        "severity":              at(_BUG_COL["severity"]),
        "urgency":               at(_BUG_COL["urgency"]),
        "priority":              at(_BUG_COL["priority"]),
        "submitter":             at(_BUG_COL["submitter"]),
        "submitted_at":          at(_BUG_COL["submitted_at"]),
        "env_where_present":     at(_BUG_COL["env_where_present"]),
        "linked_tc_ref":         at(_BUG_COL["linked_tc_ref"]),
        "linked_run_result_ref": at(_BUG_COL["linked_run_result_ref"]),
        "repro_steps":           at(_BUG_COL["repro_steps"]),
        "expected":              at(_BUG_COL["expected"]),
        "actual":                at(_BUG_COL["actual"]),
        "created_at":            at(_BUG_COL["created_at"]),
        "updated_at":            at(_BUG_COL["updated_at"]),
        "last_retest_run_id":    at(_BUG_COL["last_retest_run_id"]),
        "screenshot_ref":        at(_BUG_COL["screenshot_ref"]),
        "trace_ref":             at(_BUG_COL["trace_ref"]),
        "error_message":         at(_BUG_COL["error_message"]),
    }


# Fields that the PUT endpoint is allowed to modify. id + code + created_at
# are immutable; updated_at + last_retest_run_id are server-managed.
_BUG_UPDATABLE = {
    "name_cs", "name_en", "descr_cs", "descr_en",
    "status", "severity", "urgency", "priority",
    "env_where_present", "linked_tc_ref", "linked_run_result_ref",
    "repro_steps", "expected", "actual",
    "screenshot_ref", "trace_ref", "error_message",
}


def update_bug(wb_path: Path, code: str, fields: dict) -> dict | None:
    """Update an existing bug by code. Returns the updated bug dict, or None
    if the bug doesn't exist. Ignores keys not in _BUG_UPDATABLE.
    v0.1.3 (Block 2 — bug-edit workflow).

    Raises WorkbookLockedError if Excel has the workbook open.
    Raises FileNotFoundError if the workbook is missing.
    """
    if not wb_path.exists():
        raise FileNotFoundError(f"workbook not at {wb_path}")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(wb_path))  # NOT read_only
        if "08_Bugs" not in wb.sheetnames:
            raise RuntimeError("08_Bugs sheet missing")
        ws = wb["08_Bugs"]
        target_row_idx = None
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=_BUG_COL["code"] + 1).value == code:
                target_row_idx = row_idx
                break
        if target_row_idx is None:
            return None
        # Apply only allowlist fields
        applied = 0
        for k, v in fields.items():
            if k not in _BUG_UPDATABLE:
                continue
            col_idx = _BUG_COL.get(k)
            if col_idx is None:
                continue
            ws.cell(row=target_row_idx, column=col_idx + 1, value=v)
            applied += 1
        # Always touch updated_at
        now = _dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=target_row_idx, column=_BUG_COL["updated_at"] + 1, value=now)
        wb.save(str(wb_path))
        # Return refreshed row
        return get_bug(wb_path, code)
    except PermissionError as e:
        raise WorkbookLockedError(str(e)) from e


def set_bug_retest_run_id(wb_path: Path, code: str, run_id: str) -> None:
    """Record that a retest run was launched for this bug.
    Best-effort — failure is logged but not raised (retest dispatch shouldn't
    fail because bookkeeping failed).
    v0.1.3 (Block 2 — retest workflow)."""
    if not wb_path.exists():
        return
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(wb_path))
        if "08_Bugs" not in wb.sheetnames:
            return
        ws = wb["08_Bugs"]
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=_BUG_COL["code"] + 1).value == code:
                ws.cell(row=row_idx, column=_BUG_COL["last_retest_run_id"] + 1, value=run_id)
                now = _dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
                ws.cell(row=row_idx, column=_BUG_COL["updated_at"] + 1, value=now)
                wb.save(str(wb_path))
                return
    except Exception:
        # Bookkeeping failure is non-fatal for retest workflow
        return


def append_bug(wb_path: Path, bug: dict) -> str:
    """Append one row to 08_Bugs. Returns assigned bug code (BUG-NNN).

    Phase 1 mock: returns a fake code without touching the workbook.
    Phase 2 implements real append via openpyxl.
    """
    if not wb_path.exists():
        # Mock — returns fake code
        return "BUG-MOCK-001"
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(wb_path))  # NOT read_only — we need to write
        if "08_Bugs" not in wb.sheetnames:
            raise RuntimeError("08_Bugs sheet missing")
        ws = wb["08_Bugs"]
        # Compute next bug code
        codes = [c.value for c in ws["B"][1:] if c.value and str(c.value).startswith("BUG-")]
        nums = []
        for c in codes:
            try:
                nums.append(int(str(c).split("-")[-1]))
            except (ValueError, IndexError):
                continue
        next_n = (max(nums) if nums else 0) + 1
        new_code = f"BUG-{next_n:03d}"
        next_id = (max([c.value for c in ws["A"][1:] if isinstance(c.value, int)] or [0]) + 1)
        now = _dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        new_row = [None] * 39  # match header count
        new_row[0] = next_id
        new_row[1] = new_code
        new_row[3] = bug.get("name_en", "")
        new_row[5] = bug.get("descr_en", "")
        new_row[6] = "bug"
        new_row[7] = "open"
        new_row[8] = bug.get("severity", "C")
        new_row[9] = bug.get("urgency", "C")
        new_row[10] = bug.get("priority", "P3-medium")
        new_row[11] = bug.get("submitter", "bouracka-ui")
        new_row[12] = now
        new_row[13] = bug.get("env_where_present")
        new_row[14] = bug.get("linked_tc_ref")
        new_row[15] = bug.get("linked_run_result_ref")
        new_row[16] = bug.get("repro_steps")
        new_row[17] = bug.get("expected")
        new_row[18] = bug.get("actual")
        new_row[21] = now  # created_at
        new_row[22] = now  # updated_at
        ws.append(new_row)
        wb.save(str(wb_path))
        return new_code
    except PermissionError as e:
        raise WorkbookLockedError(str(e)) from e


# ──────────────────────────────────────────────────────────────────────────
# Mocks (used in Phase 1 + when workbook missing/unreadable)
# ──────────────────────────────────────────────────────────────────────────

def _mock_envs() -> list[dict]:
    return [
        {"code": "ENV-PUB", "name_cs": "Public produkce",
         "name_en": "Public production",
         "descr_en": "Public production; dev-time recon only, never test execution.",
         "base_url": "https://www.bouracka.cz",
         "recaptcha_posture": "live",
         "schema_env": "prod-readonly", "status": "live"},
        {"code": "ENV-TST", "name_cs": "Testovací prostředí",
         "name_en": "TEST environment",
         "descr_en": "Primary test env; full validation; both drivers sign via SMS-OTP.",
         "base_url": "https://tst.bouracka.cz",
         "recaptcha_posture": "test-keys",
         "schema_env": "tst", "status": "live"},
        {"code": "ENV-DMO", "name_cs": "DEMO (autoškola, SUPIN-internal)",
         "name_en": "DEMO (driving school, SUPIN-internal)",
         "descr_en": "SUPIN-internal driving-school DEMO at tst.demo.bouracka.cz. "
                     "Reduced validation per scope C-5. Requires SUPIN network reach. "
                     "Matches workbook v0.4.3 ENV-DMO row.",
         "base_url": "https://tst.demo.bouracka.cz",
         "recaptcha_posture": "BYPASS_TOKEN_OR_DISABLED",
         "schema_env": "tst-demo", "status": "live"},
        {"code": "ENV-DMO-PUB", "name_cs": "DEMO (veřejné, demo.bouracka.cz)",
         "name_en": "DEMO (public, demo.bouracka.cz)",
         "descr_en": "Public DEMO at demo.bouracka.cz. Cíl-1 baseline target. "
                     "Reachable without SUPIN network.",
         "base_url": "https://demo.bouracka.cz",
         "recaptcha_posture": "live",
         "schema_env": "demo", "status": "live"},
    ]


def _mock_tcs(env: str | None, framework: str | None) -> list[dict]:
    base = [
        {"code": "TC-CP-A1-MAIN-DEMO",
         "name_en": "A1 — Main happy day end-to-end (DEMO)",
         "severity": "A", "urgency": "A", "priority": "P1-critical",
         "status": "live", "type": "E2E",
         "framework_targets": "playwright,cypress,selenium",
         "applies_to_demo": True, "applies_to_prod": False,
         "env_shorthand": "DEMO", "ext_ws": None, "comments_kp_en": None,
         "steps_count": 3},
        {"code": "TC-CP-A2-ALT-1",
         "name_en": "A2 ALT-1 — RP regex validation",
         "severity": "B", "urgency": "B", "priority": "P2-high",
         "status": "live", "type": "FUNC",
         "framework_targets": "playwright,cypress,selenium",
         "applies_to_demo": True, "applies_to_prod": True,
         "env_shorthand": "STANDARD", "ext_ws": None, "comments_kp_en": None,
         "steps_count": 2},
        {"code": "TC-CP-A2-ALT-4",
         "name_en": "A2 ALT-4 — GDPR consent capture",
         "severity": "B", "urgency": "A", "priority": "P2-high",
         "status": "live", "type": "FUNC",
         "framework_targets": "playwright,cypress,selenium",
         "applies_to_demo": True, "applies_to_prod": True,
         "env_shorthand": "STANDARD", "ext_ws": None, "comments_kp_en": None,
         "steps_count": 2},
        {"code": "TC-CP-A2-ALT-5",
         "name_en": "A2 ALT-5 — Slovak prefix RP recognition",
         "severity": "C", "urgency": "B", "priority": "P3-medium",
         "status": "live", "type": "FUNC",
         "framework_targets": "playwright,cypress,selenium",
         "applies_to_demo": True, "applies_to_prod": True,
         "env_shorthand": "STANDARD", "ext_ws": None, "comments_kp_en": None,
         "steps_count": 2},
        {"code": "TC-CP-A2-ALT-6",
         "name_en": "A2 ALT-6 — Police card optional path",
         "severity": "B", "urgency": "B", "priority": "P2-high",
         "status": "live", "type": "FUNC",
         "framework_targets": "playwright,cypress,selenium",
         "applies_to_demo": True, "applies_to_prod": True,
         "env_shorthand": "STANDARD", "ext_ws": None, "comments_kp_en": None,
         "steps_count": 2},
        {"code": "TC-CP-A2-ALT-7",
         "name_en": "A2 ALT-7 — Enumerations completeness",
         "severity": "B", "urgency": "A", "priority": "P2-high",
         "status": "live", "type": "FUNC",
         "framework_targets": "playwright,cypress,selenium",
         "applies_to_demo": True, "applies_to_prod": True,
         "env_shorthand": "STANDARD", "ext_ws": None, "comments_kp_en": None,
         "steps_count": 2},
        {"code": "TC-CP-A2-ALT-8",
         "name_en": "A2 ALT-8 — DEMO banner presence",
         "severity": "B", "urgency": "B", "priority": "P2-high",
         "status": "live", "type": "FUNC",
         "framework_targets": "playwright,cypress,selenium",
         "applies_to_demo": True, "applies_to_prod": False,
         "env_shorthand": "DEMO", "ext_ws": None, "comments_kp_en": None,
         "steps_count": 1},
        {"code": "TC-CP-A2-ALT-9",
         "name_en": "A2 ALT-9 — Post-reports drift detection (soft-pass)",
         "severity": "A", "urgency": "A", "priority": "P1-critical",
         "status": "live", "type": "FUNC",
         "framework_targets": "playwright,cypress,selenium",
         "applies_to_demo": True, "applies_to_prod": True,
         "env_shorthand": "STANDARD", "ext_ws": None, "comments_kp_en": None,
         "steps_count": 3},
        {"code": "TC-CP-A2-ALT-10",
         "name_en": "A2 ALT-10 — SPA post-probe",
         "severity": "B", "urgency": "A", "priority": "P2-high",
         "status": "live", "type": "FUNC",
         "framework_targets": "playwright,cypress,selenium",
         "applies_to_demo": True, "applies_to_prod": True,
         "env_shorthand": "STANDARD", "ext_ws": None, "comments_kp_en": None,
         "steps_count": 2},
    ]
    if env == "demo":
        base = [t for t in base if t["applies_to_demo"]]
    if env == "prod-readonly":
        base = [t for t in base if t["applies_to_prod"]]
    if framework:
        base = [t for t in base if framework in t["framework_targets"].lower()]
    return base


def _mock_bugs(status: str | None, severity: str | None) -> list[dict]:
    base = [
        {"code": "BUG-027",
         "name_en": "mim-alliance-card partner badge stretches full-width",
         "status": "open", "severity": "B", "urgency": "B", "priority": "P3-medium",
         "env_where_present": "ENV-DMO", "linked_tc_ref": None,
         "linked_run_result_ref": None, "screenshot_ref": None,
         "error_message": None},
        {"code": "BUG-028",
         "name_en": "mim-alliance-card__title wraps to 2 lines on RCMT",
         "status": "open", "severity": "B", "urgency": "B", "priority": "P3-medium",
         "env_where_present": "ENV-DMO", "linked_tc_ref": None,
         "linked_run_result_ref": None, "screenshot_ref": None,
         "error_message": None},
        {"code": "BUG-MOCK-001",
         "name_en": "Mock bug — replace via Phase 2 real workbook read",
         "status": "open", "severity": "C", "urgency": "C", "priority": "P3-medium",
         "env_where_present": "ENV-TST", "linked_tc_ref": "TC-CP-A2-ALT-1",
         "linked_run_result_ref": None, "screenshot_ref": None,
         "error_message": "(synthetic — Phase 1 mock)"},
    ]
    if status:
        base = [b for b in base if b["status"] == status]
    if severity:
        base = [b for b in base if b["severity"] == severity]
    return base


# ──────────────────────────────────────────────────────────────────────────
# v0.1.5-dev5: list_steps / get_step / get_bug_evidence readers
# ──────────────────────────────────────────────────────────────────────────

# Synthetic step actions for Phase 1 mock — keyed by TC code.
_MOCK_STEP_ACTIONS: dict[str, list[str]] = {
    "TC-CP-A1-MAIN-DEMO": [
        "Navigate to Bouračka start page",
        "Fill in accident form data for both drivers",
        "Submit report and verify PDF generation",
    ],
    "TC-CP-A2-ALT-1": [
        "Navigate to RP (rodné číslo) input field",
        "Enter invalid format; verify validation error is displayed",
    ],
    "TC-CP-A2-ALT-4": [
        "Open Bouračka form; verify GDPR consent checkbox is present",
        "Check box; submit; verify consent captured in POST payload",
    ],
    "TC-CP-A2-ALT-5": [
        "Enter Slovak RČ prefix (e.g. 9999999/9999); verify field accepts it",
        "Verify no false-positive validation error for SK prefix",
    ],
    "TC-CP-A2-ALT-6": [
        "Navigate to officer data section; verify police-card field optional",
        "Submit without police card; verify form accepts submission",
    ],
    "TC-CP-A2-ALT-7": [
        "Open vehicle-type dropdown; enumerate all options",
        "Verify every expected option (per TC spec annex) is present",
    ],
    "TC-CP-A2-ALT-8": [
        "Load DEMO environment URL; verify DEMO watermark banner is visible",
    ],
    "TC-CP-A2-ALT-9": [
        "Load post-accident reports page after run",
        "Compare report hash against previous run baseline",
        "Verify drift-detection status (soft-pass if hash differs within tolerance)",
    ],
    "TC-CP-A2-ALT-10": [
        "POST accident data; capture SPA transition to results page",
        "Verify results URL contains new run_id path segment",
    ],
}

_STEP_KEYS = (
    "id", "step_code", "tc_ref", "ordinal", "action_cs", "action_en",
    "expected_cs", "expected_en", "framework_hint", "assertion_lib_ref",
    "is_decision_point", "comments_kp_en", "created_at", "updated_at", "notes",
)


def _mock_steps(tc_code: str | None) -> list[dict]:
    out: list[dict] = []
    for code, actions in _MOCK_STEP_ACTIONS.items():
        if tc_code and code != tc_code:
            continue
        for i, action in enumerate(actions, 1):
            out.append({
                "id": None, "step_code": f"STP-{code}-{i:02d}",
                "tc_ref": code, "ordinal": i, "action_cs": action, "action_en": None,
                "expected_cs": None, "expected_en": None, "framework_hint": None,
                "assertion_lib_ref": None, "is_decision_point": None,
                "comments_kp_en": None, "created_at": None, "updated_at": None,
                "notes": None,
            })
    out.sort(key=lambda s: (s["tc_ref"], s["ordinal"]))
    return out


def _legacy_steps_from_summary(wb, tc_code: str | None) -> list[dict]:
    """Synthesize steps from 02_TestCases.steps_summary when 02e_TestSteps missing."""
    if "02_TestCases" not in wb.sheetnames:
        return []
    ws = wb["02_TestCases"]
    cols = _column_map(ws)
    c_code = cols.get("item_code", 1)
    c_summary = cols.get("steps_summary")
    if c_summary is None:
        return []
    out: list[dict] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue
        code = _safe_get(r, c_code)
        if not code:
            continue
        if tc_code and code != tc_code:
            continue
        summary = _safe_get(r, c_summary, "") or ""
        lines = [ln.strip() for ln in str(summary).split("\n") if ln.strip()]
        for i, line in enumerate(lines, 1):
            out.append({
                "id": None, "step_code": f"STP-{code}-{i:02d}",
                "tc_ref": code, "ordinal": i, "action_cs": line, "action_en": None,
                "expected_cs": None, "expected_en": None, "framework_hint": None,
                "assertion_lib_ref": None, "is_decision_point": None,
                "comments_kp_en": None, "created_at": None, "updated_at": None,
                "notes": "(inferred from legacy steps_summary — workbook is pre-v0.4.4)",
            })
    out.sort(key=lambda s: (s["tc_ref"] or "", s["ordinal"] or 0))
    return out


def list_steps(wb_path: Path, tc_code: str | None = None) -> list[dict]:
    """Read 02e_TestSteps. Fallback to legacy steps_summary if sheet missing."""
    if not wb_path.exists():
        return _mock_steps(tc_code)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(wb_path), read_only=True, data_only=True)
        if "02e_TestSteps" not in wb.sheetnames:
            return _legacy_steps_from_summary(wb, tc_code)
        ws = wb["02e_TestSteps"]
        cols = _column_map(ws)
        c_id        = cols.get("id", 0)
        c_step_code = cols.get("step_code", 1)
        c_tc_ref    = cols.get("tc_ref", 2)
        c_ordinal   = cols.get("ordinal", 3)
        c_action_cs = cols.get("action_cs", 4)
        c_action_en = cols.get("action_en", 5)
        c_exp_cs    = cols.get("expected_cs", 6)
        c_exp_en    = cols.get("expected_en", 7)
        c_fw_hint   = cols.get("framework_hint", 8)
        c_assert    = cols.get("assertion_lib_ref", 9)
        c_decision  = cols.get("is_decision_point", 10)
        c_kp        = cols.get("comments_kp_en", 11)
        c_created   = cols.get("created_at", 12)
        c_updated   = cols.get("updated_at", 13)
        out: list[dict] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or not _safe_get(r, c_step_code):
                continue
            tc_ref_val = _safe_get(r, c_tc_ref)
            if tc_code and tc_ref_val != tc_code:
                continue
            out.append({
                "id": _safe_get(r, c_id), "step_code": _safe_get(r, c_step_code),
                "tc_ref": tc_ref_val, "ordinal": _safe_get(r, c_ordinal),
                "action_cs": _safe_get(r, c_action_cs), "action_en": _safe_get(r, c_action_en),
                "expected_cs": _safe_get(r, c_exp_cs), "expected_en": _safe_get(r, c_exp_en),
                "framework_hint": _safe_get(r, c_fw_hint), "assertion_lib_ref": _safe_get(r, c_assert),
                "is_decision_point": _safe_get(r, c_decision), "comments_kp_en": _safe_get(r, c_kp),
                "created_at": _safe_get(r, c_created), "updated_at": _safe_get(r, c_updated),
                "notes": None,
            })
        out.sort(key=lambda s: (s["tc_ref"] or "", s["ordinal"] or 0))
        return out
    except Exception:
        return _mock_steps(tc_code)


def get_step(wb_path: Path, step_code: str) -> dict | None:
    """Return single step by step_code, or None if not found."""
    for s in list_steps(wb_path):
        if s["step_code"] == step_code:
            return s
    return None


def _build_evidence_dict(
    bug_code: str, linked_tc_ref, linked_step_ref,
    screenshot_path, video_path, trace_path,
    capture_kind: str, capture_at, repo_root: Path | None,
) -> dict:
    def _on_disk(path) -> bool:
        return bool(path and repo_root and (repo_root / str(path)).exists())

    def _to_url(path) -> str | None:
        return f"/api/runs/{path}" if path else None

    return {
        "bug_code": bug_code, "linked_tc_ref": linked_tc_ref,
        "linked_step_ref": linked_step_ref,
        "evidence_screenshot_path": screenshot_path,
        "evidence_video_path": video_path,
        "evidence_trace_path": trace_path,
        "evidence_capture_kind": capture_kind,
        "evidence_capture_at": capture_at,
        "screenshot_url": _to_url(screenshot_path),
        "video_url": _to_url(video_path),
        "trace_url": _to_url(trace_path),
        "screenshot_on_disk": _on_disk(screenshot_path),
        "video_on_disk": _on_disk(video_path),
        "trace_on_disk": _on_disk(trace_path),
    }


def get_bug_evidence(wb_path: Path, bug_code: str,
                     repo_root: Path | None = None) -> dict | None:
    """Return visual-evidence record for a bug, or None if no evidence.
    Caller should check bug existence separately (get_bug) before calling this."""
    if not wb_path.exists():
        return None  # mock bugs have no evidence
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(wb_path), read_only=True, data_only=True)
        if "08_Bugs" not in wb.sheetnames:
            return None
        ws = wb["08_Bugs"]
        cols = _column_map(ws)
        c_code      = cols.get("item_code") if "item_code" in cols else cols.get("code", _BUG_COL["code"])
        c_link_step = cols.get("linked_step_ref")
        c_ev_ss     = cols.get("evidence_screenshot_path")
        c_ev_vid    = cols.get("evidence_video_path")
        c_ev_trace  = cols.get("evidence_trace_path")
        c_ev_kind   = cols.get("evidence_capture_kind")
        c_ev_at     = cols.get("evidence_capture_at")
        c_ss_legacy = cols.get("screenshot_ref", _BUG_COL.get("screenshot_ref"))
        c_tr_legacy = cols.get("trace_ref", _BUG_COL.get("trace_ref"))
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or _safe_get(r, c_code) != bug_code:
                continue
            if c_ev_ss is not None:
                ev_ss   = _safe_get(r, c_ev_ss)
                ev_vid  = _safe_get(r, c_ev_vid)
                ev_tr   = _safe_get(r, c_ev_trace)
                if not ev_ss and not ev_vid and not ev_tr:
                    return None
                return _build_evidence_dict(
                    bug_code, linked_tc_ref=_safe_get(r, cols.get("linked_tc_ref", _BUG_COL["linked_tc_ref"])),
                    linked_step_ref=_safe_get(r, c_link_step) if c_link_step else None,
                    screenshot_path=ev_ss, video_path=ev_vid, trace_path=ev_tr,
                    capture_kind=_safe_get(r, c_ev_kind) or "auto-from-fail",
                    capture_at=str(_safe_get(r, c_ev_at)) if _safe_get(r, c_ev_at) else None,
                    repo_root=repo_root,
                )
            # Legacy columns fallback
            leg_ss = _safe_get(r, c_ss_legacy) if c_ss_legacy is not None else None
            leg_tr = _safe_get(r, c_tr_legacy) if c_tr_legacy is not None else None
            if not leg_ss and not leg_tr:
                return None
            return _build_evidence_dict(
                bug_code, linked_tc_ref=_safe_get(r, _BUG_COL["linked_tc_ref"]),
                linked_step_ref=None,
                screenshot_path=leg_ss, video_path=None, trace_path=leg_tr,
                capture_kind="manual-tester", capture_at=None, repo_root=repo_root,
            )
        return None
    except Exception:
        return None

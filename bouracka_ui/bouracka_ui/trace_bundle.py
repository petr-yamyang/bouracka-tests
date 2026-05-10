"""Trace-bundle export + import for bouracka-ui.

Designed for the HP Elite air-gap scenario per
`_config/BOURACKA-UI-DESIGN-v0.1-2026-05-10.md` operator direction
2026-05-10 PM ("comprehensive log and trace package based on simple
file transfer because all limitations we are facing in a project").

Workflow:
  1. Tester runs tests on HP Elite via bouracka-ui
  2. Operator clicks "Export bundle" on /results/{rid} → downloads
     trace-bundle-<rid>.zip
  3. Bundle is shipped (USB / shared folder / email) to Pete's machine
  4. Pete uploads via /api/bundles/import → registers the run locally
     for inspection in his own bouracka-ui
  5. Pete drills into the envelope + per-framework outputs without
     needing access to HP Elite

Bundle layout:
    trace-bundle-<run_id>.zip:
    ├── manifest.json
    ├── README.md
    ├── envelope.json                     ← v0.1 schema envelope (canonical)
    ├── digest.md                         ← human-readable cross-framework digest
    ├── per-framework/
    │   ├── cypress/results.json          (+ optional screenshots/, videos/)
    │   ├── playwright/results.json       (+ optional traces/)
    │   └── selenium/results.json         (+ optional html-report/)
    ├── server-log.txt                    ← bouracka-ui captured stdout for this run
    ├── system/
    │   ├── health.json                   ← /api/health snapshot at run-time
    │   └── tool-versions.txt
    ├── workbook-snapshot/                (if include_workbook)
    │   ├── 02_TestCases.csv
    │   ├── 04_TestEnvironments.csv
    │   └── 08_Bugs.csv
    └── repro.sh                          ← one-line script to re-run this exact run
"""
from __future__ import annotations

import csv
import datetime as _dt
import io
import json
import platform
import shutil
import socket
import subprocess
import sys
import zipfile
from pathlib import Path
from typing import IO

BUNDLE_FORMAT_VERSION = "1.0"


# ──────────────────────────────────────────────────────────────────────────
# Export
# ──────────────────────────────────────────────────────────────────────────

def build_bundle(run_id: str,
                 envelope: dict,
                 envelope_path: Path,
                 repo_root: Path,
                 server_log_lines: list[str],
                 include_evidence: bool = True,
                 include_workbook: bool = True,
                 workbook_path: Path | None = None) -> bytes:
    """Build a trace bundle as an in-memory ZIP and return raw bytes.

    Caller is responsible for streaming bytes to client (FastAPI Response).
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as zf:

        # 1. Envelope (the canonical run record)
        zf.writestr("envelope.json", json.dumps(envelope, indent=2, ensure_ascii=False))

        # 2. Markdown digest if a sibling .md exists
        md_path = envelope_path.with_suffix(".md")
        if md_path.exists():
            zf.write(md_path, "digest.md")

        # 3. Per-framework reporter outputs + (optionally) evidence
        manifest_per_fw: dict = {}
        for fw in envelope.get("frameworks", []):
            files = _collect_framework_files(fw, repo_root, include_evidence)
            for src, arc_path in files:
                if src.exists():
                    zf.write(src, arc_path)
            manifest_per_fw[fw] = [arc for _, arc in files]

        # 4. Server log
        zf.writestr("server-log.txt", "\n".join(server_log_lines) + "\n")

        # 5. System info
        zf.writestr("system/health.json", json.dumps(_capture_system(), indent=2))
        zf.writestr("system/tool-versions.txt", _capture_tool_versions())

        # 6. Workbook snapshot (CSVs of relevant sheets)
        if include_workbook and workbook_path and workbook_path.exists():
            for sheet in ("02_TestCases", "04_TestEnvironments", "08_Bugs"):
                csv_text = _sheet_to_csv(workbook_path, sheet)
                if csv_text:
                    zf.writestr(f"workbook-snapshot/{sheet}.csv", csv_text)

        # 7. Repro script
        zf.writestr("repro.sh", _build_repro_script(envelope, run_id))

        # 8. Manifest + README — written LAST so they reflect the actual contents
        manifest = {
            "bundle_format_version": BUNDLE_FORMAT_VERSION,
            "run_id": run_id,
            "env": envelope.get("env"),
            "schema_version": envelope.get("schema_version"),
            "exported_at": _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "exported_by_host": socket.gethostname(),
            "include_evidence": include_evidence,
            "include_workbook": include_workbook,
            "per_framework_files": manifest_per_fw,
            "summary": envelope.get("summary"),
        }
        zf.writestr("manifest.json", json.dumps(manifest, indent=2))
        zf.writestr("README.md", _build_readme(manifest, envelope))

    return buf.getvalue()


def _collect_framework_files(fw: str, repo_root: Path,
                              include_evidence: bool) -> list[tuple[Path, str]]:
    """Map per-framework artefact paths to ZIP arcname paths."""
    out = []
    if fw == "cypress":
        for p in (repo_root / "cypress" / "cypress-results").glob("*.json"):
            out.append((p, f"per-framework/cypress/{p.name}"))
        if include_evidence:
            for sub in ("screenshots", "videos"):
                d = repo_root / "cypress" / sub
                if d.exists():
                    for p in d.rglob("*"):
                        if p.is_file():
                            rel = p.relative_to(d)
                            out.append((p, f"per-framework/cypress/{sub}/{rel}"))
    elif fw == "playwright":
        pw_results = repo_root / "playwright-report" / "results.json"
        if pw_results.exists():
            out.append((pw_results, "per-framework/playwright/results.json"))
        if include_evidence:
            d = repo_root / "playwright-report"
            for p in d.rglob("*"):
                if p.is_file() and p.suffix in (".zip", ".png", ".webm", ".html"):
                    rel = p.relative_to(d)
                    out.append((p, f"per-framework/playwright/{rel}"))
    elif fw == "selenium":
        se_results = repo_root / "selenium-report" / "results.json"
        if se_results.exists():
            out.append((se_results, "per-framework/selenium/results.json"))
        if include_evidence:
            for ext in (".html", ".png"):
                for p in (repo_root / "selenium-report").rglob(f"*{ext}"):
                    rel = p.relative_to(repo_root / "selenium-report")
                    out.append((p, f"per-framework/selenium/{rel}"))
    return out


def _capture_system() -> dict:
    return {
        "captured_at": _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "host": socket.gethostname(),
        "os": f"{platform.system()}-{platform.release()}",
        "platform": platform.platform(),
        "python": platform.python_version(),
        "fqdn": socket.getfqdn(),
    }


def _capture_tool_versions() -> str:
    """Best-effort tool-version capture as plain text."""
    lines = []
    for label, cmd in [
        ("python", [sys.executable, "--version"]),
        ("node",   ["node", "--version"]),
        ("npx",    ["npx", "--version"]),
        ("npm",    ["npm", "--version"]),
        ("cypress", ["npx", "cypress", "--version"]),
        ("playwright", ["npx", "playwright", "--version"]),
        ("pytest", [sys.executable, "-m", "pytest", "--version"]),
        ("git",    ["git", "--version"]),
    ]:
        try:
            r = subprocess.run(cmd, capture_output=True, text=True, timeout=5)
            ver = (r.stdout or r.stderr).strip().splitlines()[0] if (r.stdout or r.stderr) else ""
            lines.append(f"{label}: {ver}")
        except Exception as e:
            lines.append(f"{label}: (not available — {type(e).__name__})")
    return "\n".join(lines) + "\n"


def _sheet_to_csv(workbook_path: Path, sheet_name: str) -> str:
    """Return one workbook sheet as CSV text. Returns empty string on failure."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(workbook_path), read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return ""
        ws = wb[sheet_name]
        buf = io.StringIO()
        w = csv.writer(buf)
        for row in ws.iter_rows(values_only=True):
            w.writerow(["" if v is None else str(v) for v in row])
        return buf.getvalue()
    except Exception:
        return ""


def _build_repro_script(envelope: dict, run_id: str) -> str:
    env = envelope.get("env", "demo")
    fws = ",".join(envelope.get("frameworks", []))
    tcs = ",".join(r["tc_code"] for r in envelope.get("results", []))
    return f"""#!/usr/bin/env bash
# repro.sh — re-run this exact configuration via bouracka-ui
# Generated for run_id: {run_id}
# Env: {env}
# Frameworks: {fws}
# TCs: {tcs}
set -e

# Option A — via bouracka-ui REST API (if bouracka-ui is running locally)
curl -X POST http://localhost:8424/api/runs \\
     -H "Content-Type: application/json" \\
     -d '{{"env":"{env}","frameworks":["{fws}"],"tcs":["{tcs}".split(",")]}}'

# Option B — direct subprocess (no UI dependency)
# cd ~/Documents/VibeCodeProjects/SUPIN/bouracka-tests
# Cypress:    npx cypress run --env baseUrl=<env-url>
# Playwright: npx playwright test --grep "{tcs.replace(',', '|')}"
# Selenium:   python -m pytest selenium/tests/ -k "{' or '.join(t.replace('-','_') for t in tcs.split(','))}"
# Consolidate: python tools/consolidate_results.py --env {env} --run-id {run_id}
"""


def _build_readme(manifest: dict, envelope: dict) -> str:
    s = envelope.get("summary", {}) or {}
    return f"""# Trace bundle — {manifest['run_id']}

This is a self-describing trace bundle from the bouracka-ui test suite.
Designed for file-transfer between an air-gapped HP Elite (where tests
ran) and an inspection workstation (where Pete debugs).

## Quick reference

- **Run ID:** `{manifest['run_id']}`
- **Env:** `{manifest['env']}`
- **Schema version:** `{manifest['schema_version']}` (cross-framework-result-schema)
- **Exported at:** {manifest['exported_at']} from `{manifest['exported_by_host']}`
- **Bundle format:** v{BUNDLE_FORMAT_VERSION}

## Summary

| Metric | Count |
|--------|------:|
| Total TCs | {s.get('total_tcs', '?')} |
| Passed | {s.get('passed', '?')} |
| Failed | {s.get('failed', '?')} |
| Skipped | {s.get('skipped', '?')} |
| Drift-skipped | {s.get('drift_skip_count', '?')} |
| Soft-passed | {s.get('soft_passed', '?')} |
| Parity divergences | {s.get('parity_divergence_count', '?')} |
| Pass-rate (strict) | {s.get('pass_rate_strict', '?')} |
| Pass-rate (drift-aware) | {s.get('pass_rate_drift_aware', '?')} |

## How to inspect this bundle

### Option 1 — via bouracka-ui (recommended)

1. Have bouracka-ui running locally
2. Open `/runs` page → click "Import bundle" → select this ZIP
3. Bundle is registered locally; click through `/results/<run_id>` to
   drill in just like a local run

### Option 2 — manual inspection

Open `envelope.json` directly to see the canonical run record per
[CROSS-FRAMEWORK-RESULT-SCHEMA-v0.1.md](`_specs/CROSS-FRAMEWORK-RESULT-SCHEMA-v0.1.md`).
Open `digest.md` for a human-readable cross-framework summary.
Per-framework reporter JSONs live under `per-framework/<fw>/`.

### Option 3 — re-run

`repro.sh` contains the exact subprocess invocations to reproduce this
run on a machine with the same env access.

## File map

```
{json.dumps(manifest['per_framework_files'], indent=2, sort_keys=True)}
```

(Plus: envelope.json, digest.md, server-log.txt, system/, workbook-snapshot/, repro.sh)

---
Generated by bouracka-ui trace_bundle module. See
`_config/BOURACKA-UI-DESIGN-v0.1-2026-05-10.md` for design rationale.
"""


# ──────────────────────────────────────────────────────────────────────────
# Import
# ──────────────────────────────────────────────────────────────────────────

class BundleImportError(RuntimeError):
    """Raised when a bundle ZIP is malformed or contains incompatible data."""


def import_bundle(zip_bytes: bytes, runs_dir: Path) -> dict:
    """Extract envelope from uploaded bundle and persist it locally.

    Returns the envelope dict (so caller can answer with run_id + summary).
    """
    runs_dir.mkdir(parents=True, exist_ok=True)
    bundles_dir = runs_dir.parent / "imported-bundles"
    bundles_dir.mkdir(parents=True, exist_ok=True)

    try:
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    except zipfile.BadZipFile as e:
        raise BundleImportError(f"not a valid ZIP: {e}") from e

    # Validate manifest
    try:
        manifest = json.loads(zf.read("manifest.json"))
    except KeyError:
        raise BundleImportError("manifest.json missing — not a bouracka-ui bundle")
    except json.JSONDecodeError as e:
        raise BundleImportError(f"manifest.json malformed: {e}")

    if manifest.get("bundle_format_version") != BUNDLE_FORMAT_VERSION:
        raise BundleImportError(
            f"bundle format version {manifest.get('bundle_format_version')!r} "
            f"!= supported {BUNDLE_FORMAT_VERSION!r}"
        )

    # Extract envelope
    try:
        envelope = json.loads(zf.read("envelope.json"))
    except KeyError:
        raise BundleImportError("envelope.json missing")

    run_id = envelope.get("run_id")
    env = envelope.get("env")
    if not run_id or not env:
        raise BundleImportError("envelope.json missing run_id or env")

    # Persist envelope into runs/ so it shows up in /runs listing
    today = envelope.get("started_at", "").split("T")[0]
    if not today:
        today = _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%d")
    target_json = runs_dir / f"cross-framework-{env}-{today}.json"
    # Avoid overwrite if same path already has different run_id
    if target_json.exists():
        try:
            existing = json.loads(target_json.read_text(encoding="utf-8"))
            if existing.get("run_id") != run_id:
                # Append run-id suffix to disambiguate
                target_json = runs_dir / f"cross-framework-{env}-{today}-{run_id[-7:]}.json"
        except Exception:
            pass
    target_json.write_text(json.dumps(envelope, indent=2, ensure_ascii=False) + "\n",
                           encoding="utf-8")

    # Optionally extract digest.md alongside
    try:
        digest = zf.read("digest.md").decode("utf-8")
        target_json.with_suffix(".md").write_text(digest, encoding="utf-8")
    except KeyError:
        pass

    # Persist the full bundle for later forensics
    bundle_target = bundles_dir / f"{run_id}.zip"
    bundle_target.write_bytes(zip_bytes)

    return {
        "run_id": run_id,
        "env": env,
        "envelope_path": str(target_json),
        "bundle_path": str(bundle_target),
        "summary": envelope.get("summary", {}),
        "imported_at": _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
    }


# ──────────────────────────────────────────────────────────────────────────
# Diagnostics — no-run state snapshot for "the UI itself misbehaving"
# ──────────────────────────────────────────────────────────────────────────

def build_diagnostics_snapshot(repo_root: Path,
                                workbook_path: Path | None,
                                health: dict,
                                recent_log_lines: list[str] | None = None) -> bytes:
    """Build a no-run diagnostics snapshot ZIP.

    Includes: /api/health output, system info, tool versions, workbook sanity
    check, recent server log lines, runs/ index, environment vars (filtered).
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        ts = _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        meta = {
            "diagnostics_format_version": "1.0",
            "captured_at": ts,
            "captured_by_host": socket.gethostname(),
            "kind": "bouracka-ui-diagnostics-snapshot",
        }
        zf.writestr("manifest.json", json.dumps(meta, indent=2))
        zf.writestr("health.json", json.dumps(health, indent=2))
        zf.writestr("system/system.json", json.dumps(_capture_system(), indent=2))
        zf.writestr("system/tool-versions.txt", _capture_tool_versions())

        # Filtered env vars (PATH + bouracka-ui-related)
        env_text_lines = []
        for k, v in sorted(__import__("os").environ.items()):
            if k.startswith("BOURACKA_UI_") or k in ("PATH", "PYTHONPATH", "USER", "USERNAME", "HOME"):
                env_text_lines.append(f"{k}={v}")
        zf.writestr("system/env-vars-filtered.txt", "\n".join(env_text_lines) + "\n")

        # Workbook sheet listing
        if workbook_path and workbook_path.exists():
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(workbook_path), read_only=True, data_only=True)
                zf.writestr("workbook-info.txt",
                            f"path: {workbook_path}\nsheets:\n  " +
                            "\n  ".join(wb.sheetnames) + "\n")
            except Exception as e:
                zf.writestr("workbook-info.txt",
                            f"path: {workbook_path}\nERROR: {e}\n")
        else:
            zf.writestr("workbook-info.txt",
                        f"path: {workbook_path}\nMISSING\n")

        # Runs index
        runs_listing = []
        runs_dir = repo_root / "runs"
        if runs_dir.exists():
            for p in sorted(runs_dir.glob("cross-framework-*.json")):
                runs_listing.append(f"{p.name}\t{p.stat().st_size}")
        zf.writestr("runs-index.txt", "\n".join(runs_listing) + "\n")

        # Recent log
        if recent_log_lines:
            zf.writestr("recent-server-log.txt", "\n".join(recent_log_lines) + "\n")

        zf.writestr("README.md", f"""# Bouračka UI diagnostics snapshot

Captured: {ts}
Host: {socket.gethostname()}

This is a no-run diagnostic dump intended for debugging "the UI itself
misbehaving" without involving any specific test run. Useful for the
HP Elite air-gap scenario when a tester reports the UI is unresponsive
or returning unexpected data.

Contents:
- `manifest.json` — metadata
- `health.json` — /api/health snapshot
- `system/system.json` — host info
- `system/tool-versions.txt` — npx/node/python/pytest/git/etc. versions
- `system/env-vars-filtered.txt` — PATH + BOURACKA_UI_* + USER/HOME
- `workbook-info.txt` — workbook path + sheet listing
- `runs-index.txt` — what's in runs/
- `recent-server-log.txt` — recent server stdout (if captured)
""")

    return buf.getvalue()
